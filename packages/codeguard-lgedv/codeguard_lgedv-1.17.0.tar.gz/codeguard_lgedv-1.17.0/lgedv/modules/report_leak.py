import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

def extract_violations(md_text):
    # Updated regex to handle both formats
    leak_pattern = r'(###?\s*🚨\s*\*?\*?(MEMORY LEAK #[^:]+: [^\n*]+)\*?\*?)(.*?)(?=^###?\s*🚨\s*|^---|\Z)'
    matches = re.findall(leak_pattern, md_text, flags=re.DOTALL | re.MULTILINE)
    results = []
    
    for full_match, title, block in matches:
        # Skip placeholder entries
        clean_title = re.sub(r'\*+', '', title).strip()
        if '[number]' in clean_title or '[Brief Description]' in clean_title:
            continue
            
        # Extract basic info...
        files = ''
        func = ''
        m = re.search(r'\*\*Files Involved:\*\*\s*([^\n]+)', block)
        if m:
            files = m.group(1).strip()
        m = re.search(r'\*\*Function Name:\*\*\s*([^\n]+)', block)
        if m:
            func = m.group(1).strip()
        location = f"Files Involved: {files}<br>Function Name: {func}" if files or func else ''
        
        # Problem Description
        m = re.search(r'\*\*Problem Description:\*\*\s*(.*?)(?=\*\*Current Code|\*\*Fix Recommendation|$)', block, re.DOTALL)
        problem = m.group(1).strip() if m else ''
        
        # Current Code
        m = re.search(r'\*\*Current Code[^\*]*\*\*\s*```cpp\n?(.*?)```', block, re.DOTALL)
        current_code = m.group(1).strip() if m else ''
        
        # Fixed Code - multiple strategies
        fixed_code = ''
        
        # Strategy 1: Look for code blocks after Fix Recommendation
        patterns = [
            r'\*\*Fix Recommendation:\*\*.*?```cpp\n?(.*?)```',
            r'\*\*Fixed Code:\*\*\s*```cpp\n?(.*?)```',
            r'\*\*Solution:\*\*\s*```cpp\n?(.*?)```'
        ]
        
        for pattern in patterns:
            m = re.search(pattern, block, re.DOTALL)
            if m:
                fixed_code = m.group(1).strip()
                break
        
        # Strategy 2: If no code block, get text guidelines after Fix Recommendation (FIXED)
        if not fixed_code:
            m = re.search(r'\*\*Fix Recommendation:\*\*\s*(.*?)(?=^---|\Z)', block, re.DOTALL | re.MULTILINE)
            if m:
                fix_text = m.group(1).strip()
                # Clean up text and keep as plain text (NO COMMENTS)
                fixed_code = fix_text.replace('- ', '').strip()
        
        # Strategy 3: Look for refactored code sections in the entire markdown
        if not fixed_code:
            # Search for refactored code related to this file
            file_match = re.search(r'`([^`]+\.cpp)`', files)
            if file_match:
                filename = file_match.group(1)
                refactor_pattern = rf'## Refactored.*?`{filename}`.*?```cpp\n?(.*?)```'
                m = re.search(refactor_pattern, md_text, re.DOTALL)
                if m:
                    fixed_code = m.group(1).strip()
        
        results.append({
            'Issue': clean_title,
            'Location': location,
            'Problem': problem,
            'Current Code': current_code,
            'Fixed Code': fixed_code
        })
    
    return results

def render_code_with_lines(code, lang="cpp"):
    lines = code.splitlines() if code else []
    if not lines:
        return '<pre><code class="language-cpp">No code available</code></pre>'
    
    # Check if this is actual code or just text guidelines
    if code.startswith('//') or not any(c in code for c in ['{', '}', ';', '(', ')']):
        # This is text/guidelines, render as plain text
        return f'<div style="padding:8px;background:#f8f9fa;border-left:3px solid #28a745;"><p>{code}</p></div>'
    else:
        # This is actual code, render with line numbers
        line_numbers = ''.join(f'<span class="line-number" style="display:block;">{i+1}</span>' for i in range(len(lines)))
        code_html = '\n'.join(lines)
        return f'''
        <div style="display:flex;">
          <div style="background:#f6f8fa;color:#888;padding:6px 8px 6px 6px;font-family:monospace;font-size:13px;line-height:1.5;text-align:right;user-select:none;border-right:1px solid #e1e4e8;">
            {line_numbers}
          </div>
          <pre style="margin:0;flex:1;"><code class="language-{lang}">{code_html}</code></pre>
        </div>
        '''

def is_placeholder(v):
    return (
        v["Issue"].strip().startswith("[#]:") or
        "[number]" in v["Issue"] or
        "[Brief Description]" in v["Issue"] or
        v.get("Rule", "").strip().startswith("[EXACT_RULE_NUMBER]") or
        v["Location"].strip().startswith("[function name") or
        v.get("Severity", "").strip().startswith("[Critical") or
        v["Current Code"].strip() == "[problematic code]" or
        v["Fixed Code"].strip() == "[corrected code]" or
        v.get("Explanation", "").strip() == "[Why this violates the rule and how fix works]"
    )

def write_html_table(violations, output_path):
    # Filter out placeholder entries
    filtered_violations = [v for v in violations if not is_placeholder(v)]
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Memory Leak Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''
        <style>
        table{border-collapse:collapse;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        </style>
        ''')
        f.write('</head><body>\n')
        f.write(f'<h1>Memory Leak Report</h1>\n')
        f.write(f'<p>Total memory leaks found: {len(filtered_violations)}</p>\n')
        f.write('<table>\n')
        f.write('<tr><th>Issue</th><th>Location</th><th>Problem description</th><th>Current Code</th><th>Fixed code</th></tr>\n')
        
        for v in filtered_violations:
            f.write('<tr>')
            f.write(f'<td class="issue-col">{v["Issue"]}</td>')
            f.write(f'<td class="location-col">{v["Location"]}</td>')
            f.write(f'<td class="problem-col">{v["Problem"]}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Current Code"])}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Fixed Code"])}</td>')
            f.write('</tr>\n')
        f.write('</table>\n')
        f.write('</body></html>\n')

def write_consolidated_html_table(all_leaks, output_path, source_files):
    """Viết HTML table tổng hợp memory leaks từ nhiều file .md"""
    # Filter out placeholder entries
    filtered_leaks = [leak for leak in all_leaks if not is_placeholder(leak)]
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Consolidated Memory Leak Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''<style>
        table{border-collapse:collapse; width: 100%;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .source-file-header{background-color:#ffe4e1; font-weight:bold; text-align:center;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .source-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        </style>\n''')
        f.write('</head><body>\n')
        f.write(f'<h1>Consolidated Memory Leak Report</h1>\n')
        f.write(f'<p>Processed {len(source_files)} files: {", ".join(source_files)}</p>\n')
        f.write(f'<p>Total memory leaks: {len(filtered_leaks)}</p>\n')
        f.write('<table>\n')
        f.write('<tr><th>Source File</th><th>Issue</th><th>Location</th><th>Problem Description</th><th>Current Code</th><th>Fixed Code</th></tr>\n')
        
        # Group by source file
        leaks_by_file = {}
        for leak in filtered_leaks:
            source = leak.get('source_file', 'unknown')
            if source not in leaks_by_file:
                leaks_by_file[source] = []
            leaks_by_file[source].append(leak)
        
        for source_file, leaks in leaks_by_file.items():
            # Header row for each file
            f.write(f'<tr class="source-file-header"><td colspan="6">File: {source_file} ({len(leaks)} memory leaks)</td></tr>\n')
            for leak in leaks:
                f.write('<tr>')
                f.write(f'<td class="source-col">{source_file}</td>')
                f.write(f'<td class="issue-col">{leak["Issue"]}</td>')
                f.write(f'<td class="location-col">{leak["Location"]}</td>')
                f.write(f'<td class="problem-col">{leak["Problem"]}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(leak["Current Code"])}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(leak["Fixed Code"])}</td>')
                f.write('</tr>\n')
        f.write('</table>\n</body></html>\n')

def write_excel_report(violations, output_path):
    """Tạo báo cáo Excel từ danh sách memory leak violations"""
    # Filter out placeholder entries
    filtered_violations = [v for v in violations if not is_placeholder(v)]
    
    # Tạo DataFrame từ violations
    data = []
    idx = 1
    for v in filtered_violations:
        data.append({
            'STT': idx,
            'Issue': v["Issue"],
            'Location': v["Location"],
            'Problem Description': v["Problem"],
            'Current Code': v["Current Code"],
            'Fixed Code': v["Fixed Code"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid memory leak violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Memory Leak Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 30,  # Issue
        'C': 30,  # Location
        'D': 40,  # Problem Description
        'E': 60,  # Current Code
        'F': 60,  # Fixed Code
        'G': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng cho code
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes để giữ header khi scroll
    ws.freeze_panes = "A2"

    # Thêm Data Validation cho cột Agree
    agree_col = 'G'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")

def write_consolidated_excel_report(all_violations, output_path, source_files):
    """Tạo báo cáo Excel tổng hợp từ nhiều file .md"""
    # Filter out placeholder entries
    filtered_violations = [v for v in all_violations if not is_placeholder(v)]
    
    # Tạo DataFrame từ all_violations
    data = []
    idx = 1
    for v in filtered_violations:
        data.append({
            'STT': idx,
            'Source File': v.get('source_file', 'unknown'),
            'Issue': v["Issue"],
            'Location': v["Location"],
            'Problem Description': v["Problem"],
            'Current Code': v["Current Code"],
            'Fixed Code': v["Fixed Code"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid memory leak violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Memory Leak Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột cho consolidated report
    column_widths = {
        'A': 8,   # STT
        'B': 20,  # Source File
        'C': 25,  # Issue
        'D': 25,  # Location
        'E': 35,  # Problem Description
        'F': 50,  # Current Code
        'G': 50,  # Fixed Code
        'H': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes
    ws.freeze_panes = "A2"

    # Data Validation cho cột Agree
    agree_col = 'H'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Consolidated Excel report saved to: {output_path}")

def print_usage():
    print("Usage: python report_leak.py <input.md> [-o <output_file>] [--excel]")
    print("  --excel: Generate Excel report instead of HTML")
    print("Examples:")
    print("  python report_leak.py input.md")
    print("  python report_leak.py input.md --excel")
    print("  python report_leak.py input.md -o report_leak.xlsx --excel")
    sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()

    input_path = sys.argv[1]
    output_format = "html"  # default
    output_path = None

    # Parse arguments
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "-o" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--excel":
            output_format = "excel"
            i += 1
        else:
            i += 1

    # Set default output path based on format
    if output_path is None:
        if output_format == "excel":
            output_path = "output_leak.xlsx"
        else:
            output_path = "output_leak.html"

    # Read and process input
    with open(input_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    violations = extract_violations(md_text)

    # Generate report based on format
    if output_format == "excel":
        write_excel_report(violations, output_path)
    else:
        write_html_table(violations, output_path)