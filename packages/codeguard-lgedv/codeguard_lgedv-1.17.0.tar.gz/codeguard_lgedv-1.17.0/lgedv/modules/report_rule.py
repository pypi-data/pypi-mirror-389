import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation
import os

def extract_violations(md_text):
    # Hỗ trợ cả tiếng Anh và tiếng Việt
    issue_patterns = [
        r'## 🚨 Issue\s*\d*:? ?([^\n]*)',  # English, có thể có số
        r'## 🚨 Vấn đề\s*\d*:? ?([^\n]*)'  # Vietnamese, có thể có số
    ]
    rule_patterns = [
        r'\*\*Rule Violated:\*\*\s*([^\n]*)',  # English
        r'\*\*Rule vi phạm:\*\*\s*([^\n]*)'  # Vietnamese
    ]
    location_patterns = [
        r'\*\*Location:\*\*\s*([^\n]*)',  # English
        r'\*\*Vị trí:\*\*\s*([^\n]*)'  # Vietnamese
    ]
    severity_patterns = [
        r'\*\*Severity:\*\*\s*([^\n]*)',  # English
        r'\*\*Mức độ:\*\*\s*([^\n]*)'  # Vietnamese
    ]
    current_code_patterns = [
        r'\*\*Current Code[^\*]*\*\*[\s\r\n]*```cpp(.*?)```',  # English, cho phép mọi ký tự sau "Current Code"
        r'\*\*Code hiện tại[^\*]*\*\*[\s\r\n]*```cpp(.*?)```'  # Vietnamese
    ]
    fixed_code_patterns = [
        r'\*\*Fixed Code:\*\*[\s\r\n]*```cpp(.*?)```',  # English
        r'\*\*Code đã sửa:\*\*[\s\r\n]*```cpp(.*?)```'  # Vietnamese
    ]
    explanation_patterns = [
        r'\*\*Explanation:\*\*\s*([^\n]*)',  # English
        r'\*\*Giải thích:\*\*\s*([^\n]*)'  # Vietnamese
    ]

    # Tách các block vi phạm
    issues = re.split(r'(?=^## 🚨 (Issue|Vấn đề))', md_text, flags=re.MULTILINE)
    results = []
    for issue in issues:
        if not (issue.strip().startswith('## 🚨 Issue') or issue.strip().startswith('## 🚨 Vấn đề')):
            continue

        def extract(patterns, text, default=''):
            for pattern in patterns:
                m = re.search(pattern, text, re.DOTALL)
                if m:
                    return m.group(1).strip()
            return default

        issue_title = extract(issue_patterns, issue)
        # Loại bỏ số thứ tự ở đầu tiêu đề nếu có (ví dụ: "1: Magic Number" -> "Magic Number")
        issue_title = re.sub(r'^\d+\s*[:\-\.]?\s*', '', issue_title)
        rule = extract(rule_patterns, issue)
        location = extract(location_patterns, issue)
        severity = extract(severity_patterns, issue)
        current_code = extract(current_code_patterns, issue)
        fixed_code = extract(fixed_code_patterns, issue)
        explanation = extract(explanation_patterns, issue)
        results.append({
            'Issue': issue_title,
            'Rule': rule,
            'Location': location,
            'Severity': severity,
            'Current Code': current_code,
            'Fixed Code': fixed_code,
            'Explanation': explanation
        })
    return results

def render_code_with_lines(code, lang="cpp"):
    lines = code.splitlines() if code else []
    if not lines:
        return '<pre><code class="language-cpp"></code></pre>'
    line_numbers = ''.join(f'<span class="line-number" style="display:block;">{i+1}</span>' for i in range(len(lines)))
    code_html = '\n'.join(lines)
    return f'''
    <div style="display:flex;">
      <div style="background:#f6f8fa;color:#888;padding:6px 8px 6px 6px;font-family:monospace;font-size:13px;line-height:1.5;text-align:right;user-select:none;border-right:1px solid #e1e4e8;">
        {line_numbers}
      </div>
      <pre style="margin:0;flex:1;"><code class="language-{lang}">{code_html}</code></pre>
    </div>
    '''

def is_placeholder(v):
    return (
        v["Issue"].strip().startswith("[#]:") or
        v["Rule"].strip().startswith("[EXACT_RULE_NUMBER]") or
        v["Location"].strip().startswith("[function name") or
        v["Severity"].strip().startswith("[Critical") or
        v["Current Code"].strip() == "[problematic code]" or
        v["Fixed Code"].strip() == "[corrected code]" or
        v["Explanation"].strip() == "[Why this violates the rule and how fix works]"
    )

def write_html_table(violations, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>LGEDV Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''
        <style>
        table{border-collapse:collapse;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .code-col { max-width: 700px; min-width: 240px; }  /* tăng size */
        .code-col > div { max-width: 700px; overflow-x: auto; }
        .rule-col { max-width: 260px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .issue-col { max-width: 180px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 80px; min-width: 40px; overflow-x: auto; word-break: break-word; } /* giảm size */
        .explanation-col { max-width: 220px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        </style>
        ''')
        f.write('</head><body>\n')
        f.write('<table>\n')
        f.write('<tr><th>Issue</th><th>Rule Violated</th><th>Location</th><th>Current Code</th><th>Fixed Code</th><th>Explanation</th></tr>\n')
        idx = 1
        for v in violations:
            if is_placeholder(v):
                continue
            f.write('<tr>')
            f.write(f'<td class="issue-col">{idx}. {v["Issue"]}</td>')
            f.write(f'<td class="rule-col">{v["Rule"]}</td>')
            f.write(f'<td class="location-col">{v["Location"]}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Current Code"])}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Fixed Code"])}</td>')
            f.write(f'<td class="explanation-col">{v["Explanation"]}</td>')
            f.write('</tr>\n')
            idx += 1
        f.write('</table>\n')
        f.write('</body></html>\n')

def write_consolidated_html_table(all_violations, output_path, source_files):
    """Viết HTML table tổng hợp từ nhiều file .md"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Consolidated Rule Violations Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''<style>
        table{border-collapse:collapse; width: 100%;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        .source-file-header{background-color:#f0f8ff; font-weight:bold; text-align:center;}
        .source-file-col { max-width: 140px; min-width: 80px; overflow-x: auto; word-break: break-all; }
        .issue-col { max-width: 180px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .rule-col { max-width: 280px; min-width: 110px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 200px; min-width: 80px; overflow-x: auto; word-break: break-all; white-space: normal; }
        .code-col { width: auto; max-width: 1000px; min-width: 200px; overflow-x: auto; resize: horizontal; }
        .explanation-col { max-width: 380px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        </style>\n''')
        f.write('</head><body>\n')
        f.write(f'<h1>Consolidated Rule Violations Report</h1>\n')
        f.write(f'<p>Processed {len(source_files)} files: {", ".join(source_files)}</p>\n')
        f.write(f'<p>Total violations: {len(all_violations)}</p>\n')
        f.write('<table>\n')
        f.write('<tr>'
                '<th class="source-file-col">Source File</th>'
                '<th class="issue-col">Issue</th>'
                '<th class="rule-col">Rule Violated</th>'
                '<th class="location-col">Location</th>'
                '<th class="code-col">Current Code</th>'
                '<th class="code-col">Fixed Code</th>'
                '<th class="explanation-col">Explanation</th>'
                '</tr>\n')
        
        # Group by source file
        violations_by_file = {}
        for v in all_violations:
            source = v.get('source_file', 'unknown')
            if source not in violations_by_file:
                violations_by_file[source] = []
            violations_by_file[source].append(v)
        
        idx = 1
        for source_file, violations in violations_by_file.items():
            f.write(f'<tr class="source-file-header"><td colspan="7">File: {source_file} ({len(violations)} violations)</td></tr>\n')
            for v in violations:
                if is_placeholder(v):
                    continue
                f.write('<tr>')
                f.write(f'<td class="source-file-col">{source_file}</td>')
                f.write(f'<td class="issue-col">{idx}. {v["Issue"]}</td>')
                f.write(f'<td class="rule-col">{v["Rule"]}</td>')
                f.write(f'<td class="location-col">{v["Location"]}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(v["Current Code"])}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(v["Fixed Code"])}</td>')
                f.write(f'<td class="explanation-col">{v["Explanation"]}</td>')
                f.write('</tr>\n')
                idx += 1
        f.write('</table>\n</body></html>\n')

def write_excel_report(violations, output_path):
    """Tạo báo cáo Excel từ danh sách violations cho single file"""
    # Tạo DataFrame từ violations
    data = []
    idx = 1
    for v in violations:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Issue': v["Issue"],
            'Rule Violated': v["Rule"],
            'Location': v["Location"],
            'Severity': v.get("Severity", ""),
            'Current Code': v["Current Code"],
            'Fixed Code': v["Fixed Code"],
            'Explanation': v["Explanation"],
            'Agree': "Yes"  # Mặc định là yes
        })
        idx += 1

    if not data:
        print("No valid violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Rule Violations Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 25,  # Issue
        'C': 35,  # Rule Violated
        'D': 20,  # Location
        'E': 12,  # Severity
        'F': 60,  # Current Code
        'G': 60,  # Fixed Code
        'H': 40,  # Explanation
        'I': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng cho code
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes để giữ header khi scroll
    ws.freeze_panes = "A2"

    # Thêm Data Validation cho cột Agree
    agree_col = 'I'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")

def write_consolidated_excel_report(all_violations, output_path, source_files):
    """Tạo báo cáo Excel tổng hợp từ nhiều file .md"""
    # Tạo DataFrame từ all_violations
    data = []
    idx = 1
    for v in all_violations:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Source File': v.get('source_file', 'unknown'),
            'Issue': v["Issue"],
            'Rule Violated': v["Rule"],
            'Location': v["Location"],
            'Severity': v.get("Severity", ""),
            'Current Code': v["Current Code"],
            'Fixed Code': v["Fixed Code"],
            'Explanation': v["Explanation"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid rule violations found to export.")
        return

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Rule Violations"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 20,  # Source File
        'C': 25,  # Issue
        'D': 35,  # Rule Violated
        'E': 20,  # Location
        'F': 12,  # Severity
        'G': 60,  # Current Code
        'H': 60,  # Fixed Code
        'I': 40,  # Explanation
        'J': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng cho code
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes để giữ header khi scroll
    ws.freeze_panes = "A2"

    # Thêm Data Validation cho cột Agree
    agree_col = 'J'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Consolidated Excel report saved to: {output_path}")

def print_usage():
    print("Usage: python report_rule.py <input.md> [-o <output_file>] [--excel]")
    print("  --excel: Generate Excel report instead of HTML")
    print("Examples:")
    print("  python report_rule.py input.md")
    print("  python report_rule.py input.md --excel")
    print("  python report_rule.py input.md -o report.xlsx --excel")
    sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()

    input_path = sys.argv[1]
    output_format = "html"  # default
    output_path = None

    # Parse arguments
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "-o" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--excel":
            output_format = "excel"
            i += 1
        else:
            i += 1

    # Set default output path based on format
    if output_path is None:
        if output_format == "excel":
            output_path = "output.xlsx"
        else:
            output_path = "output.html"

    # Read and process input
    with open(input_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    violations = extract_violations(md_text)

    # Generate report based on format
    if output_format == "excel":
        write_excel_report(violations, output_path)
    else:
        write_html_table(violations, output_path)