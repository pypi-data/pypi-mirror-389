import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

def extract_violations(md_text):
    # Lọc các block RESOURCE LEAK - tìm từ ## 🚨 đến ---
    # Bỏ qua dòng format mẫu có [number] và [Brief Description]
    leak_pattern = r'## 🚨 (RESOURCE LEAK #\d+: [^\n]+)(.*?)(?=^## 🚨 |^---|\Z)'
    matches = re.findall(leak_pattern, md_text, flags=re.DOTALL | re.MULTILINE)
    results = []
    for title, block in matches:
        # Type - lấy từ **Type:**
        m = re.search(r'\*\*Type:\*\*\s*([^\n]+)', block)
        resource_type = m.group(1).strip() if m else ''
        
        # Severity - lấy từ **Severity:**
        m = re.search(r'\*\*Severity:\*\*\s*([^\n]+)', block)
        severity = m.group(1).strip() if m else ''
        
        # Files Involved - lấy từ **Files Involved:**
        m = re.search(r'\*\*Files Involved:\*\*\s*([^\n]+)', block)
        files_involved = m.group(1).strip() if m else ''
        
        # Location - lấy từ **Location:**
        m = re.search(r'\*\*Location:\*\*\s*([^\n]+)', block)
        location = m.group(1).strip() if m else ''
        
        # Problem - lấy toàn bộ từ **Problem:** đến **Current Code:**
        m = re.search(r'\*\*Problem:\*\*\s*(.*?)(?=\*\*Current Code|\*\*Fix Recommendation|$)', block, re.DOTALL)
        problem = m.group(1).strip() if m else ''
        
        # Current Code - lấy từ **Current Code:** đến ```
        m = re.search(r'\*\*Current Code[^\*]*\*\*\s*```cpp\n?(.*?)```', block, re.DOTALL)
        current_code = m.group(1).strip() if m else ''
        
        # Fix Recommendation - ưu tiên lấy block cuối cùng (Modern C++ Best Practice)
        fix_blocks = re.findall(r'```cpp\n?(.*?)```', block, re.DOTALL)
        fix_recommendation = ''
        if fix_blocks:
            # Ưu tiên lấy block cuối cùng nếu có nhiều block
            if len(fix_blocks) > 1:
                fix_recommendation = fix_blocks[-1]  # Block cuối cùng (Modern C++ Best Practice)
            else:
                # Nếu chỉ có 1 block, kiểm tra xem có khác Current Code không
                if fix_blocks[0].strip() != current_code.strip():
                    fix_recommendation = fix_blocks[0]
                else:
                    fix_recommendation = fix_blocks[0]  # Vẫn lấy nếu giống
        
        results.append({
            'Issue': title.strip(),
            'Type': resource_type,
            'Severity': severity,
            'Files Involved': files_involved,
            'Location': location,
            'Problem': problem,
            'Current Code': current_code,
            'Fix Recommendation': fix_recommendation
        })
    return results

def render_code_with_lines(code, lang="cpp"):
    lines = code.splitlines() if code else []
    if not lines:
        return '<pre><code class="language-cpp"></code></pre>'
    line_numbers = ''.join(f'<span class="line-number" style="display:block;">{i+1}</span>' for i in range(len(lines)))
    code_html = '\n'.join(lines)
    return f'''
    <div style="display:flex;">
      <div style="background:#f6f8fa;color:#888;padding:6px 8px 6px 6px;font-family:monospace;font-size:13px;line-height:1.5;text-align:right;user-select:none;border-right:1px solid #e1e4e8;">
        {line_numbers}
      </div>
      <pre style="margin:0;flex:1;"><code class="language-{lang}">{code_html}</code></pre>
    </div>
    '''

def is_placeholder(v):
    return (
        v["Issue"].strip().startswith("[#]:") or
        v["Type"].strip().startswith("[resource type]") or
        v["Severity"].strip().startswith("[Critical") or
        v["Files Involved"].strip().startswith("[list of files]") or
        v["Location"].strip().startswith("[function name") or
        v["Current Code"].strip() == "[show problematic code]" or
        v["Fix Recommendation"].strip() == "[suggested code or approach]" or
        v["Problem"].strip() == "[explanation]"
    )

def write_html_table(violations, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Resource Leak Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''
        <style>
        table{border-collapse:collapse;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .type-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        .severity-col { max-width: 100px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .files-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .fix-col { max-width: 400px; min-width: 200px; overflow-x: auto; word-break: break-word; }
        </style>
        ''')
        f.write('</head><body>\n')
        f.write('<table>\n')
        f.write('<tr><th>Issue</th><th>Type</th><th>Severity</th><th>Files Involved</th><th>Location</th><th>Problem</th><th>Current Code</th><th>Fix Recommendation</th></tr>\n')
        for v in violations:
            f.write('<tr>')
            f.write(f'<td class="issue-col">{v["Issue"]}</td>')
            f.write(f'<td class="type-col">{v["Type"]}</td>')
            f.write(f'<td class="severity-col">{v["Severity"]}</td>')
            f.write(f'<td class="files-col">{v["Files Involved"]}</td>')
            f.write(f'<td class="location-col">{v["Location"]}</td>')
            f.write(f'<td class="problem-col">{v["Problem"]}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Current Code"])}</td>')
            f.write(f'<td class="fix-col">{render_code_with_lines(v["Fix Recommendation"])}</td>')
            f.write('</tr>\n')
        f.write('</table>\n')
        f.write('</body></html>\n')

def write_consolidated_html_table(all_leaks, output_path, source_files):
    """Viết HTML table tổng hợp resource leaks từ nhiều file .md"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Consolidated Resource Leak Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''<style>
        table{border-collapse:collapse; width: 100%;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .source-file-header{background-color:#d4edda; font-weight:bold; text-align:center;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .type-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        .severity-col { max-width: 100px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .files-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .location-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .source-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        </style>''')
        f.write('</head><body>\n')
        f.write(f'<h1>Consolidated Resource Leak Report</h1>\n')
        f.write(f'<p>Processed {len(source_files)} files: {", ".join(source_files)}</p>\n')
        f.write(f'<p>Total resource leaks: {len(all_leaks)}</p>\n')
        f.write('<table>\n')
        f.write('<tr><th>Source File</th><th>Issue</th><th>Type</th><th>Severity</th><th>Files Involved</th><th>Location</th><th>Problem</th><th>Current Code</th><th>Fix Recommendation</th></tr>\n')
        
        # Group by source file
        leaks_by_file = {}
        for leak in all_leaks:
            source = leak.get('source_file', 'unknown')
            if source not in leaks_by_file:
                leaks_by_file[source] = []
            leaks_by_file[source].append(leak)
        
        for source_file, leaks in leaks_by_file.items():
            # Header row for each file
            f.write(f'<tr class="source-file-header"><td colspan="9">File: {source_file} ({len(leaks)} resource leaks)</td></tr>\n')
            for leak in leaks:
                f.write('<tr>')
                f.write(f'<td class="source-col">{source_file}</td>')
                f.write(f'<td class="issue-col">{leak["Issue"]}</td>')
                f.write(f'<td class="type-col">{leak["Type"]}</td>')
                f.write(f'<td class="severity-col">{leak["Severity"]}</td>')
                f.write(f'<td class="files-col">{leak["Files Involved"]}</td>')
                f.write(f'<td class="location-col">{leak["Location"]}</td>')
                f.write(f'<td class="problem-col">{leak["Problem"]}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(leak["Current Code"])}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(leak["Fix Recommendation"])}</td>')
                f.write('</tr>\n')
        
        f.write('</table>\n')
        f.write('</body></html>\n')

def write_excel_report(violations, output_path):
    """Tạo báo cáo Excel từ danh sách resource leak violations"""
    # Tạo DataFrame từ violations
    data = []
    idx = 1
    for v in violations:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Issue': v["Issue"],
            'Type': v["Type"],
            'Severity': v["Severity"],
            'Files Involved': v["Files Involved"],
            'Location': v["Location"],
            'Problem': v["Problem"],
            'Current Code': v["Current Code"],
            'Fix Recommendation': v["Fix Recommendation"],
            'Agree': "Yes"  # Mặc định là yes
        })
        idx += 1

    if not data:
        print("No valid resource leak violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Resource Leak Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 25,  # Issue
        'C': 20,  # Type
        'D': 12,  # Severity
        'E': 25,  # Files Involved
        'F': 20,  # Location
        'G': 40,  # Problem
        'H': 60,  # Current Code
        'I': 60,  # Fix Recommendation
        'J': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng cho code
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes để giữ header khi scroll
    ws.freeze_panes = "A2"

    # Thêm Data Validation cho cột Agree
    agree_col = 'J'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")

def write_consolidated_excel_report(all_violations, output_path, source_files):
    """Tạo báo cáo Excel tổng hợp từ nhiều file .md"""
    # Tạo DataFrame từ all_violations
    data = []
    idx = 1
    for v in all_violations:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Source File': v.get('source_file', 'unknown'),
            'Issue': v["Issue"],
            'Type': v["Type"],
            'Severity': v["Severity"],
            'Files Involved': v["Files Involved"],
            'Location': v["Location"],
            'Problem': v["Problem"],
            'Current Code': v["Current Code"],
            'Fix Recommendation': v["Fix Recommendation"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid resource leak violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Resource Leak Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột cho consolidated report
    column_widths = {
        'A': 8,   # STT
        'B': 20,  # Source File
        'C': 25,  # Issue
        'D': 15,  # Type
        'E': 12,  # Severity
        'F': 25,  # Files Involved
        'G': 20,  # Location
        'H': 35,  # Problem
        'I': 50,  # Current Code
        'J': 50,  # Fix Recommendation
        'K': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes
    ws.freeze_panes = "A2"

    # Data Validation cho cột Agree
    agree_col = 'K'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Consolidated Excel report saved to: {output_path}")

def print_usage():
    print("Usage: python report_resource.py <input.md> [-o <output_file>] [--excel]")
    print("  --excel: Generate Excel report instead of HTML")
    print("Examples:")
    print("  python report_resource.py input.md")
    print("  python report_resource.py input.md --excel")
    print("  python report_resource.py input.md -o report_resource.xlsx --excel")
    sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()

    input_path = sys.argv[1]
    output_format = "html"  # default
    output_path = None

    # Parse arguments
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "-o" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--excel":
            output_format = "excel"
            i += 1
        else:
            i += 1

    # Set default output path based on format
    if output_path is None:
        if output_format == "excel":
            output_path = "output_resource.xlsx"
        else:
            output_path = "output_resource.html"

    # Read and process input
    with open(input_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    violations = extract_violations(md_text)

    # Generate report based on format
    if output_format == "excel":
        write_excel_report(violations, output_path)
    else:
        write_html_table(violations, output_path)