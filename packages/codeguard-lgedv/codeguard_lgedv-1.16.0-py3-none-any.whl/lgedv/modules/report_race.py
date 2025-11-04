import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

def extract_races(md_text):
    # Lọc các block RACE CONDITION - tìm từ ## 🚨 đến ---
    race_pattern = r'## 🚨 (RACE CONDITION #\d+: [^\n]+)(.*?)(?=^## 🚨 |^---|\Z)'
    matches = re.findall(race_pattern, md_text, flags=re.DOTALL | re.MULTILINE)
    results = []
    for title, block in matches:
        # Type - lấy từ **Type:**
        m = re.search(r'\*\*Type:\*\*\s*([^\n]+)', block)
        race_type = m.group(1).strip() if m else ''
        
        # Severity - lấy từ **Severity:**
        m = re.search(r'\*\*Severity:\*\*\s*([^\n]+)', block)
        severity = m.group(1).strip() if m else ''
        
        # Files Involved - lấy từ **Files Involved:**
        m = re.search(r'\*\*Files Involved:\*\*\s*([^\n]+)', block)
        files_involved = m.group(1).strip() if m else ''
        
        # Function Name - lấy từ **Function Name:** nếu có, nếu không thì lấy từ **Location:**
        m = re.search(r'\*\*Function Name:\*\*\s*([^\n]+)', block)
        function_name = m.group(1).strip() if m else ''
        if not function_name:
            m = re.search(r'\*\*Location:\*\*\s*([^\n]+)', block)
            function_name = m.group(1).strip() if m else ''
        
        # Problem - lấy từ sau **Problem Description:** đến trước **Current Code:**
        m = re.search(r'\*\*Problem Description:\*\*\s*(.*?)(?=\*\*Current Code:|\*\*Fix Recommendation:|$)', block, re.DOTALL)
        problem = m.group(1).strip() if m else ''
        
        # Current Code - lấy từ sau **Current Code:** đến trước **Fix Recommendation:**
        m = re.search(r'\*\*Current Code[^\*]*\*\*\s*(.*?)(?=\*\*Fix Recommendation:|$)', block, re.DOTALL)
        current_code_text = m.group(1).strip() if m else ''
        
        # Trích xuất code từ current_code_text (lấy nội dung trong ```cpp...```)
        code_match = re.search(r'```cpp\n?(.*?)```', current_code_text, re.DOTALL)
        current_code = code_match.group(1).strip() if code_match else ''
        
        # Fix Recommendation - lấy từ sau **Fix Recommendation:** đến trước ---
        m = re.search(r'\*\*Fix Recommendation:\*\*\s*(.*?)(?=^---|\Z)', block, re.DOTALL | re.MULTILINE)
        fix_text = m.group(1).strip() if m else ''
        
        # Trích xuất code từ fix_text (lấy nội dung trong ```cpp...```)
        fix_code_match = re.search(r'```cpp\n?(.*?)```', fix_text, re.DOTALL)
        fix_recommendation = fix_code_match.group(1).strip() if fix_code_match else ''
        
        results.append({
            'Issue': title.strip(),
            'Type': race_type,
            'Severity': severity,
            'Files Involved': files_involved,
            'Function Name': function_name,
            'Problem': problem,
            'Current Code': current_code,
            'Fix Recommendation': fix_recommendation
        })
    return results

def render_code_with_lines(code, lang="cpp"):
    lines = code.splitlines() if code else []
    if not lines:
        return '<pre><code class="language-cpp"></code></pre>'
    line_numbers = ''.join(f'<span class="line-number" style="display:block;">{i+1}</span>' for i in range(len(lines)))
    code_html = '\n'.join(lines)
    return f'''
    <div style="display:flex;">
      <div style="background:#f6f8fa;color:#888;padding:6px 8px 6px 6px;font-family:monospace;font-size:13px;line-height:1.5;text-align:right;user-select:none;border-right:1px solid #e1e4e8;">
        {line_numbers}
      </div>
      <pre style="margin:0;flex:1;"><code class="language-{lang}">{code_html}</code></pre>
    </div>
    '''

def is_placeholder(v):
    """Kiểm tra xem race condition có phải là placeholder không"""
    return (
        v["Issue"].strip().startswith("[#]:") or
        v["Type"].strip().startswith("[race type]") or
        v["Severity"].strip().startswith("[Critical") or
        v["Files Involved"].strip().startswith("[list of files]") or
        v["Function Name"].strip().startswith("[function name") or
        v["Current Code"].strip() == "[show problematic code]" or
        v["Fix Recommendation"].strip() == "[suggested code or approach]" or
        v["Problem"].strip() == "[explanation]"
    )

def write_html_table(races, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Race Condition Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''
        <style>
        table{border-collapse:collapse;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .type-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        .severity-col { max-width: 100px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .files-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .function-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .fix-col { max-width: 400px; min-width: 200px; overflow-x: auto; word-break: break-word; }
        </style>
        ''')
        f.write('</head><body>\n')
        f.write('<table>\n')
        f.write('<tr><th>Issue</th><th>Type</th><th>Severity</th><th>Files Involved</th><th>Function/Location</th><th>Problem</th><th>Current Code</th><th>Fix Recommendation</th></tr>\n')
        for v in races:
            f.write('<tr>')
            f.write(f'<td class="issue-col">{v["Issue"]}</td>')
            f.write(f'<td class="type-col">{v["Type"]}</td>')
            f.write(f'<td class="severity-col">{v["Severity"]}</td>')
            f.write(f'<td class="files-col">{v["Files Involved"]}</td>')
            f.write(f'<td class="function-col">{v["Function Name"]}</td>')
            f.write(f'<td class="problem-col">{v["Problem"]}</td>')
            f.write(f'<td class="code-col">{render_code_with_lines(v["Current Code"])}</td>')
            f.write(f'<td class="fix-col">{render_code_with_lines(v["Fix Recommendation"])}</td>')
            f.write('</tr>\n')
        f.write('</table>\n')
        f.write('</body></html>\n')

def write_consolidated_html_table(all_races, output_path, source_files):
    """Viết HTML table tổng hợp race conditions từ nhiều file .md"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<!DOCTYPE html>\n<html><head><meta charset="utf-8"><title>Consolidated Race Condition Report</title>\n')
        f.write('<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>\n')
        f.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>\n')
        f.write('<script>hljs.highlightAll();</script>\n')
        f.write('''<style>
        table{border-collapse:collapse; width: 100%;}
        th,td{border:1px solid #888;padding:6px;vertical-align:top;}
        pre{margin:0;}
        .source-file-header{background-color:#fff3cd; font-weight:bold; text-align:center;}
        .code-col { max-width: 480px; min-width: 180px; }
        .code-col > div { max-width: 480px; overflow-x: auto; }
        .issue-col { max-width: 320px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .type-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        .severity-col { max-width: 100px; min-width: 80px; overflow-x: auto; word-break: break-word; }
        .files-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .function-col { max-width: 200px; min-width: 120px; overflow-x: auto; word-break: break-word; }
        .problem-col { max-width: 340px; min-width: 140px; overflow-x: auto; word-break: break-word; }
        .source-col { max-width: 150px; min-width: 100px; overflow-x: auto; word-break: break-word; }
        </style>\n''')
        f.write('</head><body>\n')
        f.write(f'<h1>Consolidated Race Condition Report</h1>\n')
        f.write(f'<p>Processed {len(source_files)} files: {", ".join(source_files)}</p>\n')
        f.write(f'<p>Total race conditions: {len(all_races)}</p>\n')
        f.write('<table>\n')
        f.write('<tr><th>Source File</th><th>Issue</th><th>Type</th><th>Severity</th><th>Files Involved</th><th>Function/Location</th><th>Problem</th><th>Current Code</th><th>Fix Recommendation</th></tr>\n')
        
        # Group by source file
        races_by_file = {}
        for race in all_races:
            source = race.get('source_file', 'unknown')
            if source not in races_by_file:
                races_by_file[source] = []
            races_by_file[source].append(race)
        
        for source_file, races in races_by_file.items():
            # Header row for each file
            f.write(f'<tr class="source-file-header"><td colspan="9">File: {source_file} ({len(races)} race conditions)</td></tr>\n')
            for race in races:
                f.write('<tr>')
                f.write(f'<td class="source-col">{source_file}</td>')
                f.write(f'<td class="issue-col">{race["Issue"]}</td>')
                f.write(f'<td class="type-col">{race["Type"]}</td>')
                f.write(f'<td class="severity-col">{race["Severity"]}</td>')
                f.write(f'<td class="files-col">{race["Files Involved"]}</td>')
                f.write(f'<td class="function-col">{race["Function Name"]}</td>')
                f.write(f'<td class="problem-col">{race["Problem"]}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(race["Current Code"])}</td>')
                f.write(f'<td class="code-col">{render_code_with_lines(race["Fix Recommendation"])}</td>')
                f.write('</tr>\n')
        f.write('</table>\n</body></html>\n')

def write_excel_report(races, output_path):
    """Tạo báo cáo Excel từ danh sách race condition violations"""
    # Tạo DataFrame từ races
    data = []
    idx = 1
    for v in races:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Issue': v["Issue"],
            'Type': v["Type"],
            'Severity': v["Severity"],
            'Files Involved': v["Files Involved"],
            'Function Name': v["Function Name"],
            'Problem': v["Problem"],
            'Current Code': v["Current Code"],
            'Fix Recommendation': v["Fix Recommendation"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid race condition violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Race Condition Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 25,  # Issue
        'C': 15,  # Type
        'D': 12,  # Severity
        'E': 25,  # Files Involved
        'F': 20,  # Function Name
        'G': 40,  # Problem
        'H': 60,  # Current Code
        'I': 60,  # Fix Recommendation
        'J': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng cho code
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes để giữ header khi scroll
    ws.freeze_panes = "A2"

    # Thêm Data Validation cho cột Agree
    agree_col = 'J'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")

def write_consolidated_excel_report(all_races, output_path, source_files):
    """Tạo báo cáo Excel tổng hợp từ nhiều file .md"""
    # Tạo DataFrame từ all_races
    data = []
    idx = 1
    for v in all_races:
        if is_placeholder(v):
            continue
        data.append({
            'STT': idx,
            'Source File': v.get('source_file', 'unknown'),
            'Issue': v["Issue"],
            'Type': v["Type"],
            'Severity': v["Severity"],
            'Files Involved': v["Files Involved"],
            'Function Name': v["Function Name"],
            'Problem': v["Problem"],
            'Current Code': v["Current Code"],
            'Fix Recommendation': v["Fix Recommendation"],
            'Agree': "Yes"  # Mặc định là Yes
        })
        idx += 1

    if not data:
        print("No valid race condition violations found to export.")
        return

    df = pd.DataFrame(data)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Race Condition Report"

    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Thêm header
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm dữ liệu
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Điều chỉnh độ rộng cột cho consolidated report
    column_widths = {
        'A': 8,   # STT
        'B': 20,  # Source File
        'C': 25,  # Issue
        'D': 15,  # Type
        'E': 12,  # Severity
        'F': 25,  # Files Involved
        'G': 20,  # Function Name
        'H': 35,  # Problem
        'I': 50,  # Current Code
        'J': 50,  # Fix Recommendation
        'K': 10   # Agree
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Đặt chiều cao hàng
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    # Freeze panes
    ws.freeze_panes = "A2"

    # Data Validation cho cột Agree
    agree_col = 'K'
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(2, len(data) + 2):
        dv.add(f"{agree_col}{row}")

    # Lưu file
    wb.save(output_path)
    print(f"Consolidated Excel report saved to: {output_path}")

def print_usage():
    print("Usage: python report_race.py <input.md> [-o <output_file>] [--excel]")
    print("  --excel: Generate Excel report instead of HTML")
    print("Examples:")
    print("  python report_race.py input.md")
    print("  python report_race.py input.md --excel")
    print("  python report_race.py input.md -o report_race.xlsx --excel")
    sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()

    input_path = sys.argv[1]
    output_format = "html"  # default
    output_path = None

    # Parse arguments
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "-o" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--excel":
            output_format = "excel"
            i += 1
        else:
            i += 1

    # Set default output path based on format
    if output_path is None:
        if output_format == "excel":
            output_path = "output_race.xlsx"
        else:
            output_path = "output_race.html"

    # Read and process input
    with open(input_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    races = extract_races(md_text)

    # Generate report based on format
    if output_format == "excel":
        write_excel_report(races, output_path)
    else:
        write_html_table(races, output_path)