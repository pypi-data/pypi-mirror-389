import re
from dataclasses import dataclass
from typing import Optional

import pandas as pd
import pooch
from openpyxl import load_workbook


class _Global_Carbon_Budget_Release(dict):

    def __init__(
        self,
        name: str,
        version: str,
        doi: str,
        doi_article: str,
        published: str,
        citation: str,
        citation_article: str,
        license: str,
        filename: str,
        url: str,
        known_hash: str,
        sheets: list,
    ):
        self.name = name
        self.version = version
        self.doi = doi
        self.doi_article = doi_article
        self.published = published
        self.citation = citation
        self.citation_article = citation_article
        self.license = license
        self.filename = filename
        self.url = url
        self.known_hash = known_hash

        for sheet in sheets:
            self[sheet["sheet_name"]] = _Global_Carbon_Budget_Sheet(
                release=self, **sheet
            )

    def _get_file_path(self):
        return pooch.retrieve(
            path=pooch.os_cache("openclimatedata/global-carbon-budget"),
            fname=self.filename,
            url=self.url,
            known_hash=self.known_hash,
        )

    def __repr__(self):
        return f"""{self.name}

'{self.filename}'

License: {self.license}
https://doi.org/{self.doi}

{self.citation}"""


class _Global_Carbon_Budget_Sheet(dict):
    def __init__(
        self,
        release: object,
        sheet_name: str,
        skiprows: int,
        columns: Optional[str] = None,
        tables: Optional[list] = None,
    ):
        self.release = release
        self.sheet_name = sheet_name
        self.skiprows = skiprows
        self.columns = columns

        if tables:
            for table in tables:
                self[table["table_name"]] = _Global_Carbon_Budget_Table(
                    sheet=self,
                    table_name=table["table_name"],
                    skiprows=table["skiprows"],
                    columns=table["columns"],
                )
        else:
            self.to_dataframe = self._to_dataframe
            self.to_long_dataframe = self._to_long_dataframe

    def __repr__(self):
        if not hasattr(self, "note"):
            file_path = self.release._get_file_path()
            wb = load_workbook(file_path)
            rows = [
                list(row)
                for row in list(
                    wb[self.sheet_name].iter_rows(
                        max_row=self.skiprows, values_only=True
                    )
                )
            ]
            note = ""
            for row in rows:
                note += ("\t".join([cell if cell else "\t" for cell in row])).strip()
                note += "\n"

            self.note = note

        return f"""{self.release.name} — {self.sheet_name}

{self.note}
"""

    def _to_dataframe(self):
        file_path = self.release._get_file_path()
        df = pd.read_excel(
            file_path,
            sheet_name=self.sheet_name,
            skiprows=self.skiprows,
            usecols=self.columns,
            index_col=0,
        )
        df.index.name = "Year"
        return df

    def _to_long_dataframe(self):
        df = self.to_dataframe()
        value_vars = df.columns
        df = df.reset_index().melt(
            id_vars=["Year"],
            value_vars=value_vars,
            var_name="Category",
            value_name="Value",
        )
        df.Category = df.Category.astype("category")
        return df


@dataclass
class _Global_Carbon_Budget_Table:
    sheet: object
    table_name: str
    skiprows: int
    columns: str

    def __repr__(self):
        return f"""{self.sheet.sheet_name} - {self.table_name} - {self.columns}"""

    def to_dataframe(self):
        file_path = self.sheet.release._get_file_path()
        df = pd.read_excel(
            file_path,
            sheet_name=self.sheet.sheet_name,
            skiprows=self.skiprows,
            usecols=self.columns,
            index_col=0,
        )
        # Remove suffixes from duplicated columns names (added by Pandas).
        df.columns = df.columns.map(lambda x: re.sub(r"\.\d$", "", x))
        df.name = self.table_name
        df.index.name = "Year"
        return df

    def to_long_dataframe(self):
        df = self.to_dataframe()
        value_vars = df.columns
        df = df.reset_index().melt(
            id_vars=["Year"],
            value_vars=value_vars,
            var_name="Category",
            value_name="Value",
        )
        df.Category = df.Category.astype("category")
        return df
