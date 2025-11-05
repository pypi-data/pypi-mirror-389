# -*- coding:UTF-8 -*-
import os
from typing import List, Optional
import glob
import openpyxl
import xlrd
import os


def get_files(directory: str, extensions: Optional[List[str]] = None, recursive: bool = False) -> List[str]:
    """
    获取指定目录下所有指定后缀的文件列表

    :param directory: 要搜索的目录路径
    :param extensions: 需要筛选的文件后缀列表（例如 ['.csv', '.txt']），如果为 None 则返回所有文件
    :param recursive: 是否递归查找子目录
    :return: 符合条件的文件路径列表
    """
    files_list: List[str] = []

    if recursive:
        # 递归查找所有子目录
        for root, _, files in os.walk(directory):
            for file in files:
                if extensions is None or any(file.endswith(ext) for ext in extensions):
                    files_list.append(os.path.join(root, file))
    else:
        # 仅查找当前目录
        for file in os.listdir(directory):
            if extensions is None or any(file.endswith(ext) for ext in extensions):
                files_list.append(os.path.join(directory, file))

    return files_list


def del_files(directory: str, extensions: Optional[List[str]] = None, recursive: bool = False):
    """
    删除指定目录下所有指定后缀的文件

    :param directory: 要搜索的目录路径
    :param extensions: 需要筛选的文件后缀列表（例如 ['.csv', '.txt']），如果为 None 则返回所有文件
    :param recursive: 是否递归查找子目录
    """
    # 构建搜索模式
    if recursive:
        search_pattern = os.path.join(directory, '**', '*')
    else:
        search_pattern = os.path.join(directory, '*')

    # 扩展搜索模式为具体文件后缀
    for extension in extensions:
        files_to_delete = glob.glob(f"{search_pattern}.{extension}", recursive=recursive)

        # 删除文件
        for file in files_to_delete:
            try:
                os.remove(file)
            except Exception as e:
                print(f"Error deleting {file}: {e}")


def xlsx_to_html(file_path: str) -> str:
    """
    将xlsx转换为html  也就是表格 table 包含跨行跨列的处理
    :param file_path:
    :return:
    """
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Extract data and merge information
    data = list(ws.iter_rows(values_only=True))

    merge_map = {}
    for merged_cell in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell.bounds
        merge_map[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)

    # Determine non-empty rows from top
    top_non_empty_index = 0
    while top_non_empty_index < len(data) and all(cell is None for cell in data[top_non_empty_index]):
        top_non_empty_index += 1
    data = data[top_non_empty_index:]

    # Adjust merge map for removed top rows
    adjusted_merge_map = {}
    for (min_row, min_col), (rowspan, colspan) in merge_map.items():
        if min_row >= top_non_empty_index + 1:
            adjusted_merge_map[(min_row - top_non_empty_index, min_col)] = (rowspan, colspan)

    # Determine non-empty rows from bottom
    while data and all(cell is None for cell in data[-1]):
        data.pop()

    # Determine non-empty columns from left
    if data:
        max_col_index = max(len(row) for row in data)
        left_col_index = 0
        while left_col_index < max_col_index:
            if all(row[left_col_index] is None for row in data if left_col_index < len(row)):
                left_col_index += 1
            else:
                break
        # Trim left columns
        data = [row[left_col_index:] for row in data]

    # Adjust merge map for removed left columns
    final_merge_map = {}
    for (min_row, min_col), (rowspan, colspan) in adjusted_merge_map.items():
        if min_col >= left_col_index + 1:
            final_merge_map[(min_row, min_col - left_col_index)] = (rowspan, colspan)

    # Determine non-empty columns from right
    if data:
        max_col_index = max(len(row) for row in data)
        right_col_index = max_col_index
        while right_col_index > 0:
            if all(row[right_col_index - 1] is None for row in data if right_col_index - 1 < len(row)):
                right_col_index -= 1
            else:
                break
        # Trim right columns
        data = [row[:right_col_index] for row in data]

    # Create the HTML table, skipping empty rows and columns
    html = '<table border="1">'

    skip_cells = set()
    for row_idx, row in enumerate(data, start=1):
        html += '  <tr>'
        for col_idx, value in enumerate(row, start=1):
            if (row_idx, col_idx) in skip_cells:
                continue

            rowspan, colspan = 1, 1
            if (row_idx, col_idx) in final_merge_map:
                rowspan, colspan = final_merge_map[(row_idx, col_idx)]
                for r in range(row_idx, row_idx + rowspan):
                    for c in range(col_idx, col_idx + colspan):
                        if (r, c) != (row_idx, col_idx):
                            skip_cells.add((r, c))

            html += f'    <td rowspan="{rowspan}" colspan="{colspan}">{value if value is not None else ""}</td>'
        html += '  </tr>'

    html += '</table>'

    return html


def xls_to_html(file_path: str) -> str:
    """
    将xls转换为html  也就是表格 table 包含跨行跨列的处理
    :param file_path:
    :return:
    """
    # Open the workbook and select the first sheet
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    ws = wb.sheet_by_index(0)

    # Extract data and merge information
    data = [ws.row_values(rowx) for rowx in range(ws.nrows)]

    # Get merged cells information
    merge_map = {}
    for merged_cell in ws.merged_cells:
        min_row, max_row, min_col, max_col = merged_cell
        merge_map[(min_row + 1, min_col + 1)] = (max_row - min_row, max_col - min_col)

    # Determine non-empty rows from top
    top_non_empty_index = 0
    while top_non_empty_index < len(data) and all(cell == "" for cell in data[top_non_empty_index]):
        top_non_empty_index += 1
    data = data[top_non_empty_index:]

    # Adjust merge map for removed top rows
    adjusted_merge_map = {}
    for (min_row, min_col), (rowspan, colspan) in merge_map.items():
        if min_row >= top_non_empty_index + 1:
            adjusted_merge_map[(min_row - top_non_empty_index, min_col)] = (rowspan, colspan)

    # Determine non-empty rows from bottom
    while data and all(cell == "" for cell in data[-1]):
        data.pop()

    # Determine non-empty columns from left
    if data:
        max_col_index = max(len(row) for row in data)
        left_col_index = 0
        while left_col_index < max_col_index:
            if all(row[left_col_index] == "" for row in data if left_col_index < len(row)):
                left_col_index += 1
            else:
                break
        # Trim left columns
        data = [row[left_col_index:] for row in data]

    # Adjust merge map for removed left columns
    final_merge_map = {}
    for (min_row, min_col), (rowspan, colspan) in adjusted_merge_map.items():
        if min_col >= left_col_index + 1:
            final_merge_map[(min_row, min_col - left_col_index)] = (rowspan, colspan)

    # Determine non-empty columns from right
    if data:
        max_col_index = max(len(row) for row in data)
        right_col_index = max_col_index
        while right_col_index > 0:
            if all(row[right_col_index - 1] == "" for row in data if right_col_index - 1 < len(row)):
                right_col_index -= 1
            else:
                break
        # Trim right columns
        data = [row[:right_col_index] for row in data]

    # Create the HTML table, skipping empty rows and columns
    html = '<table border="1">'

    skip_cells = set()
    for row_idx, row in enumerate(data, start=1):
        html += '  <tr>'
        for col_idx, value in enumerate(row, start=1):
            if (row_idx, col_idx) in skip_cells:
                continue

            rowspan, colspan = 1, 1
            if (row_idx, col_idx) in final_merge_map:
                rowspan, colspan = final_merge_map[(row_idx, col_idx)]
                for r in range(row_idx, row_idx + rowspan):
                    for c in range(col_idx, col_idx + colspan):
                        if (r, c) != (row_idx, col_idx):
                            skip_cells.add((r, c))

            html += f'    <td rowspan="{rowspan}" colspan="{colspan}">{value if value else ""}</td>'
        html += '  </tr>'

    html += '</table>'

    return html


def excel_to_html(file_path: str) -> str:
    try:
        if file_path.endswith('.xlsx'):
            return xlsx_to_html(file_path)
        elif file_path.endswith('.xls'):
            return xls_to_html(file_path)
    except  Exception as e:
        print(f"Error excel_to_html {file_path}: {e}")


def extract_images_from_xlsx(file_path: str):
    """
    从xlsx文件中提取图片并保存为png文件
    :param file_path: xlsx文件路径
    :return:
    """
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(file_path)

    # 获取文件名（不带扩展名）
    base_name = os.path.splitext(os.path.basename(file_path))[0]

    # 获取文件所在的目录
    directory = os.path.dirname(file_path)

    # 用于计数图片的变量
    image_count = 1

    # 遍历所有工作表
    for sheet in wb:
        # 遍历工作表中的所有图片
        for image in sheet._images:
            # 获取图片数据
            image_data = image._data()

            # 生成新的文件名，例如：xxx1.png, xxx2.png
            image_filename = f"{base_name}_{image_count}.png"
            image_path = os.path.join(directory, image_filename)

            # 将图片数据写入文件
            with open(image_path, 'wb') as img_file:
                img_file.write(image_data)

            print(f"Saved image: {image_path}")
            image_count += 1


def extract_images_from_excel(file_path: str):
    try:
        if file_path.endswith('.xlsx'):
            extract_images_from_xlsx(file_path)
    except  Exception as e:
        print(f"Error extract_images_from_excel {file_path}: {e}")
