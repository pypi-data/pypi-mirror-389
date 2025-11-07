from copy import copy
import openpyxl
import os
import csv
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.utils import range_boundaries, get_column_letter

class WorkbookManager(object):
    def copy_sheet(self, src_ws, tgt_ws):
        """
        Полностью копирует лист Excel с сохранением:
        - данных и формул
        - стилей и форматирования
        - скрытия колонок/строк
        - комментариев и гиперссылок
        - объединенных ячеек
        - параметров страницы
        """
        # Копируем параметры колонок (включая скрытие)
        for col_letter, col_dim in src_ws.column_dimensions.items():
            tgt_col = tgt_ws.column_dimensions[col_letter]
            tgt_col.width = col_dim.width
            tgt_col.hidden = col_dim.hidden  # параметр скрытия колонки
            tgt_col.bestFit = col_dim.bestFit
            tgt_col.outline_level = col_dim.outline_level
        
        # Копируем параметры строк (включая скрытие)
        for row_num, row_dim in src_ws.row_dimensions.items():
            tgt_row = tgt_ws.row_dimensions[row_num]
            tgt_row.height = row_dim.height
            tgt_row.hidden = row_dim.hidden  # параметр скрытия строки
            tgt_row.outline_level = row_dim.outline_level
        
        # Копируем данные и стили ячеек
        for row in src_ws.iter_rows():
            for src_cell in row:
                tgt_cell = tgt_ws.cell(
                    row=src_cell.row,
                    column=src_cell.column,
                    value=src_cell.value
                )
                
                # Копируем стиль
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.number_format = copy(src_cell.number_format)
                    tgt_cell.protection = copy(src_cell.protection)
                    tgt_cell.alignment = copy(src_cell.alignment)
                
                # Копируем комментарии и гиперссылки
                if src_cell.comment:
                    tgt_cell.comment = copy(src_cell.comment)
                if src_cell.hyperlink:
                    tgt_cell.hyperlink = copy(src_cell.hyperlink)
        
        # Копируем объединенные ячейки
        for merged_range in src_ws.merged_cells.ranges:
            tgt_ws.merge_cells(str(merged_range))
        
        # Копируем параметры страницы
        tgt_ws.page_margins = copy(src_ws.page_margins)
        tgt_ws.page_setup = copy(src_ws.page_setup)
        tgt_ws.print_options = copy(src_ws.print_options)
        tgt_ws.sheet_properties = copy(src_ws.sheet_properties)
        
        # Копируем фильтры
        if src_ws.auto_filter:
            tgt_ws.auto_filter.ref = src_ws.auto_filter.ref

        # Копируем параметры группировки (outline)
        tgt_ws.sheet_properties.outlinePr = copy(src_ws.sheet_properties.outlinePr)

    def copy_named_ranges(self, src_wb, tgt_wb):
        # Копирование каждого именованного диапазона
        for name, named_range in src_wb.defined_names.items():
            # Создаем новый именованный диапазон в целевой книге
            tgt_wb.defined_names[name] = named_range
            # print(f'{name}')
        
    def transfer_table_definitions(self, src_wb, tgt_wb):
        """
        Переносит определения таблиц без данных
        """
        
        # Создаем листы-заглушки при необходимости
        for sheetname in src_wb.sheetnames:
            if sheetname not in tgt_wb.sheetnames:
                tgt_wb.create_sheet(sheetname)
        
        # Переносим определения таблиц
        for src_sheet in src_wb.worksheets:
            tgt_sheet = tgt_wb[src_sheet.title]
            
            for table in src_sheet.tables.values():
                try:
                    new_table = Table(
                        displayName=table.displayName,
                        ref=table.ref,
                        tableStyleInfo=table.tableStyleInfo,
                        tableColumns=table.tableColumns,
                        autoFilter=table.autoFilter
                    )
                    tgt_sheet.add_table(new_table)
                except Exception as e:
                    print(f"Ошибка: {str(e)}")


    
    def copy_validation_rules(self, src_sheet, tgt_sheet, src_range, tgt_range):
        """
        Копирует правила валидации из исходного диапазона в целевой диапазон.
        
        Параметры:
        - src_sheet: Исходный лист, откуда копируются правила валидации
        - tgt_sheet: Целевой лист, куда копируются правила валидации
        - src_range: Диапазон исходных ячеек (например, "A1:C1")
        - tgt_range: Диапазон целевых ячеек (например, "A1:C1000")
        """
        # Получаем границы диапазонов
        src_min_col, src_min_row, src_max_col, src_max_row = range_boundaries(src_range)
        tgt_min_col, tgt_min_row, tgt_max_col, tgt_max_row = range_boundaries(tgt_range)
        
        # Проверяем, что количество столбцов совпадает
        src_cols = src_max_col - src_min_col + 1
        tgt_cols = tgt_max_col - tgt_min_col + 1
        if src_cols != tgt_cols:
            raise ValueError("Количество столбцов в исходном и целевом диапазонах должно совпадать")
        
        # Копируем валидацию для каждого столбца
        for col_offset in range(src_cols):
            src_col = src_min_col + col_offset
            tgt_col = tgt_min_col + col_offset
            
            # Ищем валидацию в исходном столбце
            validation = None
            for dv in src_sheet.data_validations.dataValidation:
                if dv.sqref and f"{get_column_letter(src_col)}{src_min_row}" in dv.sqref:
                    validation = dv
                    break
            
            if validation:
                # Создаем новую валидацию на основе исходной
                new_dv = DataValidation(
                    type=validation.type,
                    formula1=validation.formula1,
                    formula2=validation.formula2,
                    allow_blank=validation.allow_blank,
                    showErrorMessage=validation.showErrorMessage,
                    showInputMessage=validation.showInputMessage,
                    errorTitle=validation.errorTitle,
                    error=validation.error,
                    promptTitle=validation.promptTitle,
                    prompt=validation.prompt
                )
                
                # Определяем диапазон для новой валидации
                tgt_col_letter = get_column_letter(tgt_col)
                new_range = f"{tgt_col_letter}{tgt_min_row}:{tgt_col_letter}{tgt_max_row}"
                new_dv.add(new_range)
                
                # Добавляем валидацию в целевой лист
                tgt_sheet.add_data_validation(new_dv)

    # Вспомогательная функция для преобразования номера столбца в букву
    def get_column_letter(idx):
        """Преобразует номер столбца (начиная с 1) в буквенное обозначение (A, B, ..., Z, AA, AB, ...)"""
        letters = []
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            letters.append(chr(65 + remainder))
        return ''.join(reversed(letters))


    def write_to_range(self, sheet, excel_range, data):
        """
        Записать в указанный промежуток данные
        sheet - excel-лист
        excel_range - например A1:B2 - промежуток
        data - данные в виде list [[], []]
        """
        if len(data) == 0:
            return
        
        cells = sheet[excel_range]
        if isinstance(cells, Cell):
            cells.value = data[0][0]
            return

        # Проверка размеров
        if len(data) != len(cells) or len(data[0]) != len(cells[0]):
            raise ValueError(f"Несоответствие размеров данных и диапазона rows: {len(data)}/{len(cells)} columns: {len(data[0])}/{len(cells[0])}")
        
        # Запись
        for row_cells, row_data in zip(cells, data):
            for cell, value in zip(row_cells, row_data):
                cell.value = value

    def transform(self, meta_data, separator, output_path):
        targetWb = openpyxl.Workbook()   
        targetWb.remove(targetWb.active)
        
        for row in meta_data:
            if os.path.exists(row['template_name']):
                templateWb = openpyxl.load_workbook(row['template_name'])
                templateSheet = templateWb[row['sheet_name']]

            if row['sheet_name'] not in targetWb.sheetnames:
                targetSheet = targetWb.create_sheet(row['sheet_name'])
                self.copy_sheet(templateSheet, targetSheet)
                
            if os.path.exists(row['csv_data_name']) and row['mapping'].strip() != '':
                with open(row['csv_data_name'], mode='r', encoding='utf-8') as file:
                    reader = csv.reader(file, delimiter=separator)  # автоматически использует первую строку как ключи
                    data = list(reader)[1:]  # преобразуем в список словарей

                    first_row = CellRange(row['mapping'])

                    write_data_range = CellRange(row['mapping'])
                    write_data_range.expand(down=len(data) - 1)
                    
                    self.copy_validation_rules(
                        src_sheet=templateSheet, 
                        tgt_sheet=targetSheet, 
                        src_range=first_row.coord, 
                        tgt_range=write_data_range.coord
                    )

                    self.write_to_range(targetSheet, write_data_range.coord, data)


        self.copy_named_ranges(templateWb, targetWb)
        self.transfer_table_definitions(templateWb, targetWb)

        targetWb.save(output_path)
        return targetWb
