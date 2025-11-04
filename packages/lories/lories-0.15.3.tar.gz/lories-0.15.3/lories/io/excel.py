# -*- coding: utf-8 -*-
"""
lories.io.excel
~~~~~~~~~~~~~~~


"""

import logging
import os
import warnings
from typing import Optional

from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from tables import NaturalNameWarning

import pandas as pd

warnings.filterwarnings("ignore", category=NaturalNameWarning)
logger = logging.getLogger(__name__)


# noinspection PyShadowingBuiltins
def write(
    path: str,
    sheet_name: str,
    data: pd.DataFrame,
    index: Optional[bool] = None,
) -> None:
    if data.dropna(axis="index", how="all").dropna(axis="columns", how="all").empty:
        return None
    if isinstance(data.index, pd.DatetimeIndex):
        data.index = data.index.tz_localize(None)
    if data.index.name is not None:
        data.index.name = data.index.name.title()
    if index is None:
        index = len(data.index) > 1

    header_length = data.columns.nlevels

    border_side = Side(border_style=None)
    border = Border(
        left=border_side,
        right=border_side,
        bottom=border_side,
        top=border_side,
    )
    font = Font(name="Calibri Light", size=12, color="333333")

    if not path.endswith(".xlsx"):
        path += ".xlsx"

    dir = os.path.dirname(path)
    if not os.path.isdir(dir):
        os.makedirs(dir, exist_ok=True)

    kwargs = {
        "engine": "openpyxl",
    }
    if not os.path.exists(path):
        kwargs["mode"] = "w"
    else:
        kwargs["mode"] = "r+"
        kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(path, **kwargs) as writer:
        data.to_excel(writer, sheet_name=sheet_name, float_format="%.2f")  # , encoding=encoding)
        data_book = writer.book

        data_sheet = data_book[sheet_name]
        if header_length > 1:
            data_sheet[2][0].value = data_sheet[3][0].value
            data_sheet.delete_rows(3, 1)

        if not index:
            data_sheet.delete_cols(1)
            for data_merged in data_sheet.merged_cells:
                data_merged.shift(col_shift=-1)

        for column in range(1, len(data_sheet[header_length]) + 1):
            column_letter = get_column_letter(column)
            if column == 1 and index:
                column_length = len(data_sheet[column_letter]) + 1
                column_widths = [len(str(i)) for i in data.index.values]
            else:
                column_length = header_length
                column_widths = [len(str(i)) for i in data.iloc[:, column - int(index) - 1].values]

            for column_cell, *_ in data_sheet.iter_rows(
                min_row=1,
                max_row=column_length,
                min_col=column,
                max_col=column,
            ):
                column_cell.border = border
                column_cell.font = font

                if column_cell.row <= header_length:
                    header_cell_str = str(column_cell.value)
                    for header_cell_line in header_cell_str.split("\n"):
                        header_column_width = len(header_cell_line)
                        if header_cell_str in data.columns:
                            header_column_width /= len(data[[header_cell_str]].columns)
                        column_widths.append(header_column_width)

                    if "\n" in str(column_cell.value):
                        # header_alignment = copy(column_cell.alignment)
                        # header_alignment.wrapText = True
                        # header_alignment.vertical = 'center'
                        # header_alignment.horizontal = 'center'
                        # column_cell.alignment = header_alignment
                        column_cell.alignment.wrapText = True
                        data_sheet.row_dimensions[column_cell.row].height = 33

            data_sheet.column_dimensions[column_letter].width = max(column_widths) + 2
