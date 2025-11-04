'''
uses py 3.10 w/ pyqt6
'''

import os, sys
from PyQt6.QtWidgets import *
from ... import monkeyUnityv1_8 as mky
from ...utils.workers import RunCalibrationWorker
import openpyxl, shutil, filecmp
import subprocess
from pathlib import Path

# import vid_play_v0_7 as vid_player # save for later.
script_name_vid_player = 'vid_play_v0.71.py'

class TabTool(QWidget):
    def __init__(self, log_area:QTextEdit):
        super().__init__()
        self.initUI()
        self.setupConnections()
        self.log_area = log_area
        
    def initUI(self):
        # l = QVBoxLayout()
        lt = QFormLayout(self)
        # grp = QGroupBox()
        self.edt_xlsx_path = QLineEdit()
        self.btn_xlsx_browse = QPushButton('Browse')
        ltt = QHBoxLayout()
        ltt.addWidget(self.edt_xlsx_path)
        ltt.addWidget(self.btn_xlsx_browse)
        lt.addRow("Exp note path:", ltt)
        self.btn_xlsx_fill = QPushButton('Fill file names')
        lt.addRow(self.btn_xlsx_fill)
        self.btn_vid_player = QPushButton('Event Marker / Vid Player')
        lt.addRow(self.btn_vid_player)
        self.btn_calib = QPushButton('Run calibration')
        lt.addRow(self.btn_calib)
        self.btn_csv = QPushButton('Collect CSVs')
        lt.addRow(self.btn_csv)
        self.btn_fullauto = QPushButton('Full Auto - Fire at Will!')
        self.btn_fullauto.setEnabled(False)
        lt.addRow(self.btn_fullauto)
    
    def setupConnections(self):
        self.btn_xlsx_fill.clicked.connect(self.fillExpNote)
        self.btn_vid_player.clicked.connect(self.runVidPlayer)
        self.btn_calib.clicked.connect(self.btnCalibClicked)
        self.btn_csv.clicked.connect(self.collectCSV)
        #self.btn_fullauto.clicked.connect(self.fullAuto)
        
    def fillExpNote(self):
        return
        path = self.edt_xlsx_path.text()
        wb = openpyxl.load_workbook(path)
        wb.save(os.path.join(os.path.dirname(path), f'originalCopy_{os.path.basename(path)}'))
        if wb:
            ws = wb.active
            hr = None
            for row in ws.iter_rows():
                if any(cell.value == self.header_key.text() for cell in row):
                    hr = row[0].row
                break
            if not hr:
                print(f'No header found with key {self.edt_xlsx_path.text()}')
                return
            
            camnum_idx = {}
            for c in ws[hr]:
                if c.value in self.cam_header:
                    camnum_idx[c.value] = c.column
            
            for r in ws.iter_rows(min_row = hr + 1, max_row= ws.max_row):
                for cname, cidx in camnum_idx.items():
                    cell = r[cidx - 1]
                    v = cell.value
                    if v in mky.list_skipped_file_name: # sign for actually no file
                        continue
                    if v is None:
                        row = cell.row
                        while row > hr:
                            row -= 1
                            acell = ws.cell(row = row, column = cidx)
                            if acell.value is not None:
                                ahcell = ws.cell(row = row, column = 0)
                                hcell = ws.cell(row = cell.row, column = 0)
                                r[cidx - 1].value = acell.value + (hcell.value - ahcell.value)
                                break
                    # r[cidx - 1].value = 'x'
                print(r)
            
        wb.save(os.path.join(os.path.dirname(path), f'{os.path.basename(path)}'))
        wb.close()
    
    def btnCalibClicked(self):
        print('Trying to run calibration...')
        if len(mky.pm.calib_idx) != 0:
            vid_path_ = []
            folder_name = []
            for i, v in enumerate(mky.pm.vid_path):
                if i in mky.pm.calib_idx:
                    path = Path(v)
                    vid_path_.append(str(path))
                    folder_name.append(path.name)
            if len(vid_path_)>0:
                print('Copying videos; window may freeze!')
                self.sendToCalib(vid_path_=vid_path_, folder_name=folder_name)
        else:
            print('No calib found in current task set.')

    def sendToCalib(self, vid_path_: list, folder_name=['Calib']):
        '''
        if a calibration is found recorded, set up for calib and run it
        '''
        for vp, f in zip(vid_path_, folder_name):
            fname = os.path.join(mky.pm.ani_base_path, f, 'calibration')
            os.makedirs(fname, exist_ok=True)

            calib_file = os.path.join(fname, 'calibration.toml')
            if os.path.exists(calib_file):
                os.remove(calib_file)
                print(f'Removed existing calib: {calib_file}')

            try:
                vp_ = str.replace(vp, '//share.files.pitt.edu/RnelShare', 'P:') # RNEL-specific!! should be cleaned
                for v in Path(vp_).iterdir():
                    if mky.vid_type in v.name:
                        shutil.copy(v, Path(fname))
            except Exception as e:
                print(vid_path_)
                print(f'Failed sending to calib for {os.path.basename(vp)}: {e}')
        shutil.copy(mky.ani_cfg_mothercopy, mky.pm.ani_base_path)
        self.run_calib_worker = RunCalibrationWorker(mky.pm.ani_base_path)
        self.run_calib_worker.log_signal.connect(self.log_area.append)
        self.run_calib_worker.finished.connect(self.run_calib_worker.deleteLater)
        self.run_calib_worker.start()

    def calibDone(self):
        # collect calibration.toml
        for dirpath, dirnames, filenames in os.walk(mky.pm.ani_base_path):
            if "calibration" in dirnames:
                calibration_path = os.path.join(dirpath, "calibration", "calibration.toml")
                if os.path.isfile(calibration_path):
                    parent_folder = os.path.basename(dirpath)
                    shutil.copy(calibration_path, os.path.join(r'C:\Users\mkrig\Documents\calib history', f'calibration-{parent_folder}.toml')) # FIXME change path
                    shutil.copy(calibration_path, r'C:\Users\mkrig\Documents')
        self.run_calib_worker.deleteLater()
        mky.collectCalib()

    def runVidPlayer(self):
        p = str(Path(__file__).resolve().parent.parent.parent)
        print(p)
        p = os.path.join(p, 'utils', script_name_vid_player)
        subprocess.Popen([sys.executable, p])
        self.log_area.append(f'Started standalone event marker.')

    def collectCSV(self):
        src = Path(mky.pm.ani_base_path)
        dst = Path(mky.pm.data_path) / 'clean'  #TODO clean hardcoded
        dst.mkdir(parents=True, exist_ok=True)

        for f in src.rglob('*.csv'):
            tgt = dst / f.name
            if tgt.exists():
                if filecmp.cmp(f, tgt, shallow=False):
                    print(f'Skipped existing CSV: {f.stem}')
                    continue
                stem, suffix = f.stem, f.suffix
                i = 1
                while True:     # keep all the versions and rename new ones
                    tgt_new = dst / f"{stem}_{i}{suffix}"
                    if not tgt_new.exists():
                        tgt = tgt_new
                        break
                    i += 1

            shutil.copy(f, tgt)
            print(f'Copied CSV: {f.stem}')

        else:
            print('No CSV found. Nothing is copied.')

    def fullAuto(self):
        try:
            pass
            
        except Exception as e:
            self.log_area.append(f"Error: {str(e)}")
            QMessageBox.critical(self, "Error", str(e))
