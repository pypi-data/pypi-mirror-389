
from __future__ import annotations
import os
import re
import json
import uuid
from typing import Dict, Any, List, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

WORKBOOKS: Dict[str, str] = {}  # workbook_id -> file path

def ensure_workspace(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path

def register_workbook(path: str) -> str:
    """Register a workbook path and return a workbook_id."""
    if not os.path.isfile(path):
        # allow relative to cwd; else treat as in workspace
        cwd_path = os.path.join(os.getcwd(), path)
        if os.path.isfile(cwd_path):
            path = cwd_path
        else:
            # try workspace/ prefix
            ws_path = os.path.join(os.getcwd(), "workspace", path)
            if os.path.isfile(ws_path):
                path = ws_path
            else:
                raise FileNotFoundError(f"Workbook not found: {path}")
    wk_id = f"wk_{uuid.uuid4().hex[:8]}"
    WORKBOOKS[wk_id] = path
    return wk_id

def openpyxl_sheet_names(path: str) -> List[str]:
    wb = load_workbook(path)
    return wb.sheetnames

def sheet_to_model(path: str, sheet_name: str) -> Dict[str, Any]:
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    df = df.fillna("")
    cols = [str(c).strip() for c in df.columns]
    rows = []
    for idx, row in df.iterrows():
        r = {str(c): (str(row[c]).strip()) for c in df.columns}
        r["row_idx"] = int(idx) + 2  # account for header
        rows.append(r)
    hints = {
        "hasEvidenceColumn": any(c.lower() == "evidence" for c in cols),
        "hasDecisionColumn": any(c.lower() == "decision" for c in cols),
    }
    return {"columns": cols, "rows": rows, "hints": hints}

def guess_relevant_rules(topic_text: str, all_rules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    t = topic_text.lower()
    candidates = []
    for rule in all_rules:
        key = (rule.get("clause","") + " " + rule.get("text","")).lower()
        if any(k in t for k in ["9.3", "10.2", "incident", "nonconformity", "corrective", "management review", "outputs"]):
            candidates.append(rule)
    return candidates or all_rules

def classify_gaps(row: Dict[str, Any], rule: Dict[str, Any]) -> Tuple[bool, List[str], List[Dict[str, Any]], str]:
    gaps = []
    recs = []
    decision = row.get("Decision","").strip()
    evidence = row.get("Evidence","").strip()
    req_out = rule.get("required_outputs", [])
    conforms = True

    # Evidence/Decision checks
    if "Decision" in req_out and not decision:
        conforms = False
        gaps.append("Decision missing")
        recs.append({"type":"set_value", "column":"Decision",
                     "value":"<ADD MANAGEMENT DECISION HERE>"})
    if "Evidence" in req_out and not evidence:
        conforms = False
        gaps.append("Evidence missing")
        recs.append({"type":"set_value", "column":"Evidence",
                     "value":"<ATTACH EVIDENCE: logs, CAPA IDs, QMS changes>"})

    # Severity: Major if all required outputs missing, else Minor
    severity = "Minor"
    if len(req_out) >= 2 and (not decision and not evidence):
        severity = "Major"

    # Structural recommendations
    cols = set(k.lower() for k in row.keys())
    if "evidence" not in cols:
        recs.insert(0, {"type":"add_column_if_missing","column":"Evidence"})
        conforms = False
        gaps.append("Evidence column missing")
    if "decision" not in cols:
        recs.insert(0, {"type":"add_column_if_missing","column":"Decision"})
        conforms = False
        gaps.append("Decision column missing")

    return conforms, gaps, recs, severity

def apply_patches(path: str, sheet: str, patches: List[Dict[str, Any]]) -> str:
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in workbook.")
    ws = wb[sheet]

    # Add column helper
    def add_column_if_missing(col_name: str):
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1)]
        if col_name not in headers:
            ws.insert_cols(ws.max_column + 1)
            ws.cell(row=1, column=ws.max_column).value = col_name
        # return its index
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1)]
        return headers.index(col_name) + 1

    # Find column index by name (case-sensitive header match)
    def col_index(col_name: str) -> int:
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1)]
        if col_name in headers:
            return headers.index(col_name) + 1
        # lazily add if truly missing
        return add_column_if_missing(col_name)

    for p in patches:
        op = p.get("op") or p.get("type")
        if op == "add_column_if_missing":
            add_column_if_missing(p["column"])
        elif op == "set_value":
            row_idx = int(p["row_idx"])
            col = col_index(p["column"])
            ws.cell(row=row_idx, column=col).value = p.get("value","")
        elif op == "conditional_format":
            # simple fill highlight
            rng = p.get("range")
            if rng:
                yellow = PatternFill(start_color="FFF8C6", end_color="FFF8C6", fill_type="solid")
                bold = Font(bold=True)
                for row in ws[rng]:
                    for cell in row:
                        cell.fill = yellow
                        cell.font = bold
        else:
            # ignore unknown ops
            pass

    out_path = os.path.join(os.path.dirname(path), "MoM_patched.xlsx")
    wb.save(out_path)
    return out_path
