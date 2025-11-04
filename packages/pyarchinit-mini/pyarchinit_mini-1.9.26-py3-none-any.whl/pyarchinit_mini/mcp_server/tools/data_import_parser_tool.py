"""
Data Import Parser Tool

Automatic parser for importing archaeological data from various formats:
- CSV, Excel (XLS, XLSX), JSON, XML formats
- Auto-detect field mappings from column headers
- Validate data before import
- Support batch operations
- Handle errors gracefully
- Provide detailed import statistics and logs
"""

import logging
import os
import json
import csv
import base64
import tempfile
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
from .base_tool import BaseTool, ToolDescription

logger = logging.getLogger(__name__)


class DataImportParserTool(BaseTool):
    """Automatic data import parser with format detection and field mapping"""

    # Field mapping dictionary for common headers
    FIELD_MAPPINGS = {
        # Site table mappings
        "site_table": {
            "sito": ["sito", "site", "site_name", "nome_sito", "sitio"],
            "nazione": ["nazione", "country", "nation", "pais", "paese"],
            "regione": ["regione", "region", "regiao", "estado"],
            "comune": ["comune", "municipality", "municipio", "city", "città"],
            "provincia": ["provincia", "province", "state"],
            "descrizione": ["descrizione", "description", "descripcion", "desc"],
            "definizione_sito": ["definizione_sito", "site_definition", "tipo_sito", "site_type"],
        },
        # US table mappings
        "us_table": {
            "sito": ["sito", "site", "site_name"],
            "area": ["area", "settore", "sector"],
            "us": ["us", "su", "stratigraphic_unit", "unit"],
            "unita_tipo": ["unita_tipo", "unit_type", "tipo", "type"],
            "d_stratigrafica": ["d_stratigrafica", "stratigraphic_desc", "descrizione_stratigrafica"],
            "d_interpretativa": ["d_interpretativa", "interpretation", "interpretazione"],
            "rapporti": ["rapporti", "relationships", "relaciones", "relations"],
        },
        # Inventario table mappings
        "inventario_materiali_table": {
            "sito": ["sito", "site", "site_name"],
            "area": ["area", "settore", "sector"],
            "us": ["us", "su", "unit"],
            "numero_inventario": ["numero_inventario", "inventory_number", "num_inv", "n_inv"],
            "tipo_reperto": ["tipo_reperto", "artifact_type", "tipo", "type"],
            "materiale": ["materiale", "material", "materia"],
            "descrizione": ["descrizione", "description", "desc"],
        }
    }

    def to_tool_description(self) -> ToolDescription:
        return ToolDescription(
            name="import_data",
            description=(
                "✅ GENERIC DATA IMPORT PARSER - Use this tool for generic Excel/CSV/JSON/XML files with AUTO field mapping.\n\n"
                "📋 **Use this tool for**:\n"
                "   • Generic Excel/CSV files with site data (columns: sito, nazione, comune, etc.)\n"
                "   • Generic US data files (columns: sito, area, us, unita_tipo, descrizione)\n"
                "   • Inventory/finds lists (columns: numero_inventario, tipo_reperto, materiale)\n"
                "   • Any archaeological data file where columns can be auto-mapped to database fields\n\n"
                "❌ **DO NOT USE for**:\n"
                "   • Harris Matrix Template Excel (multi-sheet with NODES/RELATIONSHIPS) → Use 'import_excel' tool with format='harris_template'\n"
                "   • Extended Matrix Excel (single sheet with relationship columns: is_before, covers, etc.) → Use 'import_excel' tool with format='extended_matrix'\n"
                "   • Metro C excavation files → Use 'import_excel' tool with format='extended_matrix'\n\n"
                "🔄 **This tool automatically**:\n"
                "   1. Detects file format (CSV, Excel, JSON, XML)\n"
                "   2. Auto-maps column headers to database fields (e.g., 'site' → 'sito', 'description' → 'descrizione')\n"
                "   3. Validates data before import\n"
                "   4. Imports records into appropriate table (site_table, us_table, inventario_materiali_table)\n"
                "   5. Provides detailed statistics and error reports\n\n"
                "💡 **Supported operations**:\n"
                "   • 'parse' = Analyze file structure and detect mappings\n"
                "   • 'validate' = Check data validity without importing\n"
                "   • 'preview' = Preview first N rows with detected mappings\n"
                "   • 'analyze' = Get file statistics and column analysis\n"
                "   • 'import' = Parse and import data into database\n"
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "operation": {
                        "type": "string",
                        "enum": ["parse", "import", "validate", "preview", "analyze"],
                        "description": (
                            "'parse' = Parse file and detect structure/mappings, "
                            "'import' = Parse and import data into database, "
                            "'validate' = Validate data without importing, "
                            "'preview' = Preview first N rows with detected mappings, "
                            "'analyze' = Analyze file structure and statistics"
                        )
                    },
                    "file_path": {
                        "type": "string",
                        "description": "Path to import file on server"
                    },
                    "file_content_base64": {
                        "type": "string",
                        "description": "Base64-encoded file content (alternative to file_path)"
                    },
                    "filename": {
                        "type": "string",
                        "description": "Filename (required with file_content_base64)"
                    },
                    "target_table": {
                        "type": "string",
                        "enum": ["site_table", "us_table", "inventario_materiali_table", "auto"],
                        "description": "Target database table ('auto' for auto-detection)",
                        "default": "auto"
                    },
                    "format": {
                        "type": "string",
                        "enum": ["csv", "excel", "json", "xml", "auto"],
                        "description": "File format ('auto' for auto-detection)",
                        "default": "auto"
                    },
                    "field_mappings": {
                        "type": "object",
                        "description": "Custom field mappings (source_field → target_field)"
                    },
                    "skip_errors": {
                        "type": "boolean",
                        "description": "Skip rows with errors and continue importing",
                        "default": False
                    },
                    "batch_size": {
                        "type": "integer",
                        "description": "Number of rows to import per batch",
                        "default": 100
                    },
                    "preview_rows": {
                        "type": "integer",
                        "description": "Number of rows to preview",
                        "default": 5
                    },
                    "encoding": {
                        "type": "string",
                        "description": "File encoding (default: utf-8)",
                        "default": "utf-8"
                    },
                    "delimiter": {
                        "type": "string",
                        "description": "CSV delimiter (default: ',')",
                        "default": ","
                    }
                },
                "required": ["operation"],
            },
        )

    async def execute(self, arguments: Dict[str, Any]) -> Dict[str, Any]:
        """Execute data import parser operation"""
        try:
            operation = arguments.get("operation")

            logger.info(f"Executing data import operation: {operation}")

            # Get file path or decode base64 content
            file_path = await self._get_file_path(arguments)
            if not file_path:
                return self._format_error("file_path or file_content_base64 is required")

            try:
                if operation == "parse":
                    return await self._handle_parse(file_path, arguments)
                elif operation == "import":
                    return await self._handle_import(file_path, arguments)
                elif operation == "validate":
                    return await self._handle_validate(file_path, arguments)
                elif operation == "preview":
                    return await self._handle_preview(file_path, arguments)
                elif operation == "analyze":
                    return await self._handle_analyze(file_path, arguments)
                else:
                    return self._format_error(f"Unknown operation: {operation}")

            finally:
                # Clean up temp file if created
                if arguments.get("file_content_base64") and os.path.exists(file_path):
                    try:
                        os.unlink(file_path)
                    except:
                        pass

        except Exception as e:
            logger.error(f"Data import error: {str(e)}", exc_info=True)
            return self._format_error(f"Data import failed: {str(e)}")

    async def _get_file_path(self, arguments: Dict[str, Any]) -> Optional[str]:
        """Get file path from arguments or decode base64 content"""
        file_path = arguments.get("file_path")
        file_content_base64 = arguments.get("file_content_base64")
        filename = arguments.get("filename")

        if file_path and os.path.exists(file_path):
            return file_path

        if file_content_base64:
            if not filename:
                raise ValueError("filename is required when using file_content_base64")

            # Decode and save to temp file
            file_content = base64.b64decode(file_content_base64)
            temp_file = tempfile.NamedTemporaryFile(
                delete=False,
                suffix=Path(filename).suffix
            )
            temp_file.write(file_content)
            temp_file.close()
            return temp_file.name

        return None

    async def _handle_parse(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Handle parse operation"""
        # Detect format
        file_format = await self._detect_format(file_path, arguments.get("format", "auto"))

        # Parse file
        data, headers = await self._parse_file(file_path, file_format, arguments)

        # Detect target table
        target_table = await self._detect_target_table(
            headers,
            arguments.get("target_table", "auto")
        )

        # Detect field mappings
        field_mappings = await self._detect_field_mappings(
            headers,
            target_table,
            arguments.get("field_mappings")
        )

        return self._format_success(
            result={
                "file_format": file_format,
                "target_table": target_table,
                "total_rows": len(data),
                "headers": headers,
                "field_mappings": field_mappings,
                "sample_row": data[0] if data else None
            },
            message=f"Parsed {len(data)} rows from {file_format} file"
        )

    async def _handle_import(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Handle import operation"""
        # Parse file
        file_format = await self._detect_format(file_path, arguments.get("format", "auto"))
        data, headers = await self._parse_file(file_path, file_format, arguments)

        # Detect target table
        target_table = await self._detect_target_table(
            headers,
            arguments.get("target_table", "auto")
        )

        # Detect field mappings
        field_mappings = await self._detect_field_mappings(
            headers,
            target_table,
            arguments.get("field_mappings")
        )

        # Import data
        stats = await self._import_data(
            data,
            target_table,
            field_mappings,
            arguments
        )

        return self._format_success(
            result={
                "target_table": target_table,
                "statistics": stats
            },
            message=f"Imported {stats['imported']} of {stats['total']} rows into {target_table}"
        )

    async def _handle_validate(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Handle validate operation"""
        # Parse file
        file_format = await self._detect_format(file_path, arguments.get("format", "auto"))
        data, headers = await self._parse_file(file_path, file_format, arguments)

        # Detect target table and mappings
        target_table = await self._detect_target_table(
            headers,
            arguments.get("target_table", "auto")
        )
        field_mappings = await self._detect_field_mappings(
            headers,
            target_table,
            arguments.get("field_mappings")
        )

        # Validate data
        validation_results = await self._validate_data(
            data,
            target_table,
            field_mappings
        )

        return self._format_success(
            result={
                "target_table": target_table,
                "total_rows": len(data),
                "valid_rows": validation_results["valid_count"],
                "invalid_rows": validation_results["invalid_count"],
                "errors": validation_results["errors"][:50]  # First 50 errors
            },
            message=f"Validated {len(data)} rows: {validation_results['valid_count']} valid, {validation_results['invalid_count']} invalid"
        )

    async def _handle_preview(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Handle preview operation"""
        # Parse file
        file_format = await self._detect_format(file_path, arguments.get("format", "auto"))
        data, headers = await self._parse_file(file_path, file_format, arguments)

        # Detect target table and mappings
        target_table = await self._detect_target_table(
            headers,
            arguments.get("target_table", "auto")
        )
        field_mappings = await self._detect_field_mappings(
            headers,
            target_table,
            arguments.get("field_mappings")
        )

        # Get preview rows
        preview_rows = arguments.get("preview_rows", 5)
        preview_data = data[:preview_rows]

        return self._format_success(
            result={
                "file_format": file_format,
                "target_table": target_table,
                "total_rows": len(data),
                "headers": headers,
                "field_mappings": field_mappings,
                "preview_rows": preview_data
            },
            message=f"Preview of first {len(preview_data)} rows"
        )

    async def _handle_analyze(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Handle analyze operation"""
        # Get file info
        file_size = os.path.getsize(file_path)
        file_format = await self._detect_format(file_path, arguments.get("format", "auto"))

        # Parse file
        data, headers = await self._parse_file(file_path, file_format, arguments)

        # Analyze data
        analysis = {
            "file_size_mb": round(file_size / (1024 * 1024), 2),
            "file_format": file_format,
            "total_rows": len(data),
            "total_columns": len(headers),
            "headers": headers,
            "column_analysis": {}
        }

        # Analyze each column
        for header in headers:
            values = [row.get(header) for row in data if row.get(header)]
            analysis["column_analysis"][header] = {
                "non_empty_count": len(values),
                "empty_count": len(data) - len(values),
                "unique_values": len(set(values)),
                "sample_values": list(set(values))[:5]
            }

        return self._format_success(
            result=analysis,
            message=f"Analyzed {len(data)} rows with {len(headers)} columns"
        )

    async def _detect_format(self, file_path: str, format_hint: str) -> str:
        """Detect file format"""
        if format_hint != "auto":
            return format_hint

        ext = Path(file_path).suffix.lower()
        if ext == ".csv":
            return "csv"
        elif ext in [".xls", ".xlsx"]:
            return "excel"
        elif ext == ".json":
            return "json"
        elif ext == ".xml":
            return "xml"
        else:
            return "csv"  # Default

    async def _parse_file(
        self,
        file_path: str,
        file_format: str,
        arguments: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse file and return data with headers"""
        if file_format == "csv":
            return await self._parse_csv(file_path, arguments)
        elif file_format == "excel":
            return await self._parse_excel(file_path, arguments)
        elif file_format == "json":
            return await self._parse_json(file_path, arguments)
        elif file_format == "xml":
            return await self._parse_xml(file_path, arguments)
        else:
            raise ValueError(f"Unsupported format: {file_format}")

    async def _parse_csv(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse CSV file"""
        encoding = arguments.get("encoding", "utf-8")
        delimiter = arguments.get("delimiter", ",")

        data = []
        headers = []

        with open(file_path, "r", encoding=encoding) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            for row in reader:
                data.append(dict(row))

        return data, headers

    async def _parse_excel(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse Excel file"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Get data
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            return data, headers

        except ImportError:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

    async def _parse_json(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse JSON file"""
        with open(file_path, "r") as f:
            json_data = json.load(f)

        # Handle different JSON structures
        if isinstance(json_data, list):
            data = json_data
        elif isinstance(json_data, dict) and "data" in json_data:
            data = json_data["data"]
        else:
            data = [json_data]

        # Get headers from first row
        headers = list(data[0].keys()) if data else []

        return data, headers

    async def _parse_xml(
        self,
        file_path: str,
        arguments: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse XML file"""
        import xml.etree.ElementTree as ET

        tree = ET.parse(file_path)
        root = tree.getroot()

        data = []
        headers = set()

        # Extract data from XML
        for item in root:
            row_dict = {}
            for child in item:
                headers.add(child.tag)
                row_dict[child.tag] = child.text
            data.append(row_dict)

        return data, list(headers)

    async def _detect_target_table(
        self,
        headers: List[str],
        target_hint: str
    ) -> str:
        """Detect target table from headers"""
        if target_hint != "auto":
            return target_hint

        # Count matches for each table
        scores = {}
        for table_name, field_map in self.FIELD_MAPPINGS.items():
            score = 0
            for field, aliases in field_map.items():
                for header in headers:
                    if header.lower() in [alias.lower() for alias in aliases]:
                        score += 1
            scores[table_name] = score

        # Return table with highest score
        if scores:
            return max(scores, key=scores.get)
        return "us_table"  # Default

    async def _detect_field_mappings(
        self,
        headers: List[str],
        target_table: str,
        custom_mappings: Optional[Dict[str, str]]
    ) -> Dict[str, str]:
        """Detect field mappings"""
        if custom_mappings:
            return custom_mappings

        mappings = {}
        field_map = self.FIELD_MAPPINGS.get(target_table, {})

        for header in headers:
            header_lower = header.lower()
            for field, aliases in field_map.items():
                if header_lower in [alias.lower() for alias in aliases]:
                    mappings[header] = field
                    break

        return mappings

    async def _validate_data(
        self,
        data: List[Dict[str, Any]],
        target_table: str,
        field_mappings: Dict[str, str]
    ) -> Dict[str, Any]:
        """Validate data"""
        valid_count = 0
        invalid_count = 0
        errors = []

        for i, row in enumerate(data):
            try:
                # Map fields
                mapped_row = {
                    field_mappings.get(k, k): v
                    for k, v in row.items()
                    if k in field_mappings
                }

                # Basic validation
                if target_table == "site_table":
                    if not mapped_row.get("sito"):
                        raise ValueError("sito is required")
                elif target_table == "us_table":
                    if not mapped_row.get("sito") or not mapped_row.get("us"):
                        raise ValueError("sito and us are required")

                valid_count += 1

            except Exception as e:
                invalid_count += 1
                errors.append({
                    "row": i + 1,
                    "error": str(e),
                    "data": row
                })

        return {
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "errors": errors
        }

    async def _import_data(
        self,
        data: List[Dict[str, Any]],
        target_table: str,
        field_mappings: Dict[str, str],
        arguments: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Import data into database"""
        imported = 0
        failed = 0
        errors = []
        skip_errors = arguments.get("skip_errors", False)

        # Get appropriate service
        service = self._get_service_for_table(target_table)

        for i, row in enumerate(data):
            try:
                # Map fields
                mapped_row = {
                    field_mappings.get(k, k): v
                    for k, v in row.items()
                    if k in field_mappings
                }

                # Import row
                if target_table == "site_table":
                    service.create_site(mapped_row)
                elif target_table == "us_table":
                    service.create_us(mapped_row)
                elif target_table == "inventario_materiali_table":
                    service.create_inventario(mapped_row)

                imported += 1

            except Exception as e:
                failed += 1
                errors.append({
                    "row": i + 1,
                    "error": str(e)
                })

                if not skip_errors:
                    break

        return {
            "total": len(data),
            "imported": imported,
            "failed": failed,
            "errors": errors[:50]  # First 50 errors
        }

    def _get_service_for_table(self, table_name: str):
        """Get service instance for table"""
        if table_name == "site_table":
            from ...services.site_service import SiteService
            return SiteService(self.db_manager)
        elif table_name == "us_table":
            from ...services.us_service import USService
            return USService(self.db_manager)
        elif table_name == "inventario_materiali_table":
            from ...services.inventario_service import InventarioService
            return InventarioService(self.db_manager)
        else:
            raise ValueError(f"Unsupported table: {table_name}")
