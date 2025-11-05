"""
Excel reading functionality for Simplex Foundation projects.

This module contains all the functions needed to read Excel files and convert them
into Simplex Foundation Project objects.
"""

from typing import List
from openpyxl import load_workbook
import simplex.foundation.model.project_info
import simplex.foundation.model.foundation_loading
import simplex.foundation.model.reinforcement
import simplex.foundation.model.foundation
import simplex.foundation.model.settings
import simplex.foundation.model.soil
import simplex.foundation.model.project
import simplex.core.materials.materials
import simplex.core.geometry.point

def _read_project_info_from_excel(workbook) -> simplex.foundation.model.project_info.ProjectInfo:
    """
    Read project information from the 'Info' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        ProjectInfo: A ProjectInfo object populated with data from the Excel file
    """
    try:
        sheet = workbook['Info']
    except KeyError:
        try:
            sheet = workbook['info']
        except KeyError:
            raise ValueError("Excel file must contain an 'Info' or 'info' sheet")
    
    # Initialize ProjectInfo with default values
    project_info = simplex.foundation.model.project_info.ProjectInfo()
    
    # Read the data from specific cells based on the Excel template structure
    # Row 1 contains headers (Field, Value)
    # Row 2: Project field
    # Row 3: Name field
    # Row 4: Description field
    # Row 5: Location field
    # Row 6: Company field
    # Row 7: Signature field
    # Row 8: Comments field
    
    # Read values from specific cells (column 2 contains the values)
    project_info.project = sheet.cell(row=2, column=2).value or ""
    project_info.name = sheet.cell(row=3, column=2).value or ""
    project_info.description = sheet.cell(row=4, column=2).value or ""
    project_info.location = sheet.cell(row=5, column=2).value or ""
    project_info.company = sheet.cell(row=6, column=2).value or ""
    project_info.signature = sheet.cell(row=7, column=2).value or ""
    project_info.comments = sheet.cell(row=8, column=2).value or ""
    
    return project_info

def _read_foundation_loading_from_excel(workbook) -> List[simplex.foundation.model.foundation_loading.FoundationLoading]:
    """
    Read foundation loading from the 'Loading' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        List[FoundationLoading]: A list of FoundationLoading objects populated with data from the Excel file
    """
    try:
        # Get the 'Loading' sheet
        sheet = workbook['Loading']
    except KeyError:
        raise ValueError("Excel file must contain a 'Loading' sheet")
    
    # First, determine the foundation type to know which loading class to use
    foundation_type = _get_foundation_type(workbook)
    
    # Read the header row to understand the structure
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
    
    print(f"Found headers: {headers}")
    
    # Create dynamic header mapping based on foundation type
    if foundation_type.upper() == 'LINE':
        # Line foundation headers (per unit length)
        header_mapping = {
            'Name': 'Name',
            'Type': 'Type',
            'Hx [kN/m]': 'Hx',
            'Hy [kN/m]': 'Hy',
            'N [kN/m]': 'N',
            'Mx [kNm/m]': 'Mx',
            'My [kNm/m]': 'My',
            'SW [-]': 'SW'
        }
    else:
        # Point foundation headers (total forces)
        header_mapping = {
            'Name': 'Name',
            'Type': 'Type',
            'Hx [kN]': 'Hx',
            'Hy [kN]': 'Hy',
            'N [kN]': 'N',
            'Mx [kNm]': 'Mx',
            'My [kNm]': 'My',
            'SW [-]': 'SW'
        }
    
    # Read loading data starting from row 2
    loading_list = []
    for row in range(2, sheet.max_row + 1):
        # Check if the row has any data
        row_has_data = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                row_has_data = True
                break
        
        if not row_has_data:
            continue
        
        # Read the loading data
        loading_data = {}
        for col in range(1, min(len(headers) + 1, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                header_name = headers[col - 1]
                mapped_header = header_mapping.get(header_name, header_name)
                loading_data[mapped_header] = cell_value
        
        if not loading_data:
            continue
        
        # Create the appropriate loading object
        loading_obj = _create_loading_object(loading_data, foundation_type)
        if loading_obj:
            loading_list.append(loading_obj)
    
    return loading_list

def _get_foundation_type(workbook) -> str:
    """
    Determine the foundation type from the Foundation sheet.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        str: The foundation type ('Line' or 'Point')
    """
    try:
        sheet = workbook['Foundation']
        foundation_type = sheet['B2'].value
        if foundation_type:
            return str(foundation_type).strip()
        else:
            return "Point"  # Default to Point foundation
    except (KeyError, Exception):
        return "Point"  # Default to Point foundation

def _get_soil_type(workbook) -> str:
    """
    Determine the soil type from the Soil sheet.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        str: The soil type ('Complex' or 'Simple')
    """
    try:
        sheet = workbook['Soil']
        soil_type = sheet['B2'].value
        if soil_type:
            return str(soil_type).strip()
        else:
            return "Simple"  # Default to Simple soil
    except (KeyError, Exception):
        return "Simple"  # Default to Simple soil

def _create_loading_object(loading_data: dict, foundation_type: str) -> simplex.foundation.model.foundation_loading.FoundationLoading:
    """
    Create a FoundationLoading object based on the foundation type and loading data.
    
    Args:
        loading_data: Dictionary containing loading values
        foundation_type: Type of foundation ('Line' or 'Point')
        
    Returns:
        FoundationLoading: The appropriate loading object
    """
    from simplex.foundation.model.foundation_loading import PointFoundationLoading, LineFoundationLoading, LoadType
    
    # Parse the load type
    load_type_str = str(loading_data.get('Type', 'CHARACTERISTIC')).upper()
    load_type_map = {
        'CHARACTERISTIC': LoadType.CHARACTERISTIC,
        'FREQUENT': LoadType.FREQUENT,
        'QUASI_PERMANENT': LoadType.QUASI_PERMANENT,
        'ULTIMATE': LoadType.ULTIMATE,
        'ACCIDENTAL': LoadType.ACCIDENTAL,
        #'FIRE': LoadType.FIRE,
        'SEISMIC': LoadType.SEISMIC
    }
    load_type = load_type_map.get(load_type_str, LoadType.CHARACTERISTIC)
    
    # Get common values
    name = str(loading_data.get('Name', 'Loading'))
    hx = float(loading_data.get('Hx', 0))
    hy = float(loading_data.get('Hy', 0))
    n = float(loading_data.get('N', 0))
    sw = float(loading_data.get('SW', 1))
    
    if foundation_type.upper() == 'LINE':
        # For line foundation, we need My but not Mx
        my = float(loading_data.get('My', 0))
        return LineFoundationLoading(
            name=name,
            type=load_type,
            hx=hx,
            hy=hy,
            n=n,
            my=my,
            sw=sw
        )
    else:
        # For point foundation, we need both Mx and My
        mx = float(loading_data.get('Mx', 0))
        my = float(loading_data.get('My', 0))
        return PointFoundationLoading(
            name=name,
            type=load_type,
            hx=hx,
            hy=hy,
            n=n,
            mx=mx,
            my=my,
            sw=sw
        )

def _read_reinforcement_from_excel(workbook, foundation_type: str = None) -> simplex.foundation.model.reinforcement.Rebars:
    """
    Read reinforcement from the 'Reinforcement' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        foundation_type: Type of foundation ('Line' or 'Point') to determine reinforcement filtering
        
    Returns:
        Rebars: A Rebars object populated with data from the Excel file
    """
    try:
        # Get the 'Reinforcement' sheet
        sheet = workbook['Reinforcement']
    except KeyError:
        raise ValueError("Excel file must contain a 'Reinforcement' sheet")
    
    # Read the header row to understand the structure
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
    
    print(f"Found reinforcement headers: {headers}")
    
    # Group reinforcement layers by direction
    x_direction_layers = []
    y_direction_layers = []
    
    # Read reinforcement data starting from row 2
    for row in range(2, sheet.max_row + 1):
        # Check if the row has any data
        row_has_data = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                row_has_data = True
                break
        
        if not row_has_data:
            continue
        
        # Read the reinforcement data
        reinf_data = {}
        for col in range(1, min(len(headers) + 1, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                reinf_data[headers[col - 1]] = cell_value
        
        if not reinf_data:
            continue
        
        # Create the reinforcement layer
        reinf_layer = _create_reinf_layer(reinf_data)
        if reinf_layer:
            direction = str(reinf_data.get('Direction', '')).upper()
            if direction == 'X':
                x_direction_layers.append(reinf_layer)
            elif direction == 'Y':
                # For Line foundations, skip Y-direction reinforcement
                if foundation_type and foundation_type.upper() == 'LINE':
                    print(f"Warning: Skipping Y-direction reinforcement for Line foundation")
                    continue
                y_direction_layers.append(reinf_layer)
    
    # Create the Rebars object
    from simplex.foundation.model.reinforcement import Rebars
    
    rebars = Rebars(
        x_direction=x_direction_layers,
        y_direction=y_direction_layers if y_direction_layers else None
    )
    
    return rebars

def _create_reinf_layer(reinf_data: dict) -> simplex.foundation.model.reinforcement.ReinfLayer:
    """
    Create a ReinfLayer object from reinforcement data.
    
    Args:
        reinf_data: Dictionary containing reinforcement values
        
    Returns:
        ReinfLayer: The reinforcement layer object
    """
    from simplex.foundation.model.reinforcement import ReinfLayer, Zone
    from simplex.core.materials.materials import Reinforcement
    
    # Parse the zone
    zone_str = str(reinf_data.get('Zone', '')).upper()
    zone_map = {
        'START': Zone.START,
        'END': Zone.END,
        'BOTTOM': Zone.BOTTOM,
        'RIGHT': Zone.RIGHT,
        'TOP': Zone.TOP,
        'LEFT': Zone.LEFT
    }
    zone = zone_map.get(zone_str)
    
    # Parse the material
    material_str = str(reinf_data.get('Material', 'B500')).upper()
    material_map = {
        'B500': Reinforcement.B500,
        'B500A': Reinforcement.B500A,
        'B500B': Reinforcement.B500B,
        'B500C': Reinforcement.B500C,
        'B500K': Reinforcement.B500K,
        'B600KX': Reinforcement.B600KX,
        'B700K': Reinforcement.B700K,
        'G250': Reinforcement.G250,
        'K': Reinforcement.K,
        'N': Reinforcement.N,
        'R': Reinforcement.R,
        'S235JRG2': Reinforcement.S235JRG2,
        'Ss260': Reinforcement.Ss260,
        'Y': Reinforcement.Y,
        'Z': Reinforcement.Z,
        'A500HW': Reinforcement.A500HW,
        'A700HW': Reinforcement.A700HW,
        'Ks600': Reinforcement.Ks600,
        'Ps500': Reinforcement.Ps500,
        '_250': Reinforcement._250,
        '_420': Reinforcement._420
    }
    material_method = material_map.get(material_str, Reinforcement.B500)
    material = material_method()  # Call the method to get the material instance
    
    # Get numeric values
    diameter = float(reinf_data.get('Diameter', 0))
    spacing = float(reinf_data.get('Spacing', 0))
    concrete_cover = float(reinf_data.get('Cover', 0))
    
    return ReinfLayer(
        diameter=diameter,
        spacing=spacing,
        concrete_cover=concrete_cover,
        zone=zone,
        material=material
    )

def _read_foundation_from_excel(workbook) -> simplex.foundation.model.foundation.FoundationBase:
    """
    Read foundation from the 'Foundation' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        FoundationBase: A RectangularFoundation or LineFoundation object populated with data from the Excel file
    """
    try:
        # Get the 'Foundation' sheet
        sheet = workbook['Foundation']
    except KeyError:
        raise ValueError("Excel file must contain a 'Foundation' sheet")
    
    # Read foundation data from specific cells based on the Excel template structure
    # Row 1 contains headers (Field, Value, Units)
    # Row 2: Type field
    # Row 3: Lx bottom field
    # Row 4: Ly bottom field
    # Row 5: Lx top field
    # Row 6: Ly top field
    # Row 7: Height field
    # Row 8: Eccentricity x field
    # Row 9: Eccentricity y field
    # Row 10: Material field
    # Row 11: Top of footing field
    # Row 12: Position x field
    # Row 13: Position y field
    
    foundation_data = {}
    
    # Read values from specific cells (column 2 contains the values)
    foundation_data['Type'] = sheet.cell(row=2, column=2).value or "Point"
    foundation_data['Lx bottom'] = sheet.cell(row=3, column=2).value
    foundation_data['Ly bottom'] = sheet.cell(row=4, column=2).value
    foundation_data['Lx top'] = sheet.cell(row=5, column=2).value
    foundation_data['Ly top'] = sheet.cell(row=6, column=2).value
    foundation_data['Height'] = sheet.cell(row=7, column=2).value
    foundation_data['Eccentricity x'] = sheet.cell(row=8, column=2).value
    foundation_data['Eccentricity y'] = sheet.cell(row=9, column=2).value
    foundation_data['Material'] = sheet.cell(row=10, column=2).value
    foundation_data['Top of footing'] = sheet.cell(row=11, column=2).value
    foundation_data['Position x'] = sheet.cell(row=12, column=2).value
    foundation_data['Position y'] = sheet.cell(row=13, column=2).value
    
    # Determine foundation type
    foundation_type = str(foundation_data.get('Type', 'Point')).strip()
    
    print(f"Creating {foundation_type} foundation...")
    
    # Read reinforcement data
    reinforcement = None
    try:
        reinforcement = _read_reinforcement_from_excel(workbook, foundation_type)
    except Exception as e:
        print(f"Warning: Could not read reinforcement: {e}")
    
    if foundation_type.upper() == 'LINE':
        return _create_line_foundation(foundation_data, reinforcement)
    else:
        return _create_rectangular_foundation(foundation_data, reinforcement)

def _create_line_foundation(foundation_data: dict, reinforcement: simplex.foundation.model.reinforcement.Rebars = None) -> simplex.foundation.model.foundation.LineFoundation:
    """
    Create a LineFoundation object from foundation data.
    
    Args:
        foundation_data: Dictionary containing foundation values
        reinforcement: Optional reinforcement object
        
    Returns:
        LineFoundation: The line foundation object
    """
    from simplex.foundation.model.foundation import LineFoundation
    from simplex.core.geometry.point import Point2d
    from simplex.core.materials.materials import Concrete
    
    # Parse concrete material
    material_str = str(foundation_data.get('Material', 'C25/30')).replace('/', '_')
    material_map = {
        'C12_15': Concrete.C12_15,
        'C16_20': Concrete.C16_20,
        'C20_25': Concrete.C20_25,
        'C25_30': Concrete.C25_30,
        'C28_35': Concrete.C28_35,
        'C30_37': Concrete.C30_37,
        'C32_40': Concrete.C32_40,
        'C35_45': Concrete.C35_45,
        'C40_50': Concrete.C40_50,
        'C45_55': Concrete.C45_55,
        'C50_60': Concrete.C50_60,
        'C54_65': Concrete.C54_65,
        'C55_67': Concrete.C55_67,
        'C58_70': Concrete.C58_70,
        'C60_75': Concrete.C60_75,
        'C70_85': Concrete.C70_85,
        'C80_95': Concrete.C80_95,
        'C90_105': Concrete.C90_105
    }
    material_method = material_map.get(material_str, Concrete.C25_30)
    material = material_method()
    
    # Get numeric values, handling '-' as 0 or default
    def parse_value(value, default=0.0):
        if value == '-' or value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    lx_bottom = parse_value(foundation_data.get('Lx bottom'), 1.4)
    lx_top = parse_value(foundation_data.get('Lx top'), 1.4)
    height = parse_value(foundation_data.get('Height'), 0.4)
    eccentricity_x = parse_value(foundation_data.get('Eccentricity x'), 0.0)
    top_of_footing = parse_value(foundation_data.get('Top of footing'), 0.0)
    position_x = parse_value(foundation_data.get('Position x'), 0.0)
    position_y = parse_value(foundation_data.get('Position y'), 0.0)
    
    return LineFoundation(
        lx_bottom=lx_bottom,
        lx_top=lx_top,
        height=height,
        eccentricity_x=eccentricity_x,
        material=material,
        top_of_footing=top_of_footing,
        position=Point2d(position_x, position_y),
        reinforcement=reinforcement
    )

def _read_settings_from_excel(workbook) -> simplex.foundation.model.settings.Settings:
    """
    Read settings from the 'Settings' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        Settings: A Settings object populated with data from the Excel file
    """
    try:
        # Get the 'Settings' sheet
        sheet = workbook['Settings']
    except KeyError:
        raise ValueError("Excel file must contain a 'Settings' sheet")
    
    print(f"Creating settings from Excel data...")
    
    # Create the settings object directly from the workbook
    from simplex.foundation.model.settings import Settings
    
    # Parse code settings directly from workbook
    code_settings = _create_code_settings_from_workbook(workbook)
    
    # Parse concrete settings directly from workbook
    concrete_settings = _create_concrete_settings_from_workbook(workbook)
    
    # Parse soil settings directly from workbook
    soil_settings = _create_soil_settings_from_workbook(workbook)
    
    return Settings(
        code=code_settings,
        concrete_sett=concrete_settings,
        soil_sett=soil_settings
    )

def _read_design_settings_from_excel(workbook) -> 'simplex.foundation.design.auto_design.DesignSettings':
    """
    Read design settings from the 'DesignSettings' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        DesignSettings: A DesignSettings object populated with data from the Excel file
    """
    try:
        # Get the 'DesignSettings' sheet
        sheet = workbook['DesignSettings']
    except KeyError:
        raise ValueError("Excel file must contain a 'DesignSettings' sheet")
    
    print("Reading design settings from Excel template format...")
    
    # Initialize variables with defaults
    concrete_settings = None
    foundation_settings = None
    limit_utilisation = 1.0  # Default value

    
    # Read steel/concrete design settings
    try:
        # Read steel design values from row 3
        reinf_diameters_str = str(sheet.cell(row=3, column=1).value)
        spacing_min = float(sheet.cell(row=3, column=2).value)
        spacing_max = float(sheet.cell(row=3, column=3).value)
        spacing_step = float(sheet.cell(row=3, column=4).value)
        
        # Parse reinforcement diameters
        if reinf_diameters_str.strip():
            reinf_dia_list = [int(x.strip()) for x in reinf_diameters_str.split(',')]
        else:
            reinf_dia_list = [12, 16, 20]  # Default
        
        # Create Concrete settings for reinforcement design
        from simplex.foundation.design.auto_design import ConcreteDesign, Interval
        concrete_settings = ConcreteDesign(
            step=spacing_step,  # Convert mm to m for step
            rnfr_dia=reinf_dia_list,
            spacing_limits=Interval(spacing_min, spacing_max)
        )
        
        print(f"Steel/Concrete settings: diameters={reinf_dia_list}, spacing={spacing_min}-{spacing_max}mm, step={spacing_step}mm")
        
    except Exception as e:
        raise ValueError(f"Warning: Could not read steel design settings: {e}")
    
    # Read foundation geometry design settings
    try:
        # Read foundation design values from row 8
        step = float(sheet.cell(row=8, column=1).value)
        width_min = float(sheet.cell(row=8, column=2).value)
        width_max = float(sheet.cell(row=8, column=3).value)
        length_min = float(sheet.cell(row=8, column=4).value)
        length_max = float(sheet.cell(row=8, column=5).value)
        height_min = float(sheet.cell(row=8, column=6).value)
        height_max = float(sheet.cell(row=8, column=7).value)
        lx_ly_equal = bool(sheet.cell(row=8, column=8).value)

        
        # Create Foundation settings
        from simplex.foundation.design.auto_design import FoundationDesign, Interval
        foundation_settings = FoundationDesign(
            step=step,
            width=Interval(width_min, width_max),
            length=Interval(length_min, length_max),
            height=Interval(height_min, height_max),
            equal_length_width=lx_ly_equal
        )
        
        print(f"Foundation settings: step={step}m, width={width_min}-{width_max}m, length={length_min}-{length_max}m, height={height_min}-{height_max}m, ratio={lx_ly_equal}")
        
    except Exception as e:
        raise ValueError(f"Warning: Could not read foundation design settings: {e}")
    
    # Read general settings (limit utilisation)
    try:
        limit_util_value = sheet.cell(row=13, column=1).value
        if limit_util_value is not None:
            limit_utilisation = float(limit_util_value)
        print(f"General settings: limit utilisation={limit_utilisation}")
        
    except Exception as e:
        raise ValueError(f"Warning: Could not read general settings: {e}")
    
    # Create the DesignSettings object
    from simplex.foundation.design.auto_design import DesignSettings
    design_settings = DesignSettings(
        limit_utilisation=limit_utilisation,
        foundation_settings=foundation_settings,
        concrete_settings=concrete_settings
    )
    
    return design_settings

def _create_code_settings_from_workbook(workbook) -> simplex.foundation.model.settings.CodeSettings:
    """Create CodeSettings directly from workbook."""
    from simplex.foundation.model.settings import (
        CodeSettings, Annex, Consequence, Reliability
    )
    
    sheet = workbook['Settings']
    
    # Read values directly from specific cells
    annex_value = sheet.cell(row=2, column=2).value
    consequence_value = sheet.cell(row=3, column=2).value
    reliability_value = sheet.cell(row=4, column=2).value
    
    # Parse annex
    annex_str = str(annex_value or 'ANNEX_COMMON').upper()
    annex_map = {
        'ANNEX_COMMON': Annex.ANNEX_COMMON,
        'ANNEX_DENMARK': Annex.ANNEX_DENMARK,
        'ANNEX_SWEDEN': Annex.ANNEX_SWEDEN,
        'ANNEX_NORWAY': Annex.ANNEX_NORWAY,
        'ANNEX_FINLAND': Annex.ANNEX_FINLAND,
        'ANNEX_GREAT_BRITAIN': Annex.ANNEX_GREAT_BRITAIN
    }
    annex = annex_map.get(annex_str, Annex.ANNEX_COMMON)
    
    # Parse consequence
    consequence_str = str(consequence_value or 'CONSEQUENCE_CLASS_2').upper()
    consequence_map = {
        'CONSEQUENCE_CLASS_1': Consequence.CONSEQUENCE_CLASS_1,
        'CONSEQUENCE_CLASS_2': Consequence.CONSEQUENCE_CLASS_2,
        'CONSEQUENCE_CLASS_3': Consequence.CONSEQUENCE_CLASS_3
    }
    consequence = consequence_map.get(consequence_str, Consequence.CONSEQUENCE_CLASS_2)
    
    # Parse reliability
    reliability_str = str(reliability_value or 'RELIABILITY_CLASS_1').upper()
    reliability_map = {
        'RELIABILITY_CLASS_1': Reliability.RELIABILITY_CLASS_1,
        'RELIABILITY_CLASS_2': Reliability.RELIABILITY_CLASS_2,
        'RELIABILITY_CLASS_3': Reliability.RELIABILITY_CLASS_3
    }
    reliability = reliability_map.get(reliability_str, Reliability.RELIABILITY_CLASS_1)
    
    return CodeSettings(
        annex=annex,
        consequence=consequence,
        reliability=reliability
    )

def _create_concrete_settings_from_workbook(workbook) -> simplex.foundation.model.settings.ConcreteSettings:
    """Create ConcreteSettings directly from workbook."""
    from simplex.foundation.model.settings import (
        ConcreteSettings, FoundationDistribution, Fabrication, Exposure, 
        LifeCategory, InspectionLevel
    )
    
    sheet = workbook['Settings']
    
    # Read values directly from specific cells
    crack_value = sheet.cell(row=7, column=2).value
    distribution_value = sheet.cell(row=8, column=2).value
    fabrication_value = sheet.cell(row=9, column=2).value
    low_strength_variation_value = sheet.cell(row=10, column=2).value
    consider_min_reinforcement_value = sheet.cell(row=11, column=2).value
    exposure_class_value = sheet.cell(row=12, column=2).value
    life_category_value = sheet.cell(row=13, column=2).value
    critical_element_value = sheet.cell(row=14, column=2).value
    inspection_level_value = sheet.cell(row=15, column=2).value
    
    # Parse distribution
    distribution_str = str(distribution_value or 'FOUNDATION_DISTRIBUTION_PLASTIC').upper()
    distribution_map = {
        'FOUNDATION_DISTRIBUTION_ELASTIC': FoundationDistribution.FOUNDATION_DISTRIBUTION_ELASTIC,
        'FOUNDATION_DISTRIBUTION_PLASTIC': FoundationDistribution.FOUNDATION_DISTRIBUTION_PLASTIC
    }
    distribution = distribution_map.get(distribution_str, FoundationDistribution.FOUNDATION_DISTRIBUTION_PLASTIC)
    
    # Parse fabrication
    fabrication_str = str(fabrication_value or 'FABRICATION_IN_SITU').upper()
    fabrication_map = {
        'FABRICATION_IN_SITU': Fabrication.FABRICATION_IN_SITU,
        'FABRICATION_PREFAB': Fabrication.FABRICATION_PREFAB
    }
    fabrication = fabrication_map.get(fabrication_str, Fabrication.FABRICATION_IN_SITU)
    
    # Parse exposure class
    exposure_str = str(exposure_class_value or 'EXPOSURE_CLASS_XC1').upper()
    exposure_map = {
        'EXPOSURE_CLASS_X0': Exposure.EXPOSURE_CLASS_X0,
        'EXPOSURE_CLASS_XC1': Exposure.EXPOSURE_CLASS_XC1,
        'EXPOSURE_CLASS_XC2': Exposure.EXPOSURE_CLASS_XC2,
        'EXPOSURE_CLASS_XC3': Exposure.EXPOSURE_CLASS_XC3,
        'EXPOSURE_CLASS_XC4': Exposure.EXPOSURE_CLASS_XC4,
        'EXPOSURE_CLASS_XD1': Exposure.EXPOSURE_CLASS_XD1,
        'EXPOSURE_CLASS_XD2': Exposure.EXPOSURE_CLASS_XD2,
        'EXPOSURE_CLASS_XD3': Exposure.EXPOSURE_CLASS_XD3,
        'EXPOSURE_CLASS_XS1': Exposure.EXPOSURE_CLASS_XS1,
        'EXPOSURE_CLASS_XS2': Exposure.EXPOSURE_CLASS_XS2,
        'EXPOSURE_CLASS_XS3': Exposure.EXPOSURE_CLASS_XS3,
        'EXPOSURE_CLASS_XA1': Exposure.EXPOSURE_CLASS_XA1,
        'EXPOSURE_CLASS_XA2': Exposure.EXPOSURE_CLASS_XA2,
        'EXPOSURE_CLASS_XA3': Exposure.EXPOSURE_CLASS_XA3,
        'EXPOSURE_CLASS_XF1': Exposure.EXPOSURE_CLASS_XF1,
        'EXPOSURE_CLASS_XF2': Exposure.EXPOSURE_CLASS_XF2,
        'EXPOSURE_CLASS_XF3': Exposure.EXPOSURE_CLASS_XF3,
        'EXPOSURE_CLASS_XF4': Exposure.EXPOSURE_CLASS_XF4
    }
    exposure_class = exposure_map.get(exposure_str, Exposure.EXPOSURE_CLASS_XC1)
    
    # Parse life category
    life_str = str(life_category_value or 'LIFE_CATEGORY_L50').upper()
    life_map = {
        'LIFE_CATEGORY_L20': LifeCategory.LIFE_CATEGORY_L20,
        'LIFE_CATEGORY_L50': LifeCategory.LIFE_CATEGORY_L50,
        'LIFE_CATEGORY_L100': LifeCategory.LIFE_CATEGORY_L100
    }
    life_category = life_map.get(life_str, LifeCategory.LIFE_CATEGORY_L50)
    
    # Parse inspection level
    inspection_str = str(inspection_level_value or 'INSPECTION_LEVEL_NORMAL').upper()
    inspection_map = {
        'INSPECTION_LEVEL_RELAXED': InspectionLevel.INSPECTION_LEVEL_RELAXED,
        'INSPECTION_LEVEL_NORMAL': InspectionLevel.INSPECTION_LEVEL_NORMAL,
        'INSPECTION_LEVEL_TIGHTENED': InspectionLevel.INSPECTION_LEVEL_TIGHTENED
    }
    inspection_level = inspection_map.get(inspection_str, InspectionLevel.INSPECTION_LEVEL_NORMAL)
    
    # Parse boolean values
    def parse_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('true', '1', 'yes', 'on')
        return default
    
    low_strength_variation = parse_bool(low_strength_variation_value, False)
    consider_min_reinforcement = parse_bool(consider_min_reinforcement_value, True)
    critical_element = parse_bool(critical_element_value, False)
    
    # Parse numeric values
    crack = float(crack_value or 0.001)
    
    return ConcreteSettings(
        crack=crack,
        distribution=distribution,
        fabrication=fabrication,
        low_strength_variation=low_strength_variation,
        consider_min_reinforcement=consider_min_reinforcement,
        exposure_class=exposure_class,
        life_category=life_category,
        critical_element=critical_element,
        inspection_level=inspection_level
    )

def _create_soil_settings_from_workbook(workbook) -> simplex.foundation.model.settings.SoilSettings:
    """Create SoilSettings directly from workbook."""
    from simplex.foundation.model.settings import (
        SoilSettings, DesignApproach, GeotechnicalCategory, SoilPunchingType
    )
    
    sheet = workbook['Settings']
    
    # Read values directly from specific cells
    des_appr_value = sheet.cell(row=18, column=2).value
    geo_cat_value = sheet.cell(row=19, column=2).value
    check_settl_value = sheet.cell(row=20, column=2).value
    abs_sttl_value = sheet.cell(row=21, column=2).value
    punching_value = sheet.cell(row=22, column=2).value
    
    # Parse design approach
    des_appr_str = str(des_appr_value or 'DESIGN_APPROACH_2').upper()
    des_appr_map = {
        'DESIGN_APPROACH_1': DesignApproach.DESIGN_APPROACH_1,
        'DESIGN_APPROACH_2': DesignApproach.DESIGN_APPROACH_2,
        'DESIGN_APPROACH_3': DesignApproach.DESIGN_APPROACH_3
    }
    des_appr = des_appr_map.get(des_appr_str, DesignApproach.DESIGN_APPROACH_2)
    
    # Parse geotechnical category
    geo_cat_str = str(geo_cat_value or 'GEOTECHNICAL_CATEGORY_1').upper()
    geo_cat_map = {
        'GEOTECHNICAL_CATEGORY_1': GeotechnicalCategory.GEOTECHNICAL_CATEGORY_1,
        'GEOTECHNICAL_CATEGORY_2': GeotechnicalCategory.GEOTECHNICAL_CATEGORY_2,
        'GEOTECHNICAL_CATEGORY_3': GeotechnicalCategory.GEOTECHNICAL_CATEGORY_3
    }
    geo_cat = geo_cat_map.get(geo_cat_str, GeotechnicalCategory.GEOTECHNICAL_CATEGORY_1)
    
    # Parse soil punching type
    punching_str = str(punching_value or 'SOIL_PUNCHING_TYPE_1_2').upper()
    punching_map = {
        'SOIL_PUNCHING_TYPE_WIDTH_1_2': SoilPunchingType.SOIL_PUNCHING_TYPE_WIDTH_1_2,
        'SOIL_PUNCHING_TYPE_1_2': SoilPunchingType.SOIL_PUNCHING_TYPE_1_2,
        'SOIL_PUNCHING_TYPE_1_3': SoilPunchingType.SOIL_PUNCHING_TYPE_1_3,
        'SOIL_PUNCHING_TYPE_1_4': SoilPunchingType.SOIL_PUNCHING_TYPE_1_4
    }
    punching = punching_map.get(punching_str, SoilPunchingType.SOIL_PUNCHING_TYPE_1_2)
    
    # Parse boolean values
    def parse_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('true', '1', 'yes', 'on')
        return default
    
    check_settl = parse_bool(check_settl_value, True)
    
    # Parse numeric values
    abs_sttl = float(abs_sttl_value or 20)
    
    return SoilSettings(
        des_appr=des_appr,
        geo_cat=geo_cat,
        check_settl=check_settl,
        abs_sttl=abs_sttl,
        punching=punching
    )

def _create_rectangular_foundation(foundation_data: dict, reinforcement: simplex.foundation.model.reinforcement.Rebars = None) -> simplex.foundation.model.foundation.RectangularFoundation:
    """
    Create a RectangularFoundation object from foundation data.
    
    Args:
        foundation_data: Dictionary containing foundation values
        reinforcement: Optional reinforcement object
        
    Returns:
        RectangularFoundation: The rectangular foundation object
    """
    from simplex.foundation.model.foundation import RectangularFoundation
    from simplex.core.geometry.point import Point2d
    from simplex.core.materials.materials import Concrete
    
    # Parse concrete material
    material_str = str(foundation_data.get('Material', 'C25/30')).replace('/', '_')
    material_map = {
        'C12_15': Concrete.C12_15,
        'C16_20': Concrete.C16_20,
        'C20_25': Concrete.C20_25,
        'C25_30': Concrete.C25_30,
        'C28_35': Concrete.C28_35,
        'C30_37': Concrete.C30_37,
        'C32_40': Concrete.C32_40,
        'C35_45': Concrete.C35_45,
        'C40_50': Concrete.C40_50,
        'C45_55': Concrete.C45_55,
        'C50_60': Concrete.C50_60,
        'C54_65': Concrete.C54_65,
        'C55_67': Concrete.C55_67,
        'C58_70': Concrete.C58_70,
        'C60_75': Concrete.C60_75,
        'C70_85': Concrete.C70_85,
        'C80_95': Concrete.C80_95,
        'C90_105': Concrete.C90_105
    }
    material_method = material_map.get(material_str, Concrete.C25_30)
    material = material_method()
    
    # Get numeric values, handling '-' as 0 or default
    def parse_value(value, default=0.0):
        if value == '-' or value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    lx_bottom = parse_value(foundation_data.get('Lx bottom'), 1.4)
    ly_bottom = parse_value(foundation_data.get('Ly bottom'), 1.4)
    lx_top = parse_value(foundation_data.get('Lx top'), 0.3)
    ly_top = parse_value(foundation_data.get('Ly top'), 0.3)
    height = parse_value(foundation_data.get('Height'), 0.4)
    eccentricity_x = parse_value(foundation_data.get('Eccentricity x'), 0.0)
    eccentricity_y = parse_value(foundation_data.get('Eccentricity y'), 0.0)
    top_of_footing = parse_value(foundation_data.get('Top of footing'), 0.0)
    position_x = parse_value(foundation_data.get('Position x'), 0.0)
    position_y = parse_value(foundation_data.get('Position y'), 0.0)
    
    return RectangularFoundation(
        lx_bottom=lx_bottom,
        ly_bottom=ly_bottom,
        lx_top=lx_top,
        ly_top=ly_top,
        height=height,
        eccentricity_x=eccentricity_x,
        eccentricity_y=eccentricity_y,
        material=material,
        top_of_footing=top_of_footing,
        position=Point2d(position_x, position_y),
        reinforcement=reinforcement
    )

def _read_soil_complex_from_excel(workbook) -> simplex.foundation.model.soil.SoilComplex:
    """
    Read soil complex from the 'SoilComplex' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        SoilComplex: A SoilComplex object populated with data from the Excel file
    """
    try:
        # Get the 'SoilComplex' sheet
        sheet = workbook['Soil']
    except KeyError:
        raise ValueError("Excel file must contain a 'SoilComplex' sheet")
    
    # Read the header row to understand the structure
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=12, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
    
    
    # Read soil materials and their properties
    soil_materials = []
    top_of_layers = []
    
    # Read data starting from row 2
    for row in range(13, sheet.max_row + 1):
        # Check if the row has any data
        row_has_data = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                row_has_data = True
                break
        
        if not row_has_data:
            continue
        
        # Read the soil material data
        soil_data = {}
        for col in range(1, min(len(headers) + 1, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                header_name = headers[col - 1]
                soil_data[header_name] = cell_value
        
        if not soil_data:
            continue
        
        # Create soil material based on type
        soil_material = _create_soil_material(soil_data)
        if soil_material:
            soil_materials.append(soil_material)
            
            # Get top of layer (assuming it's in the "Top of Layer [m]" column)
            top_of_layer = soil_data.get('Top of Layer [m]', 0.0)
            if isinstance(top_of_layer, str):
                try:
                    top_of_layer = float(top_of_layer)
                except ValueError:
                    top_of_layer = 0.0
            top_of_layers.append(top_of_layer)
    
    # Get depth limit and ground water from the last columns
    depth_limit = float(sheet.cell(row=12, column=14).value or -10.0)
    ground_water = float(sheet.cell(row=13, column=14).value or -2.0)
    
    
    # Create borehole
    from simplex.foundation.model.soil import Borehole
    borehole = Borehole(
        soil_materials=soil_materials,
        top_of_layers=top_of_layers
    )
    
    # Create soil complex
    from simplex.foundation.model.soil import SoilComplex
    soil_complex = SoilComplex(
        borehole=borehole,
        depth_limit=depth_limit,
        ground_water=ground_water
    )
    
    return soil_complex

def _create_soil_material(soil_data: dict) -> simplex.foundation.model.soil.SoilMaterial:
    """
    Create a SoilMaterial object from soil data.
    
    Args:
        soil_data: Dictionary containing soil material values
        
    Returns:
        SoilMaterial: The soil material object
    """
    from simplex.foundation.model.soil import SoilMaterial
    
    name = str(soil_data.get('Name', 'Unknown Soil'))
    soil_type = str(soil_data.get('Type', 'Drained')).strip()
    
    # Parse numeric values, handling '-' as None
    def parse_value(value, default=0.0):
        if value == '-' or value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    phik = parse_value(soil_data.get('phik'))
    ck = parse_value(soil_data.get('ck'))
    cuk = parse_value(soil_data.get('cuk'))
    rk = parse_value(soil_data.get('rk'))
    gamma = parse_value(soil_data.get('gamma'), 18.0)
    gamma_eff = parse_value(soil_data.get('gamma eff'), 10.0)
    m0 = parse_value(soil_data.get('m0'), 20000.0)
    
    # Create soil material based on type
    if soil_type.upper() == 'DRAINED':
        return SoilMaterial.drained(
            name=name,
            phik=phik or 35.0,
            ck=ck or 0.0,
            gamma=gamma,
            gamma_eff=gamma_eff,
            m0=m0
        )
    elif soil_type.upper() == 'UNDRAINED':
        return SoilMaterial.undrained(
            name=name,
            cuk=cuk or 0.0,
            gamma=gamma,
            gamma_eff=gamma_eff,
            m0=m0
        )
    elif soil_type.upper() == 'COMBINED':
        return SoilMaterial.combined(
            name=name,
            phik=phik or 0.0,
            ck=ck or 0.0,
            cuk=cuk or 0.0,
            gamma=gamma,
            gamma_eff=gamma_eff,
            m0=m0
        )
    elif soil_type.upper() == 'ROCK':
        return SoilMaterial.rock(
            name=name,
            rk=rk or 0.0,
            gamma=gamma,
            gamma_eff=gamma_eff,
            m0=m0
        )
    else:
        # Default to drained
        return SoilMaterial.drained(
            name=name,
            phik=phik or 35.0,
            ck=ck or 0.0,
            gamma=gamma,
            gamma_eff=gamma_eff,
            m0=m0
        )

def _read_soil_simple_from_excel(workbook) -> simplex.foundation.model.soil.SoilSimple:
    """
    Read soil simple from the 'Soil' sheet of the Excel workbook.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        SoilSimple: A SoilSimple object populated with data from the Excel file
    """
    try:
        # Get the 'Soil' sheet
        sheet = workbook['Soil']
    except KeyError:
        raise ValueError("Excel file must contain a 'Soil' sheet")
    
    # Create soil simple object
    from simplex.foundation.model.soil import SoilSimple
    soil_simple = SoilSimple(
        allowed_soil_pressure_sls=float(sheet.cell(row=5, column=2).value or 200.0),
        allowed_soil_pressure_uls=float(sheet.cell(row=6, column=2).value or 300.0),
        friction_coefficient=float(sheet.cell(row=7, column=2).value or 0.1)
    )
    
    return soil_simple

def _read_soil_from_excel(workbook) -> simplex.foundation.model.soil.SoilComplex | simplex.foundation.model.soil.SoilSimple:
    """
    Read soil from the Excel workbook, automatically determining whether to use
    complex or simple soil based on the soil type specified in the Excel file.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        SoilComplex or SoilSimple: A soil object populated with data from the Excel file
        
    Raises:
        ValueError: If neither complex nor simple soil can be read successfully
    """
    # Determine the soil type from the Excel file
    soil_type = _get_soil_type(workbook)
    
    print(f"Detected soil type: {soil_type}")
    
    if soil_type.upper() == 'COMPLEX':
        try:
            soil = _read_soil_complex_from_excel(workbook)
            print("Successfully read SoilComplex")
            return soil
        except (ValueError, Exception) as e:
            print(f"Could not read SoilComplex: {e}")
            # Fall back to simple soil
            try:
                soil = _read_soil_simple_from_excel(workbook)
                print("Successfully read SoilSimple as fallback")
                return soil
            except (ValueError, Exception) as e:
                raise ValueError(f"Could not read SoilSimple as fallback: {e}")
    else:
        # Default to simple soil
        try:
            soil = _read_soil_simple_from_excel(workbook)
            print("Successfully read SoilSimple")
            return soil
        except (ValueError, Exception) as e:
            print(f"Could not read SoilSimple: {e}")
            # Fall back to complex soil
            try:
                soil = _read_soil_complex_from_excel(workbook)
                print("Successfully read SoilComplex as fallback")
                return soil
            except (ValueError, Exception) as e:
                raise ValueError(f"Could not read SoilComplex as fallback: {e}")