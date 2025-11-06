import openpyxl
from copy import copy

# INPUT EXAMPLE
# data_dict
# {
#     "公司1": {
#         "entityType": "ORGANISATION",
#         "entityClass": "Organisation",
#         "Screening Result": 1,
#         "Note": "2 screening results found and 0 matched",
#         "English Name": "com 1",
#         "Percent": None
#     },
#     "公司2": {
#         "entityType": "ORGANISATION",
#         "entityClass": "Shareholders",
#         "Screening Result": 1,
#         "Note": "2 screening results found and 0 matched",
#         "English Name": "com 2",
#         "Percent": 100
#     },
#     "公司3": {
#         "entityType": "ORGANISATION",
#         "entityClass": "Shareholders",
#         "Screening Result": 1,
#         "Note": "2 screening results found and 0 matched",
#         "English Name": "com 3",
#         "Percent": 70
#     },
#     "公司4": {
#         "entityType": "ORGANISATION",
#         "entityClass": "Shareholders",
#         "Screening Result": 1,
#         "Note": "2 screening results found and 0 matched",
#         "English Name": "com 3",
#         "Percent": 60
#     },
# }

def find_first(sheet, text, case_sensitive = False):
    """
    在 sheet 中查找首次出现 text 的单元格
    返回 (row, col)  行、列都从 1 开始
    找不到返回 (None, None)
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and type(cell.value) == str:
                if text in cell.value and case_sensitive:
                    return cell.row, cell.column
                elif text.strip().lower() in cell.value.strip().lower() and not case_sensitive:
                    return cell.row, cell.column
            else:
                continue
    return None, None

def insert_row_below_with_formatting(ws, row_idx):
    """
    在当前行下方插入一行，并完全复制当前行的内容和格式
    
    Args:
        ws: 工作表对象
        row_idx: 当前行的位置（在该行下方插入新行）
    """
    # 在当前行下方插入新行
    ws.insert_rows(row_idx + 1)
    
    # 复制当前行的所有内容和格式
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=row_idx, column=col)
        target_cell = ws.cell(row=row_idx + 1, column=col)
        
        # 复制值
        # target_cell.value = source_cell.value
        
        # 复制样式（如果有）
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    # 复制行高
    ws.row_dimensions[row_idx + 1].height = ws.row_dimensions[row_idx].height

def write_conflict_checklist(data_dict, project_name:str, template_path, output_path):
    """
    将数据写入冲突检查清单模板
    
    Args:
        data_dict: 包含公司数据的字典
        template_path: 模板文件路径
        output_path: 输出文件路径
    """
    
    # 加载模板
    wb = openpyxl.load_workbook(template_path)
    
    # 处理ConflictCheck工作表
    process_conflict_check_sheet(wb, project_name, data_dict)
    
    # 处理world check result工作表
    process_world_check_sheet(wb, data_dict)
    
    # 保存文件
    wb.save(output_path)
    print(f"文件已保存到: {output_path}")

def process_conflict_check_sheet(wb, project_name, data_dict):
    """处理ConflictCheck工作表"""
    ws = wb['ConflictCheck']

    ws[f'A1'] = f"Project Name:  {project_name}"
    
    # 按实体类型分组
    organisations = []
    shareholders = []
    
    for chinese_name, data in data_dict.items():
        if data['entityClass'] == 'Organisation':
            organisations.append((chinese_name, data))
        elif data['entityClass'] == 'Shareholders':
            shareholders.append((chinese_name, data))
    
    # 写入A.公司部分 
    org_start_row, org_start_col = find_first(ws, "[Organisation]")
    for i, (chinese_name, data) in enumerate(organisations):
        row = org_start_row + i
        insert_row_below_with_formatting(ws, row)
        ws[f'A{row}'] = chinese_name
        ws[f'B{row}'] = data['English Name']
        

    # 写入B.主要股东部分
    shareholder_start_row, shareholder_start_col = find_first(ws, "[Shareholders]")
    for i, (chinese_name, data) in enumerate(shareholders):
        row = shareholder_start_row + i
        insert_row_below_with_formatting(ws, row)
        ws[f'A{row}'] = chinese_name
        ws[f'B{row}'] = data['English Name']
        if data['Percent'] is not None:
            ws[f'C{row}'] = float(data['Percent']) / 100  # 转换为小数形式
        

def process_world_check_sheet(wb, data_dict):
    """处理world check result工作表"""
    ws = wb['world check result']
    
    # 按实体类型分组
    organisations = []
    shareholders = []
    
    for chinese_name, data in data_dict.items():
        if data['entityClass'] == 'Organisation':
            organisations.append((chinese_name, data))
        elif data['entityClass'] == 'Shareholders':
            shareholders.append((chinese_name, data))
    
    # 更新组织计数
    ws['C1'] = f"{len(organisations)} Organization"
    
    # 写入组织数据
    org_start_row = 3
    for i, (chinese_name, data) in enumerate(organisations):
        row = org_start_row + i
        insert_row_below_with_formatting(ws, row)
        ws[f'B{row}'] = i+1
        ws[f'C{row}'] = chinese_name
        ws[f'D{row}'] = data['English Name']
        ws[f'E{row}'] = data['Screening Result']
        ws[f'F{row}'] = data['Note']


    # 更新股东计数
    shareholder_title_row = 5 + len(organisations)  # 根据组织数量动态调整位置
    ws[f'C{shareholder_title_row}'] = f"{len(shareholders)} Major Shareholders"
    
    # 写入股东数据
    shareholder_start_row = shareholder_title_row + 2
    for i, (chinese_name, data) in enumerate(shareholders):
        row = shareholder_start_row + i
        insert_row_below_with_formatting(ws, row)
        ws[f'B{row}'] = i+1
        ws[f'C{row}'] = chinese_name
        ws[f'D{row}'] = data['English Name']
        ws[f'E{row}'] = data['Screening Result']
        ws[f'F{row}'] = data['Note']


def insert_rows_with_formatting(ws, start_row, num_rows):
    """
    在指定位置插入行并复制格式
    
    Args:
        ws: 工作表对象
        start_row: 开始插入的行
        num_rows: 要插入的行数
    """
    ws.insert_rows(start_row, num_rows)
    
    # 尝试复制上一行的格式 (如果有上一行的话)
    if start_row > 1:
        source_row = start_row - 1
        for row in range(start_row, start_row + num_rows):
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=row, column=col)
                
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)