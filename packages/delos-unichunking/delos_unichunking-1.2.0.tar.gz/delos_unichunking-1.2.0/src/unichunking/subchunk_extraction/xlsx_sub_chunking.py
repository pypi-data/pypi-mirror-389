"""Extract subchunks from XLSX file."""

from pathlib import Path
from typing import Any

import openpyxl
from xlcalculator import Evaluator

from unichunking.subchunk_extraction.xlsx_utils import ModelCompilerSimplified
from unichunking.tools import (
    compact_matrix,
    convert_file,
    extract_media_xlsx,
)
from unichunking.types import (
    ChunkPosition,
    MatrixTable,
    SubChunk,
)
from unichunking.utils import logger


def _is_empty_line(table: list[list[str]], i: int) -> bool:
    if (not table) or i >= len(table):
        return False
    return all(not table[i][j] for j in range(len(table[i])))


def _is_empty_column(table: list[list[str]], j: int) -> bool:
    if (not table) or j >= len(table[0]):
        return False
    return all(not table[i][j] for i in range(len(table)))


def _clean_borders(table: list[list[str]]) -> list[list[str]]:
    keep_going = len(table) > 0
    while keep_going:
        if _is_empty_line(table, 0):
            table.pop(0)
        elif _is_empty_line(table, len(table) - 1):
            table.pop()
        elif _is_empty_column(table, 0):
            for i in range(len(table)):
                table[i].pop(0)
        elif _is_empty_column(table, len(table) - 1):
            for i in range(len(table)):
                table[i].pop()
        else:
            keep_going = False
        keep_going = keep_going and len(table) > 0 and len(table[0]) > 0
    return table


def _extract_table_from_sheet(
    wb: openpyxl.Workbook,
    sheet: Any,
    sheet_idx: int,
    evaluator: Evaluator,
) -> MatrixTable:
    """Extracts the content of a XLSX sheet in the form of the minimal rectangular area that contains all non-empty cells on the sheet.

    Formulae are replaced by their outputs.

    Args:
        wb: The Workbook object representing the active file.
        sheet: The Worksheet object representing the active sheet.
        sheet_idx: The index of the active sheet.
        evaluator: The Evaluator object used to compute formulae outputs.

    Returns;
        The MatrixTable corresponding to the minimal rectangular area that contains all non-empty cells on the sheet.
    """
    merged_cells = sheet.merged_cells.ranges.copy()
    for merge_range in merged_cells:
        value = sheet.cell(merge_range.min_row, merge_range.min_col).value
        sheet.unmerge_cells(str(merge_range))
        for row in sheet.iter_rows(
            min_row=merge_range.min_row,
            max_row=merge_range.max_row,
            min_col=merge_range.min_col,
            max_col=merge_range.max_col,
        ):
            for cell in row:
                cell.value = value

    table: list[list[str]] = []
    for row_idx in range(sheet.min_row, sheet.max_row + 1):
        line: list[str] = []
        for col_idx in range(sheet.min_column, sheet.max_column + 1):
            # Computing the result of a formula :
            if sheet.cell(row_idx, col_idx).data_type == "f":
                cell_adress = f"{wb.sheetnames[sheet_idx]}!{sheet.cell(row_idx, col_idx).coordinate}"
                try:
                    cell_value: str = str(evaluator.evaluate(cell_adress))  # type: ignore
                    sheet.cell(row_idx, col_idx).value = cell_value
                except Exception as e:  # noqa: BLE001
                    logger.debug(
                        "Error occured when computing formula : {}",
                        e,
                    )

            line.append(str(sheet.cell(row_idx, col_idx).value).replace("None", ""))

        table.append(line)

    return MatrixTable("", table)


def _split_xlsx_sheet(sheet_table: MatrixTable) -> list[MatrixTable]:
    """Identifies the different groups of cells on a XLSX sheet, initially represented by a single table containing all the non-empty cells.

    Iteratively removes blank lines/columns on the edges and splits the table along interior blank lines/columns.

    Args:
        sheet_table: A MatrixTable object representing the minimal rectangular area that contains all non-empty cells on the sheet.

    Returns:
        A list containing a MatrixTable object for each cell group on the XLSX sheet.
    """
    tables_temp = [sheet_table]
    tables_done: list[MatrixTable] = []

    while len(tables_temp) > 0:
        table = tables_temp.pop().content

        table = _clean_borders(table)

        next_iter = not table
        if next_iter:
            continue

        for i in range(len(table)):
            if _is_empty_line(table, i):
                sub_tab_1 = MatrixTable("", table[:i])
                sub_tab_2 = MatrixTable("", table[i + 1 :])
                tables_temp.extend([sub_tab_2, sub_tab_1])
                next_iter = True
                break

        if next_iter:
            continue

        for j in range(len(table[0])):
            if _is_empty_column(table, j):
                sub_tab_1 = MatrixTable(
                    "",
                    [[table[i][k] for k in range(j)] for i in range(len(table))],
                )
                sub_tab_2 = MatrixTable(
                    "",
                    [
                        [table[i][k] for k in range(j + 1, len(table[0]))]
                        for i in range(len(table))
                    ],
                )
                tables_temp.extend([sub_tab_2, sub_tab_1])
                next_iter = True
                break

        if next_iter:
            continue

        tables_done.append(MatrixTable("", table))

    return tables_done


async def extract_subchunks_xlsx(
    path: Path,
    extension: str,
) -> tuple[list[SubChunk], list[MatrixTable], list[str], Path, bool]:
    """Filetype-specific function : extracts subchunks from a XLSX file.

    For XLS & ODS : local file converted to XLSX, subchunks extracted from XLSX.
    Keeps a named sheets structure.
    (ex : "New sheet 'Sheet1' : ... ; End of sheet 'Sheet1'").

    Args:
        path: Path to the local file.
        extension: File extension.

    Returns:
        A tuple containing three lists and a Path:
        - List of subchunks, containing actual text or markers pointing to table/chart/image objects.
        - List of table/chart objects, of class MatrixTable.
        - List of image objects.
        - Path to the processed XLSX file, different from the initial path if a conversion occured.

    """
    filename = path.name
    if extension != "xlsx":
        new_path = await convert_file(path, "xlsx")
        if new_path is not None:
            path = new_path
        else:
            return [], [], [], Path(), False

    wb = openpyxl.load_workbook(path)

    compiler = ModelCompilerSimplified()
    compiler.read_and_parse_simplified(str(path))
    evaluator = Evaluator(compiler.model)

    subchunks: list[SubChunk] = []
    tables: list[MatrixTable] = []
    images: list[str] = []
    image_bank: list[bytes] = []

    for sheet_idx in range(len(wb.worksheets)):
        sheet = wb.worksheets[sheet_idx]
        subchunks.append(
            SubChunk(
                subchunk_id=len(subchunks),
                content=f"** New sheet '{wb.sheetnames[sheet_idx]}' : ",
                page=sheet_idx,
                position=ChunkPosition(0, 0, 0, 0),
                file_name=filename,
            ),
        )

        table = _extract_table_from_sheet(
            wb=wb,
            sheet=sheet,
            sheet_idx=sheet_idx,
            evaluator=evaluator,
        )

        splitted_tables = _split_xlsx_sheet(table)

        for table in splitted_tables:
            compact_matrix(
                table=table,
                do_delete_empty_lines=False,
                do_delete_empty_columns=False,
                do_merge_fake_lines=False,
                do_merge_fake_columns=False,
            )

            is_table = True
            if table.abort():
                subchunk_content = table.read_as_str()
                is_table = False
            else:
                subchunk_content = f"{len(tables)}"
                tables.append(table)
            if subchunk_content.strip():
                subchunks.append(
                    SubChunk(
                        subchunk_id=len(subchunks),
                        content=subchunk_content,
                        page=sheet_idx,
                        position=ChunkPosition(0, 0, 0, 0),
                        file_name=filename,
                        content_type="table" if is_table else "text",
                    ),
                )

        charts, images_bytes = extract_media_xlsx(path, page_idx=sheet_idx + 1)

        for chart in charts:
            subchunks.append(
                SubChunk(
                    subchunk_id=len(subchunks),
                    content=f"{len(tables)}",
                    page=sheet_idx,
                    position=ChunkPosition(0, 0, 0, 0),
                    file_name=filename,
                    content_type="table",
                ),
            )
            tables.append(chart)

        for image_bytes in images_bytes:
            if image_bytes in image_bank:
                subchunks.append(
                    SubChunk(
                        subchunk_id=len(subchunks),
                        content=f"{image_bank.index(image_bytes)}",
                        page=sheet_idx,
                        position=ChunkPosition(0, 0, 0, 0),
                        file_name=filename,
                        content_type="image",
                    ),
                )
            else:
                subchunks.append(
                    SubChunk(
                        subchunk_id=len(subchunks),
                        content=f"{len(images)}",
                        page=sheet_idx,
                        position=ChunkPosition(0, 0, 0, 0),
                        file_name=filename,
                        content_type="image",
                    ),
                )
                images.append(f"IMAGE {len(images)} ")
                image_bank.append(image_bytes)

        subchunks.append(
            SubChunk(
                subchunk_id=len(subchunks),
                content=f" ; End of sheet '{wb.sheetnames[sheet_idx]}' ** ",
                page=sheet_idx,
                position=ChunkPosition(1, 1, 1, 1),
                file_name=filename,
            ),
        )

    wb.close()

    return subchunks, tables, images, path, True
