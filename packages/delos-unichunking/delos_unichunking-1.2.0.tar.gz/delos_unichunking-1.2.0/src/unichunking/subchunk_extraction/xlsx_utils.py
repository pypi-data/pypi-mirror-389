"""Utilities for XLSX subchunking."""

import ast
import re
from typing import Callable, cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from xlcalculator import ModelCompiler

from unichunking.utils import logger


class ModelCompilerSimplified(ModelCompiler):
    """ModelCompiler with simplified methods."""

    def _is_simple_formula(self, formula: str) -> bool:
        if not formula or not formula.startswith("="):
            return False
        simple_ops = re.match(r"^=([A-Z]+\d+|\d+)([+\-*/]([A-Z]+\d+|\d+))*$", formula)
        simple_funcs = re.match(
            r"^=(SUM|AVERAGE|MIN|MAX)\([^)]*\)$", formula, re.IGNORECASE,
        )
        return bool(simple_ops or simple_funcs)

    def _safe_eval_formula(self, formula: str, sheet: Worksheet) -> float | None:
        formula = formula.lstrip("=")

        if formula.upper().startswith("SUM("):
            rng = formula[4:-1]
            cells = sheet[rng]
            values = [
                c.value
                for row in cells
                for c in row
                if isinstance(c.value, (int, float))
            ]
            return float(sum(cast("list[float]", values))) if values else None

        if formula.upper().startswith("AVERAGE("):
            rng = formula[8:-1]
            cells = sheet[rng]
            values = [
                c.value
                for row in cells
                for c in row
                if isinstance(c.value, (int, float))
            ]
            return (
                float(sum(cast("list[float]", values)) / len(values))
                if values
                else None
            )

        if formula.upper().startswith("MIN("):
            rng = formula[4:-1]
            cells = sheet[rng]
            values = [
                c.value
                for row in cells
                for c in row
                if isinstance(c.value, (int, float))
            ]
            return float(min(cast("list[float]", values))) if values else None

        if formula.upper().startswith("MAX("):
            rng = formula[4:-1]
            cells = sheet[rng]
            values = [
                c.value
                for row in cells
                for c in row
                if isinstance(c.value, (int, float))
            ]
            return float(max(cast("list[float]", values))) if values else None

        # Fallback: arithmetic using a safe AST evaluator
        def cell_to_val(match: re.Match[str]) -> str:
            cell_value = sheet[match.group(0)].value
            if isinstance(cell_value, (int, float)):
                return str(cell_value)
            return "0"

        expr = re.sub(r"[A-Z]+\d+", lambda m: cell_to_val(m), formula)
        return self._evaluate_arithmetic_expression(expr)

    def _evaluate_arithmetic_expression(self, expr: str) -> float | None:
        """Safely evaluate a simple arithmetic expression (numbers + - * /)."""

        def to_num(x: float | str | None) -> float | None:
            if isinstance(x, (int, float)):
                return float(x)
            return None

        binop_map: dict[type[ast.AST], Callable[[float, float], float | None]] = {
            ast.Add: lambda a, b: a + b,
            ast.Sub: lambda a, b: a - b,
            ast.Mult: lambda a, b: a * b,
            ast.Div: lambda a, b: None if b == 0 else a / b,
        }

        unary_map: dict[type[ast.AST], Callable[[float], float | None]] = {
            ast.UAdd: lambda a: a,
            ast.USub: lambda a: -a,
        }

        def eval_node(node: ast.AST) -> float | None:  # noqa: PLR0911
            if isinstance(node, ast.Expression):
                return eval_node(node.body)

            if isinstance(node, ast.BinOp) and type(node.op) in binop_map:
                left, right = eval_node(node.left), eval_node(node.right)
                if left is None or right is None:
                    return None
                return binop_map[type(node.op)](left, right)

            if isinstance(node, ast.UnaryOp) and type(node.op) in unary_map:
                operand = eval_node(node.operand)
                if operand is None:
                    return None
                return unary_map[type(node.op)](operand)

            if isinstance(node, ast.Constant):
                return to_num(node.value)

            return None

        tree = ast.parse(expr, mode="eval")

        return eval_node(tree)

    def read_and_parse_simplified(
        self,
        file_name: str,
        sheet_name: str | None = None,
    ) -> "ModelCompilerSimplified":
        """Read and parse the simplified model."""
        logger.debug("Starting to read and parse the simplified model")
        wb_data = openpyxl.load_workbook(file_name, data_only=True)
        wb_formula = openpyxl.load_workbook(file_name, data_only=False)

        sheet_data = wb_data[sheet_name] if sheet_name else wb_data.active
        sheet_formula = wb_formula[sheet_name] if sheet_name else wb_formula.active

        if sheet_formula and sheet_data:
            for row in sheet_formula.iter_rows():
                for cell in row:
                    coord = f"{sheet_formula.title}!{cell.coordinate}"
                    cached_val = sheet_data[cell.coordinate].value
                    formula = (
                        cell.value
                        if isinstance(cell.value, str) and cell.value.startswith("=")
                        else None
                    )

                    if formula and self._is_simple_formula(formula):
                        try:
                            value = self._safe_eval_formula(formula, sheet_formula)
                        except (KeyError, ValueError, TypeError, ZeroDivisionError):
                            value = cached_val
                    else:
                        value = cached_val

                    self.model.set_cell_value(coord, value)  # type: ignore

        return self
