from datetime import datetime
from enum import Enum
from openpyxl import Workbook, styles, utils
from openpyxl.worksheet.worksheet import Worksheet
from viggocorev2.common import utils as core_utils
from viggocorev2.common.subsystem import operation


def tratar_title(title: str):
    title = title.replace('Exportação de ', '')
    if len(title) > 31:
        return title[:28] + '...'
    return title


class MakeExportarXsls(operation.Update):
    def _map_column(self, column):
        aux = {
            'num': self.columns.index(column) + 1,
            'name': column.get('header_name', ''),
            'width': column.get('width', 50)
        }
        if column.get('total', False) is True:
            aux['footer'] = '=sum()'

        return aux

    def _def_header_style(self, ws, cols, style_name, row_ini=1):
        headers = [col['name'] for col in cols]
        ws.append(headers)

        header_style = styles.NamedStyle(name=style_name)
        header_style.font = styles.Font(bold=True)
        header_style.border = styles.Border(
            top=styles.Side(border_style='thin'),
            right=styles.Side(border_style='thin'),
            bottom=styles.Side(border_style='thin'),
            left=styles.Side(border_style='thin'))
        header_style.alignment = styles.Alignment(horizontal='left',
                                                  vertical='center')

        for col_num in range(len(headers)):
            col = utils.cell.get_column_letter(col_num + 1)
            ws.column_dimensions[col].width = cols[col_num]['width']

        last_col = utils.cell.get_column_letter(len(headers)) + '1'
        cell_range = ws['A1':last_col]

        for cell_rows in cell_range:
            for cell in cell_rows:
                cell.style = header_style

    def format_data_hora(self, data_hora):
        dia = data_hora.day if data_hora.day > 9 else '0' + str(data_hora.day)
        mes = (data_hora.month if data_hora.month > 9
               else '0' + str(data_hora.month))
        return '{}/{}/{}'.format(dia, mes, data_hora.year)

    def _treat_data_type(self, data, col):
        data_type = col.get('type', 'STRING')
        result = ''

        if data_type == 'STRING':
            result = data if data is not None else '-'
            if data is not None:
                result = data.name if isinstance(data, Enum) else data
        elif data_type == 'BOOLEAN':
            result = 'Sim' if data is True else 'Não'
        elif data_type == 'DATE':
            result = datetime.strftime(data, '%d/%m/%Y') \
                if data is not None \
                else '-'
        elif data_type == 'DATETIME':
            result = datetime.strftime(data, '%d/%m/%Y %H:%M') \
                if data is not None \
                else '-'
        elif data_type == 'CURRENCY' or data_type == 'FLOAT':
            decimals = col.get('decimals', None)
            if decimals is not None:
                result = core_utils.to_decimal_n(data, decimals) \
                    if data is not None \
                    else core_utils.to_decimal_n('0', decimals)
            else:
                result = core_utils.to_decimal(data) \
                    if data is not None \
                    else core_utils.to_decimal('0')
        elif data_type == 'INT':
            result = int(data) if data is not None else 0

        return result

    def _map_row(self, row):
        result = []
        for col in self.columns:
            aux = ''
            fields = col.get('field', None)
            if fields is not None:
                fields = fields.split('.')
                if len(fields) > 1:
                    aux = None if row is None else row
                    if row is not None:
                        for field in fields:
                            if aux is not None:
                                aux = aux.__getattribute__(field)
                else:
                    aux = row.__getattribute__(fields[0])

                aux = self._treat_data_type(aux, col)
            result.append(aux)

        return result

    def _make_relatorio(self):
        relatorio = []
        for e in self.entities:
            relatorio.append(self._map_row(e))

        return relatorio

    def _formatar_footer(self, col, cell, lin_ini, qtd_linhas,
                         is_to_bold=True):
        if is_to_bold:
            cell.font = styles.Font(bold=True)

        if ('footer' in col):
            if not col['footer'].startswith('='):
                cell.value = col['footer']
            else:
                formula = col['footer']
                coluna = utils.cell.get_column_letter(
                    col['num'])
                intervalo_celulas = '(' + \
                    str(coluna) + str(lin_ini) + ':' + \
                    str(coluna) + str(qtd_linhas + 1) + ')'
                formula = formula.replace(
                    '()', intervalo_celulas)
                cell.value = formula

    def _adiciona_total_linhas(self, ws, row_num, qtd_linhas):
        # adiciona linha com os totais
        cell = ws.cell(row_num, 1)
        cell.border = styles.Border(
            top=styles.Side(border_style='thin'),
            right=styles.Side(border_style='thin'),
            bottom=styles.Side(border_style='thin'),
            left=styles.Side(border_style='thin'))
        cell.font = styles.Font(bold=True)
        cell.value = f'Total de itens: {qtd_linhas}'

    def _formatar_celulas(self, ws, dados, cols, is_to_bold=True):
        qtd_linhas = len(dados)

        if qtd_linhas > 0:
            lin_ini = 2
            lin_fim = lin_ini + qtd_linhas  # A última linha é de total

            for lin in range(lin_ini, lin_fim + 1):
                # Accounting format
                for col in cols:
                    cell = ws.cell(lin, col['num'])
                    cell.border = styles.Border(
                        top=styles.Side(border_style='thin'),
                        right=styles.Side(border_style='thin'),
                        bottom=styles.Side(border_style='thin'),
                        left=styles.Side(border_style='thin'))

                    if 'fmt' in col:
                        cell.number_format = col['fmt']

                    if 'align' in col:
                        cell.alignment = styles.Alignment(
                            horizontal=col['align'])

                    if 'function' in col:
                        formula = col['function'] \
                            .replace('[row]', str(cell.row))
                        cell.value = formula

                    if (lin == lin_fim):
                        self._formatar_footer(
                            col, cell, lin_ini, qtd_linhas)
        # adiciona quantidade de linhas
        self._adiciona_total_linhas(ws, lin_fim + 1, qtd_linhas)
        return

    def _style_header(self, ws: Worksheet, cols: list) -> None:
        for i in range(1, len(cols) + 1):
            cell = ws.cell(column=i, row=1)
            cell.fill = styles.PatternFill('solid', fgColor='DEEBF7')
            cell.border = styles.Border(
                top=styles.Side(border_style='medium', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='medium', color='000000'),
                left=styles.Side(border_style='thin', color='000000'))
            cell.alignment = styles.Alignment(horizontal='center')

    def _align_horizontal_center_column(
            self, ws: Worksheet, linhas: list, coluna: int) -> None:
        for i in range(1, len(linhas) + 2):
            cell = ws.cell(column=coluna, row=i)
            cell.alignment = styles.Alignment(horizontal='center')

    def _gen_work_sheet(self, ws: Worksheet):
        # num_fmt = u'_(* #,##0.00_);[Red]_(* (#,##0.00);_(* -_0_0_);_(@'
        cols = list(map(lambda x: self._map_column(x), self.columns))
        self._def_header_style(ws, cols, self.title)

        relatorio = self._make_relatorio()
        for linha in relatorio:
            ws.append(linha)

        self._formatar_celulas(ws, relatorio, cols, False)
        self._style_header(ws, cols)
        self._align_horizontal_center_column(ws, relatorio, 2)

        return True

    def pre(self, session, entities, columns, title, file_title):
        self.entities = entities
        self.columns = columns
        self.title = title

        self.now = datetime.now()
        self.file_name = f'{file_title}-{self.now}.xlsx'
        self.download_folder = \
            self.manager.get_download_folder(entities[0].domain_id)

        return True

    def do(self, session, entities, columns, title, file_title):
        wb = Workbook()
        ws = wb.active
        ws.title = tratar_title(self.title)
        self._gen_work_sheet(ws)

        wb.save(filename=self.download_folder + '/' + self.file_name)

        return self.download_folder, self.file_name
