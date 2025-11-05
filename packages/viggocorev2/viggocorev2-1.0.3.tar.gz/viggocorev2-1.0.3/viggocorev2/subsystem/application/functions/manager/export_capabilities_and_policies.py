from viggocorev2.common import exception, utils as core_utils
from viggocorev2.subsystem.capability.resource import Capability
from viggocorev2.subsystem.role.resource import Role
from viggocorev2.subsystem.route.resource import Route
from viggocorev2.subsystem.policy.resource import Policy
from sqlalchemy import func, or_
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, styles
from viggocorev2.common.subsystem.manager_functions import make_exportar_xlsx
from datetime import datetime as datetime1


class ExportCapabilitiesAndPolicies(make_exportar_xlsx.MakeExportarXsls):

    # funções sobrescritas
    def _formatar_celulas(self, ws, dados, cols, is_to_bold=True):
        qtd_linhas = len(dados)

        if qtd_linhas > 0:
            lin_ini = 2
            lin_fim = lin_ini + qtd_linhas  # A última linha é de total

            for lin in range(lin_ini, lin_fim + 1):
                # Accounting format
                for col in cols:
                    cell = ws.cell(lin, col['num'])
                    cell.border = styles.Border(
                        top=styles.Side(border_style='thin'),
                        right=styles.Side(border_style='thin'),
                        bottom=styles.Side(border_style='thin'),
                        left=styles.Side(border_style='thin'))

                    if 'fmt' in col:
                        cell.number_format = col['fmt']

                    if 'align' in col:
                        cell.alignment = styles.Alignment(
                            horizontal=col['align'])

                    if 'function' in col:
                        formula = col['function'] \
                            .replace('[row]', str(cell.row))
                        cell.value = formula

                    if (lin == lin_fim):
                        self._formatar_footer(
                            col, cell, lin_ini, qtd_linhas)
        return

    # função para pegar o nome do arquivo
    def _get_file_name(self, **kwargs):
        file_name = kwargs.get('file_name', None)
        if file_name is not None:
            return file_name
        else:
            application = self.manager.get(id=self.id)

            file_name = f'{application.name}-{self.a_partir}.xlsx'
            return file_name

    # valida se o tipo de exportacao passado é válido
    def _validar_tipo_exportacao(self):
        tipos_permitidos = ['CAPACIDADE', 'POLITICA']
        if any(list(map(
                lambda x: x not in tipos_permitidos, self.tipo_exportacao))):
            raise exception.BadRequest(
                'Os tipos de exportação permitidos são:' +
                ', '.join(tipos_permitidos) + '.')

    # funções de buscar dados
    def _get_capabilities(self, session):
        query = session.query(Route.url, Route.method)\
            .join(Capability, Capability.route_id == Route.id)\
            .filter(Capability.application_id == self.id)\
            .filter(or_(Route.created_at >= self.a_partir,
                        Route.created_at.is_(None)))\
            .order_by(Route.url, Route.method)
        result = query.all()
        return result

    def _get_policies(self, session):
        papeis = [p.upper() for p in self.papeis]
        query = session.query(Role.name, Route.url, Route.method)\
            .join(Policy, Policy.role_id == Role.id)\
            .join(Capability, Capability.id == Policy.capability_id)\
            .join(Route, Route.id == Capability.route_id)\
            .filter(func.upper(Role.name).in_(papeis))\
            .filter(Capability.application_id == self.id)\
            .filter(or_(Route.created_at >= self.a_partir,
                        Route.created_at.is_(None)))\
            .order_by(Role.name, Route.url, Route.method)
        result = query.all()
        return result

    def _get_capabilities_and_policies(self, session):
        # inicia as listas de capacidades e políticas
        capabilities = []
        policies = []

        # se foi passada a CAPACIDADE como tipo_exportacao então vai buscar as
        # capacidades
        if 'CAPACIDADE' in self.tipo_exportacao:
            capabilities = self._get_capabilities(session=session)

        # se foi passada a POLITICA como tipo_exportacao então vai buscar as
        # políticas ligadas as capacidades desta aplicação
        if 'POLITICA' in self.tipo_exportacao:
            policies = self._get_policies(session=session)

        return (capabilities, policies)

    # funções de preenchimento do xlsx
    def _gen_work_sheet_capabilities(self, ws: Worksheet, capabilities):
        ROW_INIT = 1

        cols = [
            {'num': 1, 'name': 'URL', 'width': 50},
            {'num': 2, 'name': 'METHOD', 'width': 25}
        ]

        self._def_header_style(ws, cols, 'export_capabilities',
                               row_ini=ROW_INIT)

        for linha in capabilities:
            ws.append(list(linha))

        self._formatar_celulas(ws, capabilities, cols)
        self._style_header(ws, cols)
        self._align_horizontal_center_column(ws, capabilities, 2)

        return True

    def _gen_work_sheet_policies(self, ws: Worksheet, policies):
        ROW_INIT = 1

        cols = [
            {'num': 1, 'name': 'PAPEL', 'width': 25},
            {'num': 2, 'name': 'URL', 'width': 50},
            {'num': 3, 'name': 'METHOD', 'width': 25}
        ]

        self._def_header_style(ws, cols, 'export_policies',
                               row_ini=ROW_INIT)

        for linha in policies:
            ws.append(list(linha))

        self._formatar_celulas(ws, policies, cols)
        self._style_header(ws, cols)
        self._align_horizontal_center_column(ws, policies, 2)

        return True

    def pre(self, id, **kwargs):
        self.id = id
        if self.id is None:
            raise exception.BadRequest(
                'É obrigatório passar o id na rota da requisição.')

        self.tipo_exportacao = kwargs.pop('tipo_exportacao', [])
        if not self.tipo_exportacao:
            raise exception.BadRequest(
                'É obrigatório informar o tipo_exportacao.')

        self._validar_tipo_exportacao()

        self.papeis = kwargs.pop('papeis', [])
        if 'POLITICA' in self.tipo_exportacao and not self.papeis:
            raise exception.BadRequest(
                'É obrigatório informar os papéis caso você ' +
                'queira exportar as políticas também.')

        self.a_partir = kwargs.pop('a_partir', None)
        if self.a_partir is None:
            raise exception.BadRequest(
                'É obrigatório informar o campo "a_partir" para filtrar ' +
                'as informações que foram geradas posterior a esta data.')
        else:
            try:
                self.a_partir = datetime1.strptime(
                    self.a_partir.replace(' ', '+'), '%Y-%m-%d%z')
            except Exception:
                raise exception.BadRequest(
                    'É obrigatório passar a data no formato "yyyy-mm-dd-tz."')

        return True

    def do(self, session, **kwargs):
        capabilities, policies = self._get_capabilities_and_policies(
            session=session)

        if len(capabilities) == 0 and len(policies) == 0:
            raise exception.BadRequest(
                'Não existe dados para essa filtragem.')

        self.file_name = self._get_file_name(**kwargs)
        self.download_folder = core_utils.get_upload_folder()

        wb = Workbook()

        # Aba de capacidades
        ws_cap = wb.active  # Usa a aba ativa pra primeira
        ws_cap.title = "CAPACIDADES"
        self._gen_work_sheet_capabilities(ws_cap, capabilities)

        # Aba de políticas
        ws_pol = wb.create_sheet(title="POLITICAS")
        self._gen_work_sheet_policies(ws_pol, policies)

        wb.save(filename=self.download_folder + '/' + self.file_name)

        return self.download_folder, self.file_name
