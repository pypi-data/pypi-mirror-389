import openpyxl
from viggocorev2.common.subsystem import operation
from viggocorev2.subsystem.capability.resource import Capability
from viggocorev2.subsystem.role.resource import Role
from viggocorev2.subsystem.route.resource import Route
from sqlalchemy import func


class ImportCapabilitiesAndPolicies(operation.Operation):
    """
    Importa uma lista de fabricantes via planilha xlsx
    """

    # função genérica para pegar os dados da planilha
    def _get_sheet_data(self, sheet) -> list:
        """
        Lê os dados da planilha e os converte em uma lista de dicionários

        :param sheet: tabela de dados
        :return: lista de dicionários
        :rtype: list
        """

        # cria lista de retorno
        result = []
        # seta linha inicial como 2, a primeira linha é dos nomes das colunas
        row = 2

        # itera sobre os dados da tabela até que a primeira coluna da linha
        # atual esteja vazia, significando o final da tabela
        while sheet.cell(row=row, column=1).value is not None:
            # define dicionário auxiliar para objeto sendo iterado
            aux = {}
            # inicia na coluna 1
            col = 1

            # itera sobre as colunas da linha atual até que o valor observado
            # seja nulo
            while sheet.cell(row=1, column=col).value is not None:
                # atualiza dicionário auxiliar com valor observado na iteração
                # e chave sendo o nome da coluna
                aux.update({
                    sheet.cell(row=1, column=col).value:
                    sheet.cell(row=row, column=col).value
                })
                col += 1

            # appende dicionário auxiliar à lista de retorno
            result.append(aux)
            row += 1

        return result

    def _get_route_ids(self, session, sheet_data):
        filtro = [
            f'{data.get("URL", None)} - {data.get("METHOD", None)}'
            for data in sheet_data]
        query = session.query(Route.id)\
            .filter(func.concat(Route.url, ' - ', Route.method).in_(filtro))
        result = query.all()
        result = [r[0] for r in result]
        return result

    def _get_roles(self, session, sheet_data):
        role_names = set([data.get("PAPEL", 'NULL').upper()
                          for data in sheet_data])
        query = session.query(Role)\
            .filter(func.upper(Role.name).in_(role_names))
        result = query.all()
        return {r.name: r.id for r in result}

    def _get_role_id_cabilities(self, session, sheet_data):
        role_id_cabilities = {}
        roles = self._get_roles(session, sheet_data)
        for role_name, role_id in roles.items():
            sheet_data_filtred = list(
                filter(lambda x: x.get('PAPEL', '') == role_name, sheet_data))
            filtro = [
                f'{data.get("URL", None)} - {data.get("METHOD", None)}'
                for data in sheet_data_filtred]
            query = session.query(Capability.id)\
                .join(Route, Route.id == Capability.route_id)\
                .filter(Capability.application_id == self.id)\
                .filter(func.concat(Route.url, ' - ', Route.method).in_(filtro))
            result = query.all()
            role_id_cabilities[role_id] = [r[0] for r in result]

        return role_id_cabilities

    # função que monta os dicionários das capacidades para serem salvas
    def _map_capabilities_dict(self, session, sheet_data):
        capabilities_dict = []
        route_ids = self._get_route_ids(session=session, sheet_data=sheet_data)
        for route_id in route_ids:
            capabilities_dict.append({
                "application_id": self.id,
                "route_id": route_id
            })
        return capabilities_dict

    # função que salva as capacidades
    def _save_capabilities(self, session, wb_obj):
        # valida se a planilha possui a tabela apropriada para a importação
        # dados e lança uma BadRequest caso a tabela não seja encontrada
        sheet_obj = list(filter(
            lambda x: x.title == 'CAPACIDADES', wb_obj.worksheets))
        if len(sheet_obj) == 0:
            return

        # converte tabela de objetos em lista de dicionários
        sheet_obj = sheet_obj[0]
        sheet_data = self._get_sheet_data(sheet_obj)
        # mapeia a lista de dicionários em objetos na estrutura do DB
        capabilities_dict = self._map_capabilities_dict(session, sheet_data)

        # salva objetos e suas dependências no DB
        for cap_dict in capabilities_dict:
            try:
                self.manager.api.capabilities().create(
                    session=session, **cap_dict)
                session.commit()
                # self.driver.update(session=session)
                self.capabilities_sucessos += 1
            except Exception as e:
                print(str(e))
                self.capabilities_falhas += 1
                session.rollback()

    # função que monta os dicionários das capacidades para serem salvas
    def _map_policies_dict(self, session, sheet_data):
        policies_dict = []
        role_cabilities = self._get_role_id_cabilities(session, sheet_data)
        for role_id in role_cabilities.keys():
            cability_ids = role_cabilities.get(role_id, [])
            for c_id in cability_ids:
                policies_dict.append({
                    "role_id": role_id,
                    "capability_id": c_id
                })
        return policies_dict

    # função que salva as politicas
    def _save_policies(self, session, wb_obj):
        # valida se a planilha possui a tabela apropriada para a importação
        # dados e lança uma BadRequest caso a tabela não seja encontrada
        sheet_obj = list(filter(
            lambda x: x.title == 'POLITICAS', wb_obj.worksheets))
        if len(sheet_obj) == 0:
            return

        # converte tabela de objetos em lista de dicionários
        sheet_obj = sheet_obj[0]
        sheet_data = self._get_sheet_data(sheet_obj)
        # mapeia a lista de dicionários em objetos na estrutura do DB
        policies_dict = self._map_policies_dict(session, sheet_data)

        # salva objetos e suas dependências no DB
        for policy_dict in policies_dict:
            try:
                self.manager.api.policies().create(
                    session=session, **policy_dict)
                self.policies_sucessos += 1
                session.commit()
            except Exception as e:
                print(str(e))
                self.policies_falhas += 1
                session.rollback()

    def pre(self, id, **kwargs):
        self.id = id
        # recupera a planilha à partir do dicionário kwargs
        self.file = kwargs.get('file', None)
        # recupera domain_id à partir do dicionário kwargs
        self.domain_id = kwargs.get('domain_id', None)

        return True

    def do(self, session, **kwargs):
        self.capabilities_sucessos = 0
        self.capabilities_falhas = 0
        self.policies_sucessos = 0
        self.policies_falhas = 0

        # inicia o arquivo xlsx
        wb_obj = openpyxl.load_workbook(self.file)

        self._save_capabilities(session=session, wb_obj=wb_obj)
        self._save_policies(session=session, wb_obj=wb_obj)

        print(f"""
        capabilities_sucessos: {self.capabilities_sucessos}
        capabilities_falhas: {self.capabilities_falhas}
        policies_sucessos: {self.policies_sucessos}
        policies_falhas: {self.policies_falhas}
        """)

        return super().do(**kwargs)
