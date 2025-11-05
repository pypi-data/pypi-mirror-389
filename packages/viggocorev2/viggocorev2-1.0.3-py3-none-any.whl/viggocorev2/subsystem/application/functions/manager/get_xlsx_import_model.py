from viggocorev2.common import utils as core_utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, styles
from viggocorev2.common.subsystem.manager_functions import make_exportar_xlsx


class GetXlsxImportModel(make_exportar_xlsx.MakeExportarXsls):

    # funções sobrescritas
    def _formatar_celulas(self, ws, dados, cols, is_to_bold=True):
        qtd_linhas = len(dados)

        if qtd_linhas > 0:
            lin_ini = 2
            lin_fim = lin_ini + qtd_linhas  # A última linha é de total

            for lin in range(lin_ini, lin_fim + 1):
                # Accounting format
                for col in cols:
                    cell = ws.cell(lin, col['num'])
                    cell.border = styles.Border(
                        top=styles.Side(border_style='thin'),
                        right=styles.Side(border_style='thin'),
                        bottom=styles.Side(border_style='thin'),
                        left=styles.Side(border_style='thin'))

                    if 'fmt' in col:
                        cell.number_format = col['fmt']

                    if 'align' in col:
                        cell.alignment = styles.Alignment(
                            horizontal=col['align'])

                    if 'function' in col:
                        formula = col['function'] \
                            .replace('[row]', str(cell.row))
                        cell.value = formula

                    if (lin == lin_fim):
                        self._formatar_footer(
                            col, cell, lin_ini, qtd_linhas)
        return

    # função para pegar o nome do arquivo
    def _get_file_name(self, **kwargs):
        return 'import-model-capabilities-and-policies.xlsx'

    # funções de preenchimento do xlsx
    def _gen_work_sheet_capabilities(self, ws: Worksheet, capabilities):
        ROW_INIT = 1

        cols = [
            {'num': 1, 'name': 'URL', 'width': 50},
            {'num': 2, 'name': 'METHOD', 'width': 25}
        ]

        self._def_header_style(ws, cols, 'export_capabilities',
                               row_ini=ROW_INIT)

        for linha in capabilities:
            ws.append(list(linha))

        self._formatar_celulas(ws, capabilities, cols)
        self._style_header(ws, cols)
        self._align_horizontal_center_column(ws, capabilities, 2)

        return True

    def _gen_work_sheet_policies(self, ws: Worksheet, policies):
        ROW_INIT = 1

        cols = [
            {'num': 1, 'name': 'PAPEL', 'width': 25},
            {'num': 2, 'name': 'URL', 'width': 50},
            {'num': 3, 'name': 'METHOD', 'width': 25}
        ]

        self._def_header_style(ws, cols, 'export_policies',
                               row_ini=ROW_INIT)

        for linha in policies:
            ws.append(list(linha))

        self._formatar_celulas(ws, policies, cols)
        self._style_header(ws, cols)
        self._align_horizontal_center_column(ws, policies, 2)

        return True

    def pre(self, **kwargs):
        return True

    def do(self, session, **kwargs):
        capabilities = []
        policies = []

        self.file_name = self._get_file_name(**kwargs)
        self.download_folder = core_utils.get_upload_folder()

        wb = Workbook()

        # Aba de capacidades
        ws_cap = wb.active  # Usa a aba ativa pra primeira
        ws_cap.title = "CAPACIDADES"
        self._gen_work_sheet_capabilities(ws_cap, capabilities)

        # Aba de políticas
        ws_pol = wb.create_sheet(title="POLITICAS")
        self._gen_work_sheet_policies(ws_pol, policies)

        wb.save(filename=self.download_folder + '/' + self.file_name)

        return self.download_folder, self.file_name
