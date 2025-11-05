"""
Enhanced data export and visualization tools for researchers.
Provides multiple export formats and publication-ready visualizations.
"""

import json
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Union
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from pandas import DataFrame


class DataExporter:
    """
    Comprehensive data export utility for microsimulation results.
    """
    
    def __init__(self, data: pd.DataFrame, scenario_name: str = "Analysis"):
        self.data = data
        self.scenario_name = scenario_name
        
    def export_csv(self, output_path: str, include_metadata: bool = True) -> str:
        """Export data to CSV with optional metadata header."""
        
        if include_metadata:
            # Create metadata header
            metadata = [
                f"# NZ Tax Microsimulation Model Export",
                f"# Scenario: {self.scenario_name}",
                f"# Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"# Records: {len(self.data)}",
                f"# Variables: {len(self.data.columns)}",
                "#",
            ]
            
            # Write metadata and data
            with open(output_path, 'w') as f:
                f.write('\n'.join(metadata) + '\n')
                self.data.to_csv(f, index=False)
        else:
            self.data.to_csv(output_path, index=False)
        
        return output_path
    
    def export_excel(self, output_path: str, include_summary: bool = True) -> str:
        """Export data to Excel with formatted sheets and optional summary."""
        
        wb = Workbook()
        
        # Main data sheet
        ws_data = wb.active
        ws_data.title = "Simulation Results"
        
        # Write headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, column_title in enumerate(self.data.columns, 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_num, row in enumerate(self.data.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws_data.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Summary sheet
        if include_summary and "disposable_income" in self.data.columns:
            ws_summary = wb.create_sheet("Summary Statistics")
            
            summary_stats = [
                ["Metric", "Value"],
                ["Total Records", len(self.data)],
                ["Mean Disposable Income", f"${self.data['disposable_income'].mean():.2f}"],
                ["Median Disposable Income", f"${self.data['disposable_income'].median():.2f}"],
                ["Standard Deviation", f"${self.data['disposable_income'].std():.2f}"],
                ["Minimum Income", f"${self.data['disposable_income'].min():.2f}"],
                ["Maximum Income", f"${self.data['disposable_income'].max():.2f}"],
            ]
            
            # Add tax revenue if available
            if "tax_liability" in self.data.columns:
                total_tax = self.data["tax_liability"].sum()
                summary_stats.extend([
                    ["Total Tax Revenue", f"${total_tax:,.2f}"],
                    ["Average Tax Rate", f"{(total_tax / self.data['disposable_income'].sum()) * 100:.2f}%"]
                ])
            
            for row_num, (metric, value) in enumerate(summary_stats, 1):
                ws_summary.cell(row=row_num, column=1, value=metric)
                ws_summary.cell(row=row_num, column=2, value=value)
            
            # Format summary headers
            ws_summary.cell(row=1, column=1).font = header_font
            ws_summary.cell(row=1, column=2).font = header_font
            ws_summary.cell(row=1, column=1).fill = header_fill
            ws_summary.cell(row=1, column=2).fill = header_fill
        
        wb.save(output_path)
        return output_path
    
    def export_json(self, output_path: str, format_type: str = "records") -> str:
        """Export data to JSON in various formats."""
        
        # Define a safe orient by using explicit method calls to avoid pandas type issues
        if format_type == "dict":
            data_dict = self.data.to_dict(orient="dict")
        elif format_type == "list":
            data_dict = self.data.to_dict(orient="list")
        elif format_type == "series":
            data_dict = self.data.to_dict(orient="series")
        elif format_type == "split":
            data_dict = self.data.to_dict(orient="split")
        elif format_type == "tight":
            data_dict = self.data.to_dict(orient="tight")
        elif format_type == "index":
            data_dict = self.data.to_dict(orient="index")
        else:  # defaults to "records"
            data_dict = self.data.to_dict(orient="records")

        export_data = {
            "metadata": {
                "scenario_name": self.scenario_name,
                "generated_at": pd.Timestamp.now().isoformat(),
                "record_count": len(self.data),
                "variables": list(self.data.columns)
            },
            "data": data_dict
        }
        
        with open(output_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return output_path
    
    def export_stata(self, output_path: str) -> str:
        """Export data to Stata format (.dta)."""
        
        # Convert any object columns to strings for Stata compatibility
        stata_data = self.data.copy()
        for col in stata_data.select_dtypes(include=['object']).columns:
            stata_data[col] = stata_data[col].astype(str)
        
        stata_data.to_stata(output_path, write_index=False, version=117)
        return output_path
    
    def export_spss(self, output_path: str) -> str:
        """Export data to SPSS format (.sav)."""
        try:
            import pyreadstat
            pyreadstat.write_sav(self.data, output_path)
            return output_path
        except ImportError:
            raise ImportError("pyreadstat required for SPSS export. Install with: pip install pyreadstat")
    
    def export_r_data(self, output_path: str) -> str:
        """Export data to R format (.RData)."""
        try:
            import pyreadr
            pyreadr.write_rdata(output_path, self.data, df_name=self.scenario_name.replace(" ", "_"))
            return output_path
        except ImportError:
            raise ImportError("pyreadr required for R export. Install with: pip install pyreadr")


class PolicyVisualization:
    """
    Create publication-ready visualizations for policy analysis.
    """
    
    def __init__(self, baseline_data: pd.DataFrame, scenario_data: Optional[pd.DataFrame] = None):
        self.baseline_data = baseline_data
        self.scenario_data = scenario_data
        
        # Set up matplotlib and seaborn defaults
        plt.style.use('seaborn-v0_8-whitegrid')
        sns.set_palette("Set2")
    
    def create_income_distribution_plot(self, output_path: str, 
                                      bins: int = 50, 
                                      max_income: Optional[float] = None) -> str:
        """Create income distribution comparison plot."""
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Filter extreme values if specified
        baseline_income = self.baseline_data["disposable_income"]
        if max_income:
            baseline_income = baseline_income[baseline_income <= max_income]
        
        ax.hist(baseline_income, bins=bins, alpha=0.7, label="Baseline", density=True)
        
        if self.scenario_data is not None:
            scenario_income = self.scenario_data["disposable_income"]
            if max_income:
                scenario_income = scenario_income[scenario_income <= max_income]
            ax.hist(scenario_income, bins=bins, alpha=0.7, label="Policy Scenario", density=True)
        
        ax.set_xlabel("Disposable Income ($)", fontsize=12)
        ax.set_ylabel("Density", fontsize=12)
        ax.set_title("Income Distribution Comparison", fontsize=14, fontweight='bold')
        ax.legend(fontsize=11)
        
        # Format x-axis as currency
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    def create_decile_impact_chart(self, output_path: str) -> str:
        """Create decile impact bar chart."""
        
        if self.scenario_data is None:
            raise ValueError("Scenario data required for impact analysis")
        
        # Calculate decile impacts
        baseline_deciles = pd.qcut(self.baseline_data["disposable_income"], 10, labels=False)
        decile_impacts = []
        
        for decile in range(10):
            decile_mask = baseline_deciles == decile
            baseline_subset = self.baseline_data[decile_mask]["disposable_income"]
            scenario_subset = self.scenario_data[decile_mask]["disposable_income"]
            
            mean_impact = scenario_subset.mean() - baseline_subset.mean()
            decile_impacts.append(mean_impact)
        
        # Create chart
        fig, ax = plt.subplots(figsize=(12, 8))
        
        deciles = range(1, 11)
        colors = ['red' if x < 0 else 'green' for x in decile_impacts]
        
        bars = ax.bar(deciles, decile_impacts, color=colors, alpha=0.7)
        
        # Add value labels on bars
        for bar, impact in zip(bars, decile_impacts):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + (50 if height > 0 else -150),
                   f'${impact:.0f}', ha='center', va='bottom' if height > 0 else 'top',
                   fontweight='bold')
        
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        ax.set_xlabel("Income Decile", fontsize=12)
        ax.set_ylabel("Average Income Change ($)", fontsize=12)
        ax.set_title("Policy Impact by Income Decile", fontsize=14, fontweight='bold')
        
        # Format y-axis as currency
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    def create_lorenz_curve(self, output_path: str) -> str:
        """Create Lorenz curve for income distribution."""
        
        def lorenz_curve(income_series):
            """Calculate Lorenz curve coordinates."""
            sorted_income = np.sort(income_series)
            cumulative_income = np.cumsum(sorted_income)
            total_income = cumulative_income[-1]
            n = len(sorted_income)
            
            # Lorenz curve points
            lorenz_x = np.arange(n + 1) / n
            lorenz_y = np.concatenate([[0], cumulative_income]) / total_income
            
            return lorenz_x, lorenz_y
        
        fig, ax = plt.subplots(figsize=(10, 10))
        
        # Baseline Lorenz curve
        baseline_x, baseline_y = lorenz_curve(self.baseline_data["disposable_income"])
        ax.plot(baseline_x, baseline_y, label="Baseline", linewidth=2)
        
        # Scenario Lorenz curve
        if self.scenario_data is not None:
            scenario_x, scenario_y = lorenz_curve(self.scenario_data["disposable_income"])
            ax.plot(scenario_x, scenario_y, label="Policy Scenario", linewidth=2)
        
        # Line of equality
        ax.plot([0, 1], [0, 1], 'k--', alpha=0.5, label="Perfect Equality")
        
        ax.set_xlabel("Cumulative Population Share", fontsize=12)
        ax.set_ylabel("Cumulative Income Share", fontsize=12)
        ax.set_title("Lorenz Curve - Income Distribution", fontsize=14, fontweight='bold')
        ax.legend(fontsize=11)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    def create_winners_losers_plot(self, output_path: str, 
                                 change_threshold: float = 50) -> str:
        """Create winners and losers visualization."""
        
        if self.scenario_data is None:
            raise ValueError("Scenario data required for winners/losers analysis")
        
        income_change = self.scenario_data["disposable_income"] - self.baseline_data["disposable_income"]
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
        
        # Distribution of income changes
        ax1.hist(income_change, bins=100, alpha=0.7, color='steelblue')
        ax1.axvline(x=0, color='red', linestyle='--', linewidth=2, label='No Change')
        ax1.axvline(x=change_threshold, color='green', linestyle='--', alpha=0.7, label=f'Winners (>${change_threshold})')
        ax1.axvline(x=-change_threshold, color='orange', linestyle='--', alpha=0.7, label=f'Losers (<-${change_threshold})')
        
        ax1.set_xlabel("Income Change ($)", fontsize=12)
        ax1.set_ylabel("Number of Individuals", fontsize=12)
        ax1.set_title("Distribution of Income Changes", fontsize=14, fontweight='bold')
        ax1.legend(fontsize=10)
        
        # Winners/losers/unchanged pie chart
        winners = (income_change > change_threshold).sum()
        losers = (income_change < -change_threshold).sum()
        unchanged = len(income_change) - winners - losers
        
        sizes = [winners, losers, unchanged]
        labels = [f'Winners\n({winners:,})', f'Losers\n({losers:,})', f'Unchanged\n({unchanged:,})']
        colors = ['green', 'red', 'gray']
        
        ax2.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax2.set_title("Population Impact Summary", fontsize=14, fontweight='bold')
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return output_path


def create_research_export_package(baseline_data: pd.DataFrame,
                                 scenario_data: pd.DataFrame,
                                 scenario_name: str,
                                 output_dir: str = "research_package") -> Dict[str, str]:
    """
    Create a comprehensive export package for researchers.
    
    Returns a dictionary of created file paths.
    """
    
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True, parents=True)
    
    # Initialize exporters
    baseline_exporter = DataExporter(baseline_data, "Baseline")
    scenario_exporter = DataExporter(scenario_data, scenario_name)
    visualizer = PolicyVisualization(baseline_data, scenario_data)
    
    created_files = {}
    
    # Data exports
    created_files["baseline_csv"] = baseline_exporter.export_csv(
        str(output_path / "baseline_data.csv")
    )
    created_files["scenario_csv"] = scenario_exporter.export_csv(
        str(output_path / f"{scenario_name.lower().replace(' ', '_')}_data.csv")
    )
    created_files["baseline_excel"] = baseline_exporter.export_excel(
        str(output_path / "baseline_data.xlsx")
    )
    created_files["scenario_excel"] = scenario_exporter.export_excel(
        str(output_path / f"{scenario_name.lower().replace(' ', '_')}_data.xlsx")
    )
    
    # Visualizations
    created_files["income_distribution"] = visualizer.create_income_distribution_plot(
        str(output_path / "income_distribution.png")
    )
    created_files["decile_impact"] = visualizer.create_decile_impact_chart(
        str(output_path / "decile_impact.png")
    )
    created_files["lorenz_curve"] = visualizer.create_lorenz_curve(
        str(output_path / "lorenz_curve.png")
    )
    created_files["winners_losers"] = visualizer.create_winners_losers_plot(
        str(output_path / "winners_losers.png")
    )
    
    # Create README file
    readme_content = f"""
# Research Data Package: {scenario_name}

This package contains microsimulation results and visualizations for policy analysis.

## Files Included

### Data Files
- `baseline_data.csv/xlsx` - Baseline policy simulation results
- `{scenario_name.lower().replace(' ', '_')}_data.csv/xlsx` - Policy scenario results

### Visualizations
- `income_distribution.png` - Comparison of income distributions
- `decile_impact.png` - Policy impact by income decile
- `lorenz_curve.png` - Income inequality visualization
- `winners_losers.png` - Population impact analysis

## Data Dictionary

Common variables included:
- `disposable_income` - Final disposable income after taxes and transfers
- `tax_liability` - Total income tax liability
- `employment_income` - Income from employment
- `self_employment_income` - Income from self-employment
- Various WFF and benefit columns (if applicable)

## Citation

Please cite the NZ Tax Microsimulation Model when using this data:
[Add appropriate citation here]

Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
    
    readme_path = output_path / "README.md"
    with open(readme_path, 'w') as f:
        f.write(readme_content)
    
    created_files["readme"] = str(readme_path)
    
    return created_files