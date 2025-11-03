"""Data source parsers for CSV, XLSX, JSON, and XML files."""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Generator, List, Optional
import json
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook


class DataSourceParser(ABC):
    """Abstract base class for data source parsers."""

    @abstractmethod
    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse data source and yield DataFrames in chunks."""
        pass

    @abstractmethod
    def get_column_names(self) -> List[str]:
        """Get list of column names."""
        pass


class CSVParser(DataSourceParser):
    """Parser for CSV files."""

    def __init__(
        self,
        file_path: Path,
        delimiter: str = ",",
        has_header: bool = True,
        encoding: str = "utf-8",
    ):
        """Initialize CSV parser.
        
        Args:
            file_path: Path to CSV file
            delimiter: Column delimiter
            has_header: Whether first row is header
            encoding: File encoding
        """
        self.file_path = file_path
        self.delimiter = delimiter
        self.has_header = has_header
        self.encoding = encoding
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse CSV file and yield DataFrames.
        
        Args:
            chunk_size: Number of rows per chunk. If None, load entire file.
            
        Yields:
            DataFrames containing parsed data
        """
        header = 0 if self.has_header else None
        
        if chunk_size:
            # Stream large files in chunks
            for chunk in pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                header=header,
                encoding=self.encoding,
                chunksize=chunk_size,
                keep_default_na=False,  # Preserve empty strings vs NaN
                na_values=[""],
            ):
                yield chunk
        else:
            # Load entire file
            df = pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                header=header,
                encoding=self.encoding,
                keep_default_na=False,
                na_values=[""],
            )
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from CSV."""
        if not self.has_header:
            # Read first row to determine number of columns
            df_sample = pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                nrows=1,
                header=None,
            )
            return [f"Column_{i}" for i in range(len(df_sample.columns))]
        
        # Read just the header
        df_sample = pd.read_csv(
            self.file_path,
            delimiter=self.delimiter,
            nrows=0,
        )
        return df_sample.columns.tolist()


class XLSXParser(DataSourceParser):
    """Parser for XLSX files."""

    def __init__(
        self,
        file_path: Path,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ):
        """Initialize XLSX parser.
        
        Args:
            file_path: Path to XLSX file
            sheet_name: Name of sheet to read. If None, reads first sheet.
            has_header: Whether first row is header
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.has_header = has_header
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"XLSX file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse XLSX file and yield DataFrames.
        
        Args:
            chunk_size: Number of rows per chunk. If None, load entire sheet.
            
        Yields:
            DataFrames containing parsed data
        """
        header = 0 if self.has_header else None
        
        if chunk_size:
            # For XLSX, we need to load and chunk manually
            # openpyxl doesn't support streaming reads efficiently
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                header=header,
                keep_default_na=False,
                na_values=[""],
            )
            
            # Yield in chunks
            for i in range(0, len(df), chunk_size):
                yield df.iloc[i : i + chunk_size]
        else:
            # Load entire sheet
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                header=header,
                keep_default_na=False,
                na_values=[""],
            )
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from XLSX."""
        if not self.has_header:
            # Read first row to determine columns
            df_sample = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                nrows=1,
                header=None,
            )
            return [f"Column_{i}" for i in range(len(df_sample.columns))]
        
        # Read just the header
        df_sample = pd.read_excel(
            self.file_path,
            sheet_name=self.sheet_name or 0,
            nrows=0,
        )
        return df_sample.columns.tolist()

    def list_sheets(self) -> List[str]:
        """List all sheet names in the workbook."""
        wb = load_workbook(self.file_path, read_only=True)
        return wb.sheetnames


class JSONParser(DataSourceParser):
    """Parser for JSON files."""

    def __init__(self, file_path: Path):
        """Initialize JSON parser.

        Args:
            file_path: Path to JSON file
        """
        self.file_path = file_path

        if not self.file_path.exists():
            raise FileNotFoundError(f"JSON file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse JSON file and yield DataFrames.

        Args:
            chunk_size: Not used, as JSON is loaded entirely

        Yields:
            DataFrames containing parsed data
        """
        with open(self.file_path, 'r', encoding='utf-8') as f:
            if self.file_path.suffix.lower() == '.jsonl':
                # JSON Lines format
                data = []
                for line in f:
                    if line.strip():
                        data.append(json.loads(line))
            else:
                # Standard JSON
                content = json.load(f)
                # Handle different JSON structures
                if isinstance(content, list):
                    data = content
                elif isinstance(content, dict):
                    # Look for array field or treat as single record
                    array_fields = [k for k, v in content.items() if isinstance(v, list)]
                    if array_fields:
                        data = content[array_fields[0]]  # Use first array field
                    else:
                        data = [content]  # Single record
                else:
                    raise ValueError("JSON must contain object or array")

        # Expand arrays and flatten nested structures
        expanded_data = []
        for record in data:
            expanded_records = self._expand_arrays(record)
            for expanded_record in expanded_records:
                expanded_data.append(expanded_record)

        # Use json_normalize to flatten the expanded data
        df = pd.json_normalize(expanded_data)
        yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from JSON."""
        # To ensure consistency, we'll parse a small sample and get the actual column names
        # This ensures get_column_names() returns the same expanded columns as parse()
        try:
            for df in self.parse():
                return df.columns.tolist()
        except Exception:
            # Fallback to original method if parsing fails
            with open(self.file_path, 'r', encoding='utf-8') as f:
                if self.file_path.suffix.lower() == '.jsonl':
                    first_line = f.readline()
                    json_obj = json.loads(first_line)
                else:
                    content = json.load(f)
                    if isinstance(content, list):
                        json_obj = content[0] if content else {}
                    elif isinstance(content, dict):
                        array_fields = [k for k, v in content.items() if isinstance(v, list)]
                        if array_fields:
                            json_obj = content[array_fields[0]][0] if content[array_fields[0]] else {}
                        else:
                            json_obj = content
                    else:
                        json_obj = {}

                if json_obj:
                    df = pd.json_normalize([json_obj])
                    return df.columns.tolist()
                else:
                    return []

    def _expand_arrays(self, obj):
        """Expand arrays in JSON objects to create separate records for each array item."""
        if not isinstance(obj, dict):
            return [obj]

        # Find all array fields in the object
        array_fields = []

        def find_arrays(data, path=""):
            if isinstance(data, dict):
                for key, value in data.items():
                    current_path = f"{path}.{key}" if path else key
                    if isinstance(value, list) and value and isinstance(value[0], dict):
                        # Found an array of objects
                        array_fields.append((current_path, value))
                    elif isinstance(value, dict):
                        find_arrays(value, current_path)

        find_arrays(obj)

        if not array_fields:
            # No arrays found, return original object
            return [obj]

        # Handle the first array field found
        array_path, array_items = array_fields[0]

        expanded_records = []
        for item in array_items:
            # Create a new record for each array item
            expanded_record = self._deep_copy_object(obj)

            # Replace the array with the single item
            self._set_nested_value(expanded_record, array_path, item)

            expanded_records.append(expanded_record)

        return expanded_records

    def _deep_copy_object(self, obj):
        """Create a deep copy of an object."""
        if isinstance(obj, dict):
            return {key: self._deep_copy_object(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._deep_copy_object(item) for item in obj]
        else:
            return obj

    def _set_nested_value(self, obj, path, value):
        """Set a value in a nested object using dot notation path."""
        parts = path.split('.')
        current = obj

        # Navigate to the parent of the target
        for part in parts[:-1]:
            if part not in current:
                current[part] = {}
            current = current[part]

        # Set the final value
        current[parts[-1]] = value


class XMLParser(DataSourceParser):
    """Parser for XML files."""

    def __init__(self, file_path: Path):
        """Initialize XML parser.

        Args:
            file_path: Path to XML file
        """
        self.file_path = file_path

        if not self.file_path.exists():
            raise FileNotFoundError(f"XML file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse XML file and yield DataFrames.

        Args:
            chunk_size: Not used, as XML is loaded entirely

        Yields:
            DataFrames containing parsed data
        """
        # Load entire file
        tree = ET.parse(self.file_path)
        root = tree.getroot()

        # Convert XML to dict
        def xml_to_dict(element):
            return {
                child.tag: xml_to_dict(child) if len(child) > 0 else child.text
                for child in element
            }

        data = [xml_to_dict(child) for child in root]

        # Convert to DataFrame
        df = pd.json_normalize(data)
        yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from XML."""
        tree = ET.parse(self.file_path)
        root = tree.getroot()

        # Convert XML to dict
        def xml_to_dict(element):
            return {
                child.tag: xml_to_dict(child) if len(child) > 0 else child.text
                for child in element
            }

        # Get column names from first record
        first_record = xml_to_dict(root[0])
        return list(first_record.keys())


def create_parser(
    file_path: Path,
    delimiter: str = ",",
    has_header: bool = True,
    sheet_name: Optional[str] = None,
) -> DataSourceParser:
    """Factory function to create appropriate parser based on file extension.
    
    Args:
        file_path: Path to data file
        delimiter: CSV delimiter
        has_header: Whether file has header row
        sheet_name: XLSX sheet name (for Excel files)
        
    Returns:
        Appropriate parser instance
        
    Raises:
        ValueError: If file format is not supported
    """
    suffix = file_path.suffix.lower()
    
    if suffix == ".csv":
        return CSVParser(file_path, delimiter=delimiter, has_header=has_header)
    elif suffix in [".xlsx", ".xls"]:
        return XLSXParser(file_path, sheet_name=sheet_name, has_header=has_header)
    elif suffix in [".json", ".jsonl"]:
        return JSONParser(file_path)
    elif suffix == ".xml":
        return XMLParser(file_path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}")
