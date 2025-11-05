# pylint: disable=line-too-long,invalid-name,consider-using-f-string
"""Module providing validation classes."""

import logging
import os
import sys
import re
import time
import itertools
import pathlib
from shutil import copyfile
import json
import jsonschema
import pandas as pd
import tempfile
import importlib
import inspect
from tqdm import tqdm
from openpyxl import reader, load_workbook
from frictionless import Package, Dialect, Resource, Schema, Checklist, Check, Field, Pipeline, Plugin, fields, settings, formats, describe, system
from ._version import __version__
from .report import ValidationReport
from .functions import setDescriptorValueDynamicString
from .functions import setDescriptorValueDynamicBoolean
from .functions import setDescriptorValueDynamicInteger
from .functions import setDescriptorValueDynamicNumber
from .functions import setDescriptorValueDynamicList

class Validate:

    """
    xlsx validator
    
    Parameters
    ----------------------
    filename : str
        location of the xlsx file that should be validated

    configuration: str
        location of the configuration file

    **kwargs
        additional optional arguments, valid keyword arguments are:

        cli: bool
            enables progressbar in output for cli usage; by default disabled
        statusRegister: dict
            provide a shared dictionary to update validation status of processes
        statusIdentifier: str
            used as key in the statusRegister to update validation status current process
            if missing, validation will terminate
        create: str
            create a configuration for the xlsx file, and store it on the provided location
        webinterfaceData: dict
            used to provide webinterface data to override default configuration settings
        name: str
            name used in validation; by default basename of xlsx file will be used
        updateFile: bool
            enables updating of xlsx file when automatically solving problems; by default disabled
        
        
    """

    ERROR_GENERAL              = "General Problem"
    ERROR_NO_WRITE_ACCESS      = "No Write Access"
    ERROR_NO_SHEET_NAMES       = "No Sheet Names"
    ERROR_NO_SHEET             = "No Sheet"
    ERROR_MISSING_SHEETS       = "Missing Sheets"
    ERROR_UNEXPECTED_SHEETS    = "Unexpected Sheets"
    ERROR_ORDER_SHEETS         = "Incorrect Order Sheets"
    ERROR_EMPTY_ROW            = "Empty Rows"
    ERROR_EMPTY_COLUMN         = "Empty Columns"
    ERROR_MISSING_COLUMNS      = "Missing Columns"
    ERROR_UNRECOGNIZED_COLUMNS = "Unrecognized Columns"
    ERROR_ORDER_COLUMNS        = "Incorrect Order Columns"
    ERROR_CONFIGURATION        = "No valid configuration"

    WARNING_TYPE               = "Incorrect Type"
    WARNING_EMPTY_ROW          = "Empty Rows"
    WARNING_EMPTY_COLUMN       = "Empty Columns"
    WARNING_NO_SCHEMA          = "No Schema"
    WARNING_MODULE             = "Module"
    WARNING_EMPTY_PACKAGE      = "Empty Package"
    

    def __init__(self, filename: str, configuration: str, **kwargs):
        #logging
        self._logger = logging.getLogger(__name__)

        reader.excel.warnings.simplefilter(action="ignore")

        #initialise
        self._wb = None
        self._checks = None
        
        #check file and define basepath for output
        filename = os.path.abspath(filename)
        assert os.path.isfile(filename) and os.access(filename, os.R_OK), \
                "File '%s' doesn't exist or isn't readable" % filename

        #include progress bar for cli call, or status update
        self._cli = bool(kwargs.get("cli", False))
        self._statusRegister = kwargs.get("statusRegister",None)
        self._statusIdentifier = kwargs.get("statusIdentifier",None)
        
        if kwargs.get("create", False):
            config_filename = str(kwargs.get("create"))
            config_location = os.path.abspath(config_filename)
            if os.path.exists(config_filename):
                raise Exception("couldn't create configuration, %s already exists" % os.path.basename(config_filename))
            else:                    
                self._filename = os.path.basename(filename)
                with tempfile.TemporaryDirectory() as tmp_dir:
                    output_filename = os.path.join(tmp_dir,self._filename)
                    copyfile(filename,output_filename)
                    self._basepath = tmp_dir
                    self._createConfiguration(config_location)
                self._basepath = None
        else:        
            #check configuration
            configurationFilename = os.path.abspath(configuration)
            assert os.path.isfile(configurationFilename) and os.access(configurationFilename, os.R_OK), \
                    "Configuration file '{}' doesn't exist or isn't readable".format(configuration)
            self._parseConfiguration(configurationFilename)
            #process webinterface settings
            self._webinterfaceData = kwargs.get("webinterfaceData",{})
            for entry in self._config.get("webinterface",[]):
                for option in entry.get("options",[]):
                    key = "option%s" % "_".join(option["setting"])
                    if option["type"] == "boolean":
                        if key in self._webinterfaceData:
                            value = int(self._webinterfaceData[key])>0
                            Validate._updateNestedDictionary(self._config, option["setting"], value)
            #load plugins
            if not self._config is None and self._pluginPath:
                if(os.path.isdir(self._pluginPath)):
                    for file in os.listdir(self._pluginPath):
                        if os.path.isfile(os.path.join(self._pluginPath, file)):
                            # try:
                            customChecks = importlib.machinery.SourceFileLoader(pathlib.Path(file).stem, 
                                                os.path.join(self._pluginPath, file)).load_module()
                            members = inspect.getmembers(customChecks, inspect.isclass)
                            for member in members:
                                if issubclass(member[1],Plugin) or 1==1:
                                    system.register(member[0],member[1]())
                                # elif issubclass(member[1],Error):
                                #     errors[member[0]] = member[1]
                            # except Exception as ex:
                            #     raise Exception("could not load plugin %s: %s" % (os.path.basename(file),ex))
            #initialise for validation
            self._package = Package()
            self._name = kwargs.get("name", os.path.basename(filename))
            self._report = ValidationReport(self._logger)
            self._expectedSheets = set()
            #start validation
            if not self._config is None:
                self._filename = os.path.basename(filename)
                if not kwargs.get("updateFile", False):
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        output_filename = os.path.join(tmp_dir,self._filename)
                        copyfile(filename,output_filename)
                        self._basepath = tmp_dir
                        self._validate()
                    self._basepath = None
                else:
                    self._basepath = os.path.dirname(filename)
                    self._validate()

    def getConfigFilename(config=None):
        internal_config_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)),"config")
        #get configuration for validation
        if not (config is None or config == False):
            config_filename = os.path.abspath(config)
            if not os.path.exists(config_filename):
                config_filename = os.path.join(internal_config_directory, config)
            if not os.path.exists(config_filename):
                raise FileNotFoundError("Can't find configuration %s" % config)
            elif os.path.isdir(config_filename):
                config_filename = os.path.join(config_filename, "config.json")
                if not os.path.exists(config_filename):
                    raise FileNotFoundError("Can't find configuration %s" % config)
        else:
            config_filename = os.path.join(internal_config_directory, "default/config.json")
        return config_filename
               
    def _createConfiguration(self, configurationLocation: str):
        #initialise
        progression_bar = tqdm(total=2, disable=(not self._cli), leave=False)
        progression_bar_size = 25
        progression_bar.set_description(str("Initialise").ljust(progression_bar_size))
        #check file
        package = Package()
        resourceFilename = os.path.join(self._basepath, self._filename)
        self._wb = load_workbook(os.path.join(resourceFilename))
        self._availableSheets = list(self._wb.sheetnames)
        #update
        progression_bar.reset(total=len(self._availableSheets)+2)        
        #initialise
        os.makedirs(configurationLocation, exist_ok=True)
        os.mkdir(os.path.join(configurationLocation,"schema"))
        configuration = {
            "settings": {
                "schemaPath": "schema",
                "pluginPath": "plugins",
                "allowAdditionalSheets": False,
                "requireSheetOrder": True,
                "adjustTypeForStringColumns": True,
                "removeEmptyRows": True,
                "removeEmptyColumns": True
            },
            "webinterface": [],
            "package": {},
            "sheets":[]
        }
        resourceNames = set()
        progression_bar.update(1)
        for i,name in enumerate(self._availableSheets):
            #set progress
            n = progression_bar_size - 10
            sheet = "[%s...]"%name[:(n-3)] if len(name)>n else "[%s]"%name
            progression_bar.set_description("Analyse %s" % sheet.ljust(n+2))
            #filter
            self._removeEmptyRowsForSheet(name)
            self._removeEmptyColumnsForSheet(name)
            #describe
            resource = "%s_%s" % (str(i).zfill(2),re.sub("[^a-zA-Z0-9]+", "", name).lower())
            sheetConfig = {"name": name, "resource": resource, "schema": {}}
            dialect = Dialect()
            dialect.add_control(formats.ExcelControl(sheet=name, preserve_formatting=False))
            schemaResource = describe(resourceFilename, type="schema", dialect=dialect)
            schemaResource.to_json(os.path.join(configurationLocation,"schema","%s.json"%resource))
            sheetConfig["schema"]["file"] = "%s.json"%resource
            configuration["sheets"].append(sheetConfig)
            #update
            progression_bar.update(1)            
        #check configuration
        progression_bar.set_description(str("Validate configuration").ljust(progression_bar_size))
        with open(os.path.join(os.path.dirname(__file__),"config.json"), encoding="UTF-8") as configurationSchema:
            schema = json.load(configurationSchema)
        jsonschema.validate(configuration,schema)
        progression_bar.update(1)
        #store
        with open(os.path.join(configurationLocation,"config.json"), "w") as f:
            json.dump(configuration, f, ensure_ascii=False)

    def _addSchemaDependencies(self, resource, name, dependencies, schema):
        #update dependencies
        for entry in schema["data"].get("foreignKeys",[]):
            if "reference" in entry and "resource" in entry["reference"]:
                if not entry["reference"]["resource"]==resource:
                    if entry["reference"]["resource"] in self._resourceSheet:
                        sheetName = self._resourceSheet[entry["reference"]["resource"]]
                        if not sheetName in dependencies:
                            dependencies.append(sheetName)
                            self._logger.debug("set dependency on '%s for '%s' based on frictionless schema",
                                sheetName,name)
        for entry in schema.get("dynamic",[]):
            if "dynamicResources" in entry:
                for value in entry["dynamicResources"].values():
                    if value["resource"] in self._resourceSheet:
                        sheetName = self._resourceSheet[value["resource"]]
                        if not sheetName in dependencies:
                            dependencies.append(sheetName)
                            self._logger.debug("set dependency on '%s' for '%s' based on dynamic frictionless schema",
                                               sheetName,name)
            if "linkedResources" in entry:
                for value in entry["linkedResources"].values():
                    if value["resource"] in self._resourceSheet:
                        sheetName = self._resourceSheet[value["resource"]]
                        if not sheetName in dependencies:
                            dependencies.append(sheetName)
                            self._logger.debug("set dependency on '%s' for '%s' based on dynamic frictionless schema",
                                sheetName,name)
        return dependencies


    def _parseConfiguration(self, configurationFilename: str):
        self._config = None
        self._allowedSheets = []
        self._resourceSheet = {}
        self._schemaPath = None
        try:
            with open(configurationFilename, encoding="UTF-8") as configurationData:
                self._config = json.load(configurationData)
            with open(os.path.join(os.path.dirname(__file__),"config.json"), encoding="UTF-8") as configurationSchema:
                schema = json.load(configurationSchema)
            jsonschema.validate(self._config,schema)
            #set paths
            self._schemaPath = self._config.get("settings",{}).get("schemaPath",None)
            if self._schemaPath:
                if not os.path.isabs(self._schemaPath):
                    self._schemaPath = os.path.abspath(os.path.join(os.path.dirname(configurationFilename),self._schemaPath))
            self._pluginPath = self._config.get("settings",{}).get("pluginPath",None)
            if self._pluginPath:
                if not os.path.isabs(self._pluginPath):
                    self._pluginPath = os.path.abspath(os.path.join(os.path.dirname(configurationFilename),self._pluginPath))
            #loop over definitions
            for i in range(len(self._config["sheets"])):
                #require unique name
                assert self._config["sheets"][i]["name"] not in self._allowedSheets, "sheet name should be unique"
                self._allowedSheets.append(self._config["sheets"][i]["name"])
                if "resource" in self._config["sheets"][i]:
                    #require unique resource
                    assert self._config["sheets"][i]["resource"] not in self._resourceSheet, "resource name should be unique"
                    self._resourceSheet[self._config["sheets"][i]["resource"]] = self._config["sheets"][i]["name"]
                    #package
                    if "package" in self._config:
                        #checklist
                        if "checklist" in self._config["package"]:
                            if "file" in self._config["package"]["checklist"]:
                                if self._schemaPath:
                                    checklistFilename = os.path.join(self._schemaPath,
                                                              self._config["package"]["checklist"]["file"])
                                else:
                                    checklistFilename = os.path.join(os.path.dirname(configurationFilename),
                                                              self._config["package"]["checklist"]["file"])
                                with open(checklistFilename, encoding="UTF-8") as checklistData:
                                    self._config["package"]["checklist"]["data"] = json.load(checklistData)
                    #schema
                    if "schema" in self._config["sheets"][i]:
                        if "file" in self._config["sheets"][i]["schema"]:
                            if self._schemaPath:
                                schemaFilename = os.path.join(self._schemaPath,
                                                          self._config["sheets"][i]["schema"]["file"])
                            else:
                                schemaFilename = os.path.join(os.path.dirname(configurationFilename),
                                                          self._config["sheets"][i]["schema"]["file"])
                            with open(schemaFilename, encoding="UTF-8") as schemaData:
                                self._config["sheets"][i]["schema"]["data"] = json.load(schemaData)
                    #checklist
                    if "checklist" in self._config["sheets"][i]:
                        if "file" in self._config["sheets"][i]["checklist"]:
                            if self._schemaPath:
                                checklistFilename = os.path.join(self._schemaPath,
                                                          self._config["sheets"][i]["checklist"]["file"])
                            else:
                                checklistFilename = os.path.join(os.path.dirname(configurationFilename),
                                                          self._config["sheets"][i]["checklist"]["file"])
                            with open(checklistFilename, encoding="UTF-8") as checklistData:
                                self._config["sheets"][i]["checklist"]["data"] = json.load(checklistData)
                    #transform
                    if "transforms" in self._config["sheets"][i]:
                        for j in range(len(self._config["sheets"][i]["transforms"])):
                            #require unique resource
                            assert self._config["sheets"][i]["transforms"][j]["resource"] not in self._resourceSheet, "resource name should be unique"
                            self._resourceSheet[self._config["sheets"][i]["transforms"][j]["resource"]] = self._config["sheets"][i]["name"]
                            self._config["sheets"][i]["transforms"][j]["name"] = "%s -> %s" % (
                                self._config["sheets"][i]["name"], self._config["sheets"][i]["transforms"][j]["resource"])
                            if "file" in self._config["sheets"][i]["transforms"][j]["pipeline"]:
                                if self._schemaPath:
                                    pipelineFilename = os.path.join(self._schemaPath,
                                                              self._config["sheets"][i]["transforms"][j]["pipeline"]["file"])
                                else:
                                    pipelineFilename = os.path.join(os.path.dirname(configurationFilename),
                                                              self._config["sheets"][i]["transforms"][j]["pipeline"]["file"])
                                with open(pipelineFilename, encoding="UTF-8") as pipelineData:
                                    self._config["sheets"][i]["transforms"][j]["pipeline"]["data"] = json.load(pipelineData)
                            if "file" in self._config["sheets"][i]["transforms"][j]["schema"]:
                                if self._schemaPath:
                                    schemaFilename = os.path.join(self._schemaPath,
                                                              self._config["sheets"][i]["transforms"][j]["schema"]["file"])
                                else:
                                    schemaFilename = os.path.join(os.path.dirname(configurationFilename),
                                                              self._config["sheets"][i]["transforms"][j]["schema"]["file"])
                                with open(schemaFilename, encoding="UTF-8") as schemaData:
                                    self._config["sheets"][i]["transforms"][j]["schema"]["data"] = json.load(schemaData)
            #reloop over definitions to get dependencies
            for i in range(len(self._config["sheets"])):
                if not "dependencies" in self._config["sheets"][i]:
                    self._config["sheets"][i]["dependencies"] = []
                #check schema
                if "resource" in self._config["sheets"][i] and "schema" in self._config["sheets"][i]:
                    self._config["sheets"][i]["dependencies"] = self._addSchemaDependencies(
                        self._config["sheets"][i]["resource"], self._config["sheets"][i]["name"], 
                        self._config["sheets"][i]["dependencies"], self._config["sheets"][i]["schema"])
                if "resource" in self._config["sheets"][i]:
                    if "transforms" in self._config["sheets"][i]:
                        for j in range(len(self._config["sheets"][i]["transforms"])):
                            if "schema" in self._config["sheets"][i]["transforms"][j]:
                                self._config["sheets"][i]["dependencies"] = self._addSchemaDependencies(
                                    self._config["sheets"][i]["resource"], self._config["sheets"][i]["name"], 
                                    self._config["sheets"][i]["dependencies"], self._config["sheets"][i]["transforms"][j]["schema"])
        except Exception as ex:
            self._logger.error("Could not parse configuration: %s", str(ex))
            self._config = None

    def _computeExpectedSheets(self):
        self._expectedSheets = set()
        #recursively get dependencies
        def getAllDependencies(dependencies):
            checkDependencies = [sheetName for sheetName in dependencies]
            newDependencies = [sheetName for sheetName in dependencies]
            while len(checkDependencies)>0:
                newCheckDependencies = []
                for entry in self._config["sheets"]:
                    if ("dependencies" in entry) and (entry["name"] in checkDependencies):
                        for dependency in entry["dependencies"]:
                            if not dependency in newDependencies:
                                newDependencies.append(dependency)
                                newCheckDependencies.append(dependency)
                checkDependencies = newCheckDependencies
            return newDependencies
        #check defined sheets
        for i in range(len(self._config["sheets"])):
            if (self._config["sheets"][i]["name"] in self._availableSheets) or not self._config["sheets"][i]["optional"]:
                self._expectedSheets.add(self._config["sheets"][i]["name"])
                if "dependencies" in self._config["sheets"][i]:
                    dependencies = []
                    for sheetName in self._config["sheets"][i]["dependencies"]:
                        dependencies.append(sheetName)
                    self._expectedSheets.update(getAllDependencies(dependencies))

    def _computeValidationOrder(self):
        validationEntries = []
        validationSheets = []
        recheckList = []
        for entry in self._config["sheets"]:
            if "resource" in entry:
                if entry["name"] in self._availableSheets:
                    if len(set(entry["dependencies"]).difference(validationSheets))==0:
                        validationEntries.append(entry)
                        validationSheets.append(entry["name"])
                    else:
                        recheckList.append(entry)
                    recheck = True
                    while recheck:
                        recheck = False
                        newRecheckList = []
                        for entry in recheckList:
                            if len(set(entry["dependencies"]).difference(validationSheets))==0:
                                validationEntries.append(entry)
                                validationSheets.append(entry["name"])
                                recheck = True
                            else:
                                newRecheckList.append(entry)
                        recheckList = newRecheckList
        if len(recheckList)>0:
            recheckSheetList = [entry["name"] for entry in recheckList]
            self._report.addReportError("general", "Can't solve dependencies for sheets","'{}'".format("', '".join(recheckSheetList)))
        return validationEntries

    def _getSheetColumnNames(self, sheetName:str):
        self._logger.debug("get column names from sheet '{}'".format(sheetName))
        columnNames = []
        if not self._wb:
            resourceFilename = os.path.join(self._basepath, self._filename)
            self._wb = load_workbook(os.path.join(resourceFilename))
        if not sheetName in self._wb.sheetnames:
            self._logger.error("can't find sheet {} in resource".format(sheetName))
        else:
            ws = self._wb[sheetName]
            for row in ws.rows:
                columnNames = [cell.internal_value for cell in row]
                break
        return columnNames

    def _checkMissingColumns(self, sheetName: str, resource:Resource, reportId:str):
        adjusted_resource = False
        if not reportId in self._report.reports:
            self._logger.error("reportId {} not found".format(reportId))
        elif not sheetName in self._availableSheets:
            self._logger.error("sheetName {} not found".format(sheetName))
        else:
            try:
                headerCase = resource.dialect.header_case
                resourceColumnNames = self._getSheetColumnNames(sheetName)
                requiredNames = resource.schema.field_names
                if headerCase:
                    missingNames = set(requiredNames).difference(resourceColumnNames)
                    unrecognizedNames = set(resourceColumnNames).difference(requiredNames)
                else:
                    resourceColumnNamesUpper = [entry.upper() for entry in resourceColumnNames]
                    requiredNamesUpper = [entry.upper() for entry in requiredNames]
                    missingNames = set([x for x,y in zip(requiredNames,requiredNamesUpper)
                                        if not y in resourceColumnNamesUpper])
                    unrecognizedNames = set([x for x,y in zip(resourceColumnNames,resourceColumnNamesUpper)
                                             if not y in requiredNamesUpper])
                if len(missingNames)>0:
                    self._report.addReportError(reportId, Validate.ERROR_MISSING_COLUMNS, "'{}'".format("', '".join(missingNames)))
                    adjusted_resource = True
                if len(unrecognizedNames)>0:
                    unrecognizedStringNames = set([(x if not x is None else "") for x in unrecognizedNames])
                    self._report.addReportError(reportId, Validate.ERROR_UNRECOGNIZED_COLUMNS, 
                                                "'{}'".format("', '".join(unrecognizedStringNames)))
                    adjusted_resource = True
                if headerCase:
                    recognizedResourceNames = [item for item in resourceColumnNames if item in requiredNames]
                    recognizedRequiredNames = [item for item in requiredNames if item in resourceColumnNames]
                else:
                    recognizedResourceNames = [item for item in resourceColumnNamesUpper if item in requiredNamesUpper]
                    recognizedRequiredNames = [item for item in requiredNamesUpper if item in resourceColumnNamesUpper]
                if not recognizedResourceNames==recognizedRequiredNames:
                    self._report.addReportError(reportId, Validate.ERROR_ORDER_COLUMNS, "expected: {}".
                                                format(", ".join(requiredNames)))
                    adjusted_resource = True
                #reconstruct schema frictionless
                if adjusted_resource:
                    originalFields = {fieldName:resource.schema.get_field(fieldName) for fieldName in requiredNames}
                    resource.schema.clear_fields()
                    for fieldName in resourceColumnNames:
                        if fieldName and fieldName in originalFields:
                            resource.schema.add_field(originalFields[fieldName])
                        else:
                            field = fields.AnyField(name=fieldName)
                            resource.schema.add_field(field)
                    resource.schema.primary_key = [fieldName for fieldName in resource.schema.primary_key
                                                   if resource.schema.has_field(fieldName)]
                    resource.schema.foreign_keys = [entry for entry in resource.schema.foreign_keys
                                                   if any([resource.schema.has_field(fieldName) for fieldName in entry["fields"]])]
            # except Exception as e:
            #     self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem checking missing columns: {}".format(str(e)))
            finally:
                pass
        return adjusted_resource

    def _removeEmptyRowsForSheet(self, sheetName:str, resource:Resource = None, reportId:str = None, allowed=True):
        try:
            if not resource is None:
                headerRows = resource.dialect.header_rows
            else:
                headerRows = [1]
            if not reportId is None:
                self._report.addReportDebug(reportId, "remove empty rows from sheet '{}'".format(sheetName))
            resourceFilename = os.path.join(self._basepath, self._filename)
            if not os.access(resourceFilename, os.W_OK):
                if not (resource is None or reportId is None):
                    self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                            "no access to {}, can't try to remove empty rows".format(resource.name))
            else:
                if not self._wb:
                    self._wb = load_workbook(os.path.join(resourceFilename))
                if not sheetName in self._wb.sheetnames:
                    if not reportId is None:
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET,
                                            "can't find {} in resource".format(sheetName))
                else:
                    ws = self._wb[sheetName]
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for row in ws.rows:
                        values = [cell for cell in row if not (cell.internal_value is None or
                                                              str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in row:
                                rowId = cell.row
                                break
                            if not id in headerRows:
                                deletablIds.append(rowId)
                    if len(deletablIds)>0:
                        if not reportId is None:
                            if allowed:
                                self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_ROW,
                                                "removed {} empty rows".format(len(deletablIds)))
                            else:
                                self._report.addReportWarning(reportId, Validate.ERROR_EMPTY_ROW,
                                                "removed {} empty rows".format(len(deletablIds)))
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i,rowId in enumerate(sortedDeletableIds):
                            deletableList.append(rowId)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(rowId-1):
                                #just continue
                                pass
                            else:
                                ws.delete_rows(deletableList[-1],len(deletableList))
                                deletableList = []
                    #always save
                    self._wb.save(filename = resourceFilename)
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "updated {} after removing {} empty rows".format(resource.name,len(deletablIds))) 
        except Exception as e:
            if not reportId is None:
                self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem removing empty rows","{}".format(str(e)))

    def _removeEmptyColumnsForSheet(self, sheetName:str, resource:Resource = None, reportId:str = None, allowed=True):
        try:
            if not reportId is None:
                self._report.addReportDebug(reportId, "remove empty columns from sheet '{}'".format(sheetName))
            resourceFilename = os.path.join(self._basepath, self._filename)
            if not os.access(resourceFilename, os.W_OK):
                self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                            "no access to {}, can't try to remove empty columns".format(resource.name))
            else:
                if not self._wb:
                    self._wb = load_workbook(os.path.join(resourceFilename))
                if not sheetName in self._wb.sheetnames:
                    if not (resource is None or reportId is None):
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET, "can't find {} in resource".format(sheetName))
                else:
                    ws = self._wb[sheetName]
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                        "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for column in ws.columns:
                        values = [cell for cell in column if not (cell.internal_value is None or
                                                                  str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in column:
                                columnId = cell.column
                                break
                            deletablIds.append(columnId)
                    if len(deletablIds)>0:
                        if not reportId is None:
                            if allowed:
                                self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_COLUMN,
                                                "removed {} empty columns".format(len(deletablIds))) 
                            else:
                                self._report.addReportError(reportId, Validate.ERROR_EMPTY_COLUMN,
                                                "removed {} empty columns".format(len(deletablIds))) 
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i in range(len(sortedDeletableIds)):
                            columnId = sortedDeletableIds[i]
                            deletableList.append(columnId)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(columnId-1):
                                #just continue
                                pass
                            else:
                                ws.delete_cols(deletableList[-1],len(deletableList))
                                deletableList = []
                    #always save
                    self._wb.save(filename = resourceFilename)
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "updated {} after removing {} empty columns".format(resource.name,len(deletablIds))) 
        except Exception as e:
            if not reportId is None:
                self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem removing empty columns: {}".format(str(e)))

    def _updateTypeForStringColumns(self, sheetName:str, resource:Resource, reportId:str):
        try:
            headerRows = resource.dialect.header_rows
            columnNames = [field.name for field in resource.schema.fields if field.name and field.type=="string"]
            if len(columnNames)>0:
                self._report.addReportDebug(reportId,
                            "convert column(s) '{}' to string in '{}'".format("', '".join(columnNames),sheetName))
                resourceFilename = os.path.join(self._basepath, self._filename)
                if not os.access(resourceFilename, os.W_OK):
                    self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                                "no access to {}, can't try to convert columns to string".format(resource.name))
                else:
                    if not self._wb:
                        self._wb = load_workbook(os.path.join(resourceFilename))
                    if not sheetName in self._wb.sheetnames:
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET,
                            "can't find {} in resource".format(sheetName))
                    else:
                        ws = self._wb[sheetName]
                        totalUpdatedNumber=0
                        for column in ws.columns:
                            if column[0].value in columnNames:
                                updatedNumber = 0
                                for row,item in enumerate(column):
                                    if row+1 in headerRows:
                                        continue
                                    if not (item.internal_value is None or isinstance(item.value,str)):
                                        item.value = str(item.value)
                                        updatedNumber+=1
                                if updatedNumber>0:
                                    self._report.addReportWarning(reportId, Validate.WARNING_TYPE,
                                        "changed type for {} entries from column '{}'".format(
                                            updatedNumber,column[0].value))
                                totalUpdatedNumber+=updatedNumber
                        #always save
                        self._wb.save(filename = resourceFilename)
                        self._report.addReportDebug(reportId,
                                                "updated {} cells to string in {}".format(
                                                totalUpdatedNumber,resource.name))
        except Exception as e:
            self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem converting sheet columns to string: {}".format(str(e)))

    def _validate(self):
        self._logger.debug("start validation")

        #set status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"started": int(time.time())})
                self._statusRegister.update({self._statusIdentifier: status})
            else:
                return

        #get available sheets
        self._report.addReport("general","General",False,ValidationReport.TYPE_GENERAL)
        #set status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"status": "initialise validation"})
                self._statusRegister.update({self._statusIdentifier: status})               
            else:
                return
        progression_bar = tqdm(total=2, disable=(not self._cli), leave=False)
        progression_bar_size = 25
        progression_bar.set_description(str("Initialise validation").ljust(progression_bar_size))
        #ry to get sheetnames
        try:
            resourceFilename = os.path.join(self._basepath, self._filename)
            self._wb = load_workbook(os.path.join(resourceFilename))
            self._availableSheets = list(self._wb.sheetnames)
        except Exception as e:
            self._report.addReportError("general", Validate.ERROR_NO_SHEET_NAMES,
                                        "problem retrieving sheetnames from '{}': {}".format(
                                            os.path.basename(self._filename),str(e)))
            if (self._statusRegister and self._statusIdentifier):
                if self._statusIdentifier in self._statusRegister:
                    status = self._statusRegister[self._statusIdentifier]
                    status.update({"ended": int(time.time())})
                    self._statusRegister.update({self._statusIdentifier: status})
            return

        #configuration
        allowAdditionalSheets = self._config["settings"].get("allowAdditionalSheets", False)
        requireSheetOrder = self._config["settings"].get("requireSheetOrder", False)        

        #compute the expected sheets, check if everything is included and compute the order for validation
        self._computeExpectedSheets()
        #update status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"step": 0, "total": len(self._expectedSheets)+2})
                self._statusRegister.update({self._statusIdentifier: status})
            else:
                return
        progression_bar.reset(total=len(self._expectedSheets)+2)
        
        missingSheets = [sheetName for sheetName in self._expectedSheets if not sheetName in self._availableSheets]
        if len(missingSheets)>0:
            self._report.addReportError("general", Validate.ERROR_MISSING_SHEETS, "'{}'".format("', '".join(missingSheets)))
        additionalSheets = [sheetName for sheetName in self._availableSheets if not sheetName in self._allowedSheets]
        if len(additionalSheets)>0:
            if not allowAdditionalSheets:
                self._report.addReportError("general", Validate.ERROR_UNEXPECTED_SHEETS, "'{}'".format("', '".join(additionalSheets)))
            else:
                self._report.addReportInfo("general", "Ignoring additional sheet(s): '{}'".format("', '".join(additionalSheets)))
        if requireSheetOrder:
            requiredOrder = [entry["name"] for entry in self._config["sheets"] if entry["name"] in self._availableSheets]
            availableOrder = [sheetName for sheetName in self._availableSheets if sheetName in requiredOrder]
            if not requiredOrder==availableOrder:
                self._report.addReportError("general", Validate.ERROR_ORDER_SHEETS, "'{}'".format("', '".join(availableOrder)))
        #get validation order (check dependencies)
        validationOrder = self._computeValidationOrder()
        #update status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"step": 0, "total": len(validationOrder)+2})
                self._statusRegister.update({self._statusIdentifier: status})
            else:
                return
        progression_bar.reset(total=len(validationOrder)+2)
        progression_bar.update(1)
        #validate resources
        errorTypes = set()
        for entry in validationOrder:
            n = progression_bar_size - 13
            sheet = "[%s...]"%entry["name"][:(n-3)] if len(entry["name"])>n else "[%s]"%entry["name"]
            #set status
            if (self._statusRegister and self._statusIdentifier):
                if self._statusIdentifier in self._statusRegister:
                    status = self._statusRegister[self._statusIdentifier]
                    status.update({"status": "Validating [%s]" % entry["name"]})
                    self._statusRegister.update({self._statusIdentifier: status})
                else:
                    return
            progression_bar.set_description("Validating %s" % sheet.ljust(n+2))
            resource = self._createResource(entry) 
            if "schema" in entry:
                resource_validation = self._validateResource(resource,entry)
                if resource_validation:
                    errorTypes.update([item.type for item in resource_validation.tasks[0].errors])
            #try to add resource
            self._package.add_resource(resource)   
            if "transforms" in entry:
                for j in range(len(entry["transforms"])):
                    try:
                        transformResource = self._createTransform(entry["transforms"][j],resource,entry["name"])
                        if "schema" in entry["transforms"][j]:
                            resource_validation = self._validateResource(transformResource,entry["transforms"][j])
                            if resource_validation:
                                errorTypes.update([item.type for item in resource_validation.tasks[0].errors])
                        #try to add resource
                        self._package.add_resource(transformResource)
                    except Exception as e:
                        pass
            #set status
            if (self._statusRegister and self._statusIdentifier):
                if self._statusIdentifier in self._statusRegister:
                    status = self._statusRegister[self._statusIdentifier]
                    status.update({"step": status["step"]+1})
                    self._statusRegister.update({self._statusIdentifier: status})
                else:
                    return
            progression_bar.update(1)  
        #set status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"step": status["step"]+1, "status": "Validating package"})
                self._statusRegister.update({self._statusIdentifier: status})
            else:
                return
        progression_bar.set_description("Validating package")
        progression_bar.update(1)
        #validate package
        packageEntry = self._config.get("package",{})
        self._validatePackage(packageEntry, skip_errors = list(errorTypes))
        #set status
        if (self._statusRegister and self._statusIdentifier):
            if self._statusIdentifier in self._statusRegister:
                status = self._statusRegister[self._statusIdentifier]
                status.update({"step": status["step"]+1, "ended": int(time.time()), "status": "Validating finished"})
                self._statusRegister.update({self._statusIdentifier: status})
            else:
                return
        progression_bar.close()

    def _createResource(self, entry):
        reportId = "resource:{}".format(entry["resource"])
        self._report.addReport(reportId, entry["name"], True, ValidationReport.TYPE_RESOURCE)
        self._report.addReportDebug(reportId,"define resource from sheet '{}'".format(entry["name"]))
        #create resource
        dialectArguments = {}
        dialectArguments["header"] = entry.get("header",self._config["settings"].get(
            "header",settings.DEFAULT_HEADER))
        dialectArguments["header_rows"] = entry.get("headerRows",self._config["settings"].get(
            "headerRows",settings.DEFAULT_HEADER_ROWS))
        dialectArguments["header_join"] = entry.get("headerJoin",self._config["settings"].get(
            "headerJoin",settings.DEFAULT_HEADER_JOIN))
        dialectArguments["header_case"] = entry.get("headerCase",self._config["settings"].get(
            "headerCase",settings.DEFAULT_HEADER_CASE))
        dialectArguments["comment_char"] = entry.get("commentChar",self._config["settings"].get(
            "commentChar",None))
        dialectArguments["comment_rows"] = entry.get("commentRows",self._config["settings"].get(
            "commentRows",[]))
        dialectArguments["skip_blank_rows"] = entry.get("skipBlankRows",self._config["settings"].get(
            "skipBlankRows",False))
        dialect = Dialect(**dialectArguments)
        dialect.add_control(formats.ExcelControl(sheet=entry["name"], preserve_formatting=False))
        resource = Resource(basepath=self._basepath, path=self._filename, dialect=dialect)
        #set name and remove if exists
        if self._package.has_resource(entry["resource"]):
            self._package.remove_resource(entry["resource"])
        resource.name = entry["resource"]
        return resource

    def _createTransform(self, entry, resource, parentName):
        reportId = "resource:{}".format(entry["resource"])
        parentId = "resource:{}".format(resource.name)
        self._report.addReport(reportId, entry["name"], True, ValidationReport.TYPE_RESOURCE_TRANSFORM, parentId, parentName)
        self._report.addReportDebug(reportId,"define transformation from sheet '{}'".format(entry["name"]))
        #create resource
        pipeline = Pipeline.from_descriptor(entry["pipeline"]["data"])
        transform = resource.to_copy()
        transform.schema = None
        transform.transform(pipeline)
        transform.name = entry["resource"]
        return transform

    def _validateResource(self, resource, entry, skip_errors:list=None):
        reportId = "resource:{}".format(entry["resource"])
        adjustTypeForStringColumns = self._config["settings"].get("adjustTypeForStringColumns", False)
        removeEmptyRows = self._config["settings"].get("removeEmptyRows", False)
        removeEmptyColumns = self._config["settings"].get("removeEmptyColumns", False)
        if resource.memory:
            resourceColumnNames = list(resource.data.data.fieldnames())
        else:
            resourceColumnNames = self._getSheetColumnNames(entry["name"])
        #define full schema
        resource.schema = Schema.from_descriptor(entry["schema"]["data"])
        #dynamic
        if "dynamic" in entry["schema"]:
            resources = {}
            for dynamic in entry["schema"]["dynamic"]:
                dynamicResources = {}
                mappings = dynamic.get("mappings",{})
                #get resources
                if "dynamicResources" in dynamic:
                    for name,dynamicResource in dynamic["dynamicResources"].items():
                        if not dynamicResource["resource"] in resources:
                            resources[dynamicResource["resource"]] = pd.DataFrame(self._package.get_resource(
                                dynamicResource["resource"]).extract().get(dynamicResource["resource"],[]))
                        resourceData = resources[dynamicResource["resource"]].copy()
                        if "condition" in dynamicResource:
                            for condition in dynamicResource["condition"]:
                                resourceData = resourceData[resourceData[condition["field"]]==condition["value"]]
                        dynamicResources[name] = resourceData
                if "linkedResources" in dynamic:
                    for name,linkedResource in dynamic["linkedResources"].items():
                        if not linkedResource["resource"] in resources:
                            resources[linkedResource["resource"]] = pd.DataFrame(self._package.get_resource(
                                linkedResource["resource"]).extract().get(linkedResource["resource"],[]))
                #create fields
                newEntryFields = []
                dynamicResourcesList = sorted(list(dynamicResources.keys()))
                dynamicResourcesIterators = [dynamicResources[key].iterrows() for key in dynamicResourcesList]
                linkedResources = {}
                recomputeLinkedResources = False
                required = dynamic.get("required",True)
                for dynamicEntry in itertools.product(*dynamicResourcesIterators):
                    dynamicResourcesEntry = dict(map(lambda k,v : (k,v[1]), dynamicResourcesList,dynamicEntry))
                    if "linkedResources" in dynamic and (len(linkedResources)==0 or recomputeLinkedResources):
                        linkedResources = {}
                        for name,linkedResource in dynamic["linkedResources"].items():
                            resourceData = resources[linkedResource["resource"]].copy()
                            if ("condition" in linkedResource) and (resourceData.shape[0]>0):
                                for condition in linkedResource["condition"]:
                                    if isinstance(condition["value"], dict):
                                        recomputeLinkedResources = True
                                        value = dynamicResourcesEntry.get(
                                                condition["value"]["dynamicResource"],{}).get(condition["value"]["field"],None)
                                    else:
                                        value = condition["value"]
                                    resourceData = resourceData[resourceData[condition["field"]].values==value]
                            linkedResources[name] = resourceData
                    for fieldEntry in dynamic["fields"]:
                        #create field
                        fieldDescriptor = {}
                        for key in ["name","type","rdfType","title","format","example","description"]:
                            fieldDescriptor = setDescriptorValueDynamicString(
                                key,fieldDescriptor,fieldEntry,
                                dynamicResourcesEntry,mappings,linkedResources,self._logger)
                        #check is missing field is allowed
                        if not fieldDescriptor.get("name","") in resourceColumnNames:
                            if required==False:
                                continue
                            elif isinstance(required,list) and not fieldDescriptor.get("name","") in required:
                                continue
                        #add constraints
                        if "constraints" in fieldEntry:
                            fieldDescriptor["constraints"] = {}
                            for key in ["enum"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicList(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["required"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicBoolean(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["minLength","maxLength"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicInteger(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["minimum","maximum"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicNumber(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["pattern"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicString(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                        newEntryFields.append(Field.from_descriptor(fieldDescriptor))
                if not dynamic.get("ordered",True):
                    newFields = []
                    for resourceColumnName in resourceColumnNames:
                        for item in newEntryFields:
                            if item.name==resourceColumnName:
                                newFields.append(item)
                    for item in newEntryFields:
                        if not item.name in resourceColumnNames:
                            newFields.append(item)
                else:
                    newFields = newEntryFields
                #update schema
                position = dynamic.get("position","after")
                field = dynamic.get("field",None)
                if field is None:
                    if position=="before":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField,position=1+i)
                    elif position=="after":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField)
                elif resource.schema.has_field(field):
                    pos = min([i+1 for i,schemaField in enumerate(resource.schema.fields)
                                    if schemaField.name==field])
                    if position=="before":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField, position=pos+i)
                    elif position=="after":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField, position=pos+i+1)
        
        #adjust schema
        if not entry["schema"].get("required",True):
            removeFields = set()
            for fieldName in resource.schema.field_names:
                if not fieldName in resourceColumnNames:
                    removeFields.add(fieldName)
            for fieldName in removeFields:
                resource.schema.remove_field(fieldName)
        if not entry["schema"].get("ordered",True):
            allFields = []
            for fieldName in resource.schema.field_names:
                allFields.append(resource.schema.get_field(fieldName))
            resource.schema.clear_fields()
            for resourceColumnName in resourceColumnNames:
                for item in allFields:
                    if item.name==resourceColumnName:
                        resource.schema.add_field(item)
            for item in allFields:
                if not item.name in resourceColumnNames:
                    resource.schema.add_field(item)
        #missing
        if "missing" in entry["schema"]:
            for missing in entry["schema"]["missing"]:
                if len(resource.schema.fields)==0:
                    for resourceColumnName in resourceColumnNames:
                        newField = Field.from_descriptor({"name": resourceColumnName})
                        resource.schema.add_field(newField)
                elif missing["position"] == "after":
                    addNewFields = False
                    for resourceColumnName in resourceColumnNames:
                        if addNewFields:
                            newField = Field.from_descriptor({"name": resourceColumnName})
                            resource.schema.add_field(newField)
                        elif resourceColumnName == resource.schema.fields[-1].name:
                            addNewFields = True
                elif missing["position"] == "before":
                    for i,resourceColumnName in enumerate(resourceColumnNames):
                        if resourceColumnName == resource.schema.fields[0].name:
                            break
                        else:
                            newField = Field.from_descriptor({"name": resourceColumnName})
                            resource.schema.add_field(newField, position=1+i)
                elif missing["position"] == "any":
                    allFields = []
                    allFieldNames = []
                    for fieldName in resource.schema.field_names:
                        allFields.append(resource.schema.get_field(fieldName))
                        allFieldNames.append(fieldName)
                    resource.schema.clear_fields()
                    i = 0
                    for resourceColumnName in resourceColumnNames:
                        if i<len(allFields) and allFields[i].name==resourceColumnName:
                            resource.schema.add_field(allFields[i])
                            i+=1
                            continue
                        elif not resourceColumnName in allFieldNames:
                            newField = Field.from_descriptor({"name": resourceColumnName})
                            resource.schema.add_field(newField)
                    for item in allFields[i:]:
                        resource.schema.add_field(item)
        #check types, empty rows and columns
        if resource.path:
            #detect too large number of columns and try to pre-filter empty columns
            if len(resourceColumnNames)>len(resource.schema.fields):
                if removeEmptyColumns:
                    self._removeEmptyColumnsForSheet(entry["name"],resource,reportId)
                else:
                    #create error if empty columns found
                    self._removeEmptyColumnsForSheet(entry["name"],resource,reportId, False)
            #initial checks
            pick_errors = ["type-error","blank-row","extra-label"]
            if len(pick_errors)>0:
                resource_validation = resource.validate(checklist=Checklist(pick_errors=pick_errors, skip_errors=skip_errors))
                if not resource_validation.valid:
                    errorTypes = set([item.type for item in resource_validation.tasks[0].errors])
                    if adjustTypeForStringColumns and "type-error" in errorTypes:
                        self._updateTypeForStringColumns(entry["name"],resource,reportId) 
                        if len(resource_validation.tasks[0].errors)>=settings.DEFAULT_LIMIT_ERRORS:                                
                            resource_validation = resource.validate(checklist=Checklist(pick_errors=pick_errors, skip_errors=skip_errors))
                            if not resource_validation.valid:
                                errorTypes = set([item.type for item in resource_validation.tasks[0].errors])   
                    if removeEmptyColumns and "extra-label" in errorTypes:
                        self._removeEmptyColumnsForSheet(entry["name"],resource,reportId)
                        if len(resource_validation.tasks[0].errors)>=settings.DEFAULT_LIMIT_ERRORS:                                
                            resource_validation = resource.validate(checklist=Checklist(pick_errors=pick_errors, skip_errors=skip_errors))
                            if not resource_validation.valid:
                                errorTypes = set([item.type for item in resource_validation.tasks[0].errors])
                    if removeEmptyRows and "blank-row" in errorTypes:
                        self._removeEmptyRowsForSheet(entry["name"],resource,reportId)                    
        #validate
        checklist = Checklist.from_descriptor(entry["checklist"]["data"]) if "checklist" in entry else Checklist()
        checklist.skip_errors = skip_errors
        resource_validation = resource.validate(checklist=checklist)
        self._report.setFrictionless(reportId, resource_validation)
        self._report.addReportDebug(reportId, resource_validation.stats)
        if resource_validation.valid and skip_errors is None:
            self._report.addReportInfo(reportId,"succesfull frictionless validation '{}' sheet".format(entry["name"]))
        #check missing columns and possibly revalidate
        elif skip_errors is None and not resource_validation.valid and resource.path:
            if self._checkMissingColumns(entry["name"],resource,reportId):
                resource_validation = resource.validate(checklist=checklist)
                self._report.setFrictionless(reportId, resource_validation)
                self._report.addReportDebug(reportId, resource_validation.stats)
        #return validation
        return resource_validation

    def _validatePackage(self, entry, skip_errors:list=None):
        reportId = "package"
        self._report.addReport(reportId, "Package", True, ValidationReport.TYPE_PACKAGE)
        #validate
        if len(self._package.resources)>0:
            checklist = Checklist.from_descriptor(entry["checklist"]["data"]) if "checklist" in entry else Checklist()
            checklist.skip_errors = skip_errors
            package_validation = self._package.validate(checklist=checklist)
            self._report.setFrictionless(reportId, package_validation)
            self._report.addReportDebug(reportId, package_validation.stats)
        else:
            self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_PACKAGE, "no sheets found to be validated")

    def createPackageJSON(self, filename:str = None):
        """
        Create frictionless package data
        """
        if not filename is None:
            try:
                return self._package.to_json(filename)
            except:
                self._logger.error("can't store package to '%s'",filename)
                return self._package.to_json()
        else:
            return self._package.to_json()

    def createReport(self, filename:str=None):
        """
        Create JSON object with report
        """
        reportObject = {"name": os.path.basename(self._name),
                        "version": __version__, 
                        "valid": self.valid, 
                        "reports": self._report.createReportObjects()}
        if not filename is None:
            with open(filename, "w") as f:
                json.dump(reportObject, f)
        return reportObject

    def createTextReport(self, filename:str=None, textWidth=100, examples=3, warnings=True):
        """
        Create text version report
        """
        #create text
        name = os.path.basename(self._name)
        reportText = "=== {} '{}' (version {}) ===\n".format(
            ("VALID" if self._report.valid else "INVALID"),name,__version__)
        reportText = reportText + self._report.createTextReport(textWidth, examples, False, warnings)
        if not filename is None:
            with open(filename, "w") as f:
                f.write(reportText)
        return reportText

    def createMarkdownReport(self, filename:str=None, textWidth=100, examples=3, warnings=True):
        """
        Create markdown version report
        """
        #create text
        name = os.path.basename(self._name)
        reportText = "# {}: {}\n\nVersion {}\n\n".format(
            ("VALID" if self._report.valid else "INVALID"),name,__version__)
        reportText = reportText + self._report.createTextReport(textWidth, examples, True, warnings)
        if not filename is None:
            with open(filename, "w") as f:
                f.write(reportText)
        return reportText

    @property
    def valid(self):
        """
        Valid
        """
        return self._report.valid

    @property
    def reports(self):
        """
        Reports
        """
        return self._report.reports

    def __getitem__(self, key):
        return self._report[key]


    def _updateNestedDictionary(data, locationPath, value):
        if len(locationPath)==1:
            data[locationPath[0]] = value
        elif len(locationPath)>1:
             Validate._updateNestedDictionary(data[locationPath[0]], locationPath[1:], value)
