"""Excel spreadsheet reading and extraction functions for AI agents.

This module provides functions for reading and extracting data from
Excel (.xlsx) spreadsheets.
"""

import os

from ..decorators import strands_tool

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped, import-not-found]

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Maximum file size: 100MB
MAX_FILE_SIZE = 100 * 1024 * 1024


@strands_tool
def read_excel_sheet(file_path: str, sheet_name: str) -> list[list[str]]:
    """Read Excel sheet as 2D list of strings.

    This function reads all data from a specific sheet and returns it
    as a 2D list where each inner list represents a row.

    Args:
        file_path: Path to Excel file to read
        sheet_name: Name of sheet to read

    Returns:
        2D list of strings representing sheet data

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, or sheet doesn't exist

    Example:
        >>> data = read_excel_sheet("/path/to/file.xlsx", "Sheet1")
        >>> data[0]  # First row
        ['Name', 'Age', 'City']
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        # Load workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        # Get sheet
        ws = wb[sheet_name]

        # Extract all data
        data = []
        for row in ws.iter_rows(values_only=True):
            # Convert row to strings, handling None values
            str_row = [str(cell) if cell is not None else "" for cell in row]
            data.append(str_row)

        wb.close()
        return data

    except Exception as e:
        raise ValueError(f"Failed to read Excel sheet: {e}")


@strands_tool
def get_excel_sheet_names(file_path: str) -> list[str]:
    """Get list of all sheet names in Excel workbook.

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> sheets = get_excel_sheet_names("/path/to/file.xlsx")
        >>> sheets
        ['Sheet1', 'Sheet2', 'Data']
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheet_names = list(wb.sheetnames)  # Cast to list[str] for type safety
        wb.close()
        return sheet_names

    except Exception as e:
        raise ValueError(f"Failed to get sheet names: {e}")


@strands_tool
def read_excel_as_dicts(
    file_path: str, sheet_name: str, header_row: int
) -> list[dict[str, str]]:
    """Read Excel sheet as list of dictionaries using header row as keys.

    This function reads a sheet and returns each row as a dictionary where
    keys come from the specified header row.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read
        header_row: Row number containing headers (1-indexed)

    Returns:
        List of dictionaries, one per data row

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If header_row invalid, file unreadable, or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> data = read_excel_as_dicts("/path/to/file.xlsx", "Sheet1", 1)
        >>> data[0]
        {'Name': 'Alice', 'Age': '30', 'City': 'NYC'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(header_row, int):
        raise TypeError("header_row must be an integer")

    if header_row < 1:
        raise ValueError("header_row must be >= 1 (1-indexed)")

    # Read sheet data
    data = read_excel_sheet(file_path, sheet_name)

    if len(data) < header_row:
        raise ValueError(
            f"header_row {header_row} out of range (sheet has {len(data)} rows)"
        )

    # Extract headers (convert to 0-indexed)
    headers = data[header_row - 1]

    # Convert rows to dicts, skipping header row
    result = []
    for i, row in enumerate(data):
        if i < header_row:
            continue  # Skip rows before and including header

        # Create dict with headers as keys
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                key = headers[j] if headers[j] else f"Column{j + 1}"
                row_dict[key] = cell
            else:
                # Handle extra cells beyond headers
                row_dict[f"Column{j + 1}"] = cell

        result.append(row_dict)

    return result


@strands_tool
def get_excel_cell_value(file_path: str, sheet_name: str, cell_reference: str) -> str:
    """Get value from single cell using A1 notation.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_reference: Cell reference in A1 notation (e.g., "B5", "AA10")

    Returns:
        Cell value as string

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If cell_reference invalid or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> value = get_excel_cell_value("/path/to/file.xlsx", "Sheet1", "B5")
        >>> value
        '42'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_reference, str):
        raise TypeError("cell_reference must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        cell_value = ws[cell_reference].value
        wb.close()

        return str(cell_value) if cell_value is not None else ""

    except KeyError:
        raise ValueError(f"Invalid cell reference: {cell_reference}")
    except Exception as e:
        raise ValueError(f"Failed to get cell value: {e}")


@strands_tool
def get_excel_cell_range(
    file_path: str, sheet_name: str, start_cell: str, end_cell: str
) -> list[list[str]]:
    """Get range of cells as 2D list using A1 notation.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        start_cell: Start cell reference (e.g., "A1")
        end_cell: End cell reference (e.g., "C10")

    Returns:
        2D list of strings for the specified range

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If cell references invalid or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> data = get_excel_cell_range("/path/to/file.xlsx", "Sheet1", "A1", "C3")
        >>> data
        [['Name', 'Age', 'City'], ['Alice', '30', 'NYC'], ['Bob', '25', 'LA']]
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(start_cell, str):
        raise TypeError("start_cell must be a string")

    if not isinstance(end_cell, str):
        raise TypeError("end_cell must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get range
        cell_range = f"{start_cell}:{end_cell}"
        data = []

        for row in ws[cell_range]:
            str_row = [
                str(cell.value) if cell.value is not None else "" for cell in row
            ]
            data.append(str_row)

        wb.close()
        return data

    except Exception as e:
        raise ValueError(f"Failed to get cell range: {e}")


@strands_tool
def search_excel_text(
    file_path: str, search_term: str, case_sensitive: bool
) -> list[dict[str, object]]:
    """Search for text across all sheets in Excel workbook.

    Args:
        file_path: Path to Excel file
        search_term: Text to search for
        case_sensitive: Whether search should be case-sensitive

    Returns:
        List of dicts with keys: sheet_name, row, column, cell_reference, value

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> matches = search_excel_text("/path/to/file.xlsx", "Python", False)
        >>> matches[0]
        {'sheet_name': 'Sheet1', 'row': 5, 'column': 2, 'cell_reference': 'B5', 'value': 'Python Developer'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(search_term, str):
        raise TypeError("search_term must be a string")

    if not isinstance(case_sensitive, bool):
        raise TypeError("case_sensitive must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        matches = []

        # Prepare search term for comparison
        search_compare = search_term if case_sensitive else search_term.lower()

        # Search all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue

                    cell_str = str(cell.value)
                    cell_compare = cell_str if case_sensitive else cell_str.lower()

                    if search_compare in cell_compare:
                        matches.append(
                            {
                                "sheet_name": sheet_name,
                                "row": row_idx,
                                "column": col_idx,
                                "cell_reference": cell.coordinate,
                                "value": cell_str,
                            }
                        )

        wb.close()
        return matches

    except Exception as e:
        raise ValueError(f"Failed to search Excel file: {e}")


@strands_tool
def get_excel_metadata(file_path: str) -> dict[str, str]:
    """Get Excel workbook metadata and properties.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary containing metadata (creator, title, subject, etc.)

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> metadata = get_excel_metadata("/path/to/file.xlsx")
        >>> metadata['creator']
        'John Doe'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        props = wb.properties

        metadata = {
            "creator": str(props.creator) if props.creator else "",
            "title": str(props.title) if props.title else "",
            "subject": str(props.subject) if props.subject else "",
            "description": str(props.description) if props.description else "",
            "keywords": str(props.keywords) if props.keywords else "",
            "category": str(props.category) if props.category else "",
            "last_modified_by": str(props.lastModifiedBy)
            if props.lastModifiedBy
            else "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
        }

        wb.close()
        return metadata

    except Exception as e:
        raise ValueError(f"Failed to get Excel metadata: {e}")


@strands_tool
def get_excel_info(file_path: str) -> dict[str, object]:
    """Get comprehensive information about Excel workbook.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with file info, sheets, dimensions, formula count, etc.

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> info = get_excel_info("/path/to/file.xlsx")
        >>> info['sheet_count']
        3
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)

        # Get sheet information
        sheet_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info.append(
                {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                }
            )

        info = {
            "file_path": file_path,
            "file_size_bytes": file_size,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheet_info,
        }

        wb.close()
        return info

    except Exception as e:
        raise ValueError(f"Failed to get Excel info: {e}")
